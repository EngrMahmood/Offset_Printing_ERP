from __future__ import annotations

import csv
import io
import re
from typing import Any


def _json_primitive(value: Any) -> str:
    if value is None:
        return ''
    return str(value)


# Same brand blue used across the web app's headers/buttons
# (core/static/core/css/erp-global-theme.css --erp-primary), so a printed
# report is visually recognizable as coming from this system rather than a
# plain data dump.
BRAND_PRIMARY = '#2f6ea7'
BRAND_PRIMARY_STRONG = '#255986'
BRAND_PRIMARY_SOFT = '#eaf1f8'

# Only applied to columns that are actually a status/issue field (matched by
# header key below) - keeps this from accidentally recoloring unrelated text
# that happens to contain one of these words (e.g. a SKU code).
_STATUS_COLOR_KEYWORDS = (
    ('RED', '#dc2626'),
    ('STUCK', '#dc2626'),
    ('AWAITING', '#d97706'),
    ('YELLOW', '#d97706'),
    ('PENDING', '#d97706'),
    ('NOT RELEASED', '#64748b'),
    ('GREEN', '#16a34a'),
)
_STATUS_COLOR_HEADERS = {'issue', 'status'}


def _status_color(header: str, value: str) -> str | None:
    if header not in _STATUS_COLOR_HEADERS or not value:
        return None
    upper = value.upper()
    for keyword, hex_color in _STATUS_COLOR_KEYWORDS:
        if keyword in upper:
            return hex_color
    return None


# Human-readable labels for the filter values that narrow a report down to a
# specific slice (which KPI, which detail mode, which pending-work stage) -
# these get folded into the export title/filename so two exports of the same
# report never look identical or overwrite each other's download.
_KPI_LABELS = {
    'order_fulfillment': 'Order Fulfillment Efficiency',
    'wastage_reduction': 'Wastage Reduction Efficiency',
    'dispatch_alignment': 'Dispatch vs Production Alignment',
    'all': 'All KPIs',
}
_DETAIL_LABELS = {
    'quarterly': 'Quarterly Detail',
    'focus': 'Improvement Focus',
}
_STAGE_LABELS = {
    'printing': 'Pending Printing',
    'packing': 'Pending Packing',
    'dispatch': 'Pending Dispatch',
    'not_released': 'Not Yet Released',
}


def _report_context_parts(payload: dict) -> list[str]:
    filters = payload.get('filters') or {}
    parts = []
    machine_filter = (filters.get('machine') or '').strip()
    if machine_filter:
        parts.append(machine_filter)
    kpi = (filters.get('kpi') or '').strip()
    if kpi:
        parts.append(_KPI_LABELS.get(kpi, kpi.replace('_', ' ').title()))
    detail = (filters.get('detail') or '').strip()
    if detail:
        parts.append(_DETAIL_LABELS.get(detail, detail.replace('_', ' ').title()))
    stage = (filters.get('stage') or '').strip()
    if stage:
        parts.append(_STAGE_LABELS.get(stage, stage.replace('_', ' ').title()))
    return parts


def build_report_title(payload: dict) -> str:
    report_title = (payload.get('report') or {}).get('title') or 'Report'
    parts = _report_context_parts(payload)
    if parts:
        report_title = f"{report_title} - {' - '.join(parts)}"
    return report_title


def _slugify(text: str) -> str:
    slug = re.sub(r'[^A-Za-z0-9]+', '-', text).strip('-').lower()
    return slug or 'report'


def build_export_filename(payload: dict, slug: str) -> str:
    base = slug.replace('_', '-').strip() or 'report'
    parts = _report_context_parts(payload)
    if parts:
        base = '-'.join([base, *[_slugify(p) for p in parts]])
    return base


def _headers_with_row_numbers(headers: list[str], labels: dict[str, str]) -> tuple[list[str], dict[str, str]]:
    """Prepend a row-number column, unless the report already has its own
    sequence column (e.g. Machine Planning's S#)."""
    if 'sequence' in headers or '_row_no' in headers:
        return headers, labels
    labels = {**labels, '_row_no': '#'}
    return ['_row_no', *headers], labels


def _cell_value(header: str, row_index: int, row: dict[str, Any]) -> str:
    if header == '_row_no':
        return str(row_index)
    return _json_primitive(row.get(header))


def _report_totals_line(payload: dict, rows: list[dict[str, Any]]) -> str | None:
    """Total Impressions / Hours summary line, Machine Planning only - other
    reports (e.g. Manual Working) have their own unrelated 'total_impressions'
    field, so this is gated on the report slug, not just key presence."""
    slug = (payload.get('report') or {}).get('slug')
    if slug != 'machine-planning' or not rows:
        return None
    total_impressions = sum(_to_number(row.get('total_impressions')) for row in rows)
    total_hours = sum(_to_number(row.get('estimated_hours')) for row in rows)
    return f'Total Impressions: {total_impressions:,.0f} — Total Hours: {total_hours:,.2f}h'


def _to_number(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _rowset_from_payload(payload: dict) -> list[dict[str, Any]]:
    data = payload.get('data') or {}
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if not isinstance(data, dict):
        return []

    # Prefer common tabular keys from report contexts. 'export_rows' lets a
    # report nominate exactly which of its tables should be exported.
    for key in (
        'export_rows',
        'wastage_rows',
        'machine_rows',
        'actual_rows',
        'status_rows',
        'by_machine',
        'by_shift',
        'by_date',
        'by_dc',
        'cutting_request_rows',
        'recent_dispatches',
        'pending_qc_jobs',
        'approved_recent',
        'rejected_jobs',
        'overdue_jobs',
        'due_soon_jobs',
        'missing_readiness',
        'missing_plate_jobs',
        'ready_jobs',
    ):
        value = data.get(key)
        if isinstance(value, list) and value and isinstance(value[0], dict):
            return value

    # Fallback to summary if no rowset is available.
    summary = data.get('summary')
    if isinstance(summary, dict):
        return [summary]
    return []


def export_as_csv(payload: dict) -> bytes:
    rows = _rowset_from_payload(payload)
    report_title = build_report_title(payload)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([report_title])
    writer.writerow(['generated_at', payload.get('generated_at', '')])
    totals_line = _report_totals_line(payload, rows)
    if totals_line:
        writer.writerow([totals_line])
    writer.writerow([])

    if not rows:
        writer.writerow(['message', 'No tabular rows available for export'])
        return output.getvalue().encode('utf-8-sig')

    headers = payload.get('headers') or (payload.get('data') or {}).get('headers')
    if not headers:
        headers = sorted({key for row in rows for key in row.keys()})
    labels = payload.get('header_labels') or (payload.get('data') or {}).get('header_labels') or {}
    headers, labels = _headers_with_row_numbers(headers, labels)
    writer.writerow([labels.get(header, header) for header in headers])
    for idx, row in enumerate(rows, start=1):
        writer.writerow([_cell_value(header, idx, row) for header in headers])
    return output.getvalue().encode('utf-8-sig')


def export_as_xlsx(payload: dict) -> bytes:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    except Exception as exc:
        raise RuntimeError('XLSX export requires openpyxl package.') from exc

    rows = _rowset_from_payload(payload)
    report_title = build_report_title(payload)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Report'
    sheet.append([report_title])
    sheet['A1'].font = Font(bold=True, size=13, color=BRAND_PRIMARY_STRONG.lstrip('#'))
    sheet.append(['generated_at', payload.get('generated_at', '')])
    totals_line = _report_totals_line(payload, rows)
    if totals_line:
        sheet.append([totals_line])
    sheet.append([])

    if not rows:
        sheet.append(['message', 'No tabular rows available for export'])
    else:
        headers = payload.get('headers') or (payload.get('data') or {}).get('headers')
        if not headers:
            headers = sorted({key for row in rows for key in row.keys()})
        labels = payload.get('header_labels') or (payload.get('data') or {}).get('header_labels') or {}
        headers, labels = _headers_with_row_numbers(headers, labels)
        sheet.append([labels.get(header, header) for header in headers])
        header_row_index = sheet.max_row
        header_fill = PatternFill('solid', fgColor=BRAND_PRIMARY.lstrip('#'))
        header_font = Font(bold=True, color='FFFFFF')
        for col_index in range(1, len(headers) + 1):
            cell = sheet.cell(row=header_row_index, column=col_index)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical='center')
        for idx, row in enumerate(rows, start=1):
            sheet.append([_cell_value(header, idx, row) for header in headers])
        sheet.freeze_panes = sheet.cell(row=header_row_index + 1, column=1).coordinate

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def export_as_pdf(payload: dict) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from xml.sax.saxutils import escape as xml_escape
    except Exception as exc:
        raise RuntimeError('PDF export requires reportlab package.') from exc

    rows = _rowset_from_payload(payload)
    report_title = build_report_title(payload)
    generated_at = payload.get('generated_at', '')
    row_count = len(rows)

    def draw_page_chrome(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(colors.HexColor(BRAND_PRIMARY))
        canvas.rect(0, doc.pagesize[1] - 5 * mm, doc.pagesize[0], 5 * mm, stroke=0, fill=1)
        canvas.setStrokeColor(colors.HexColor(BRAND_PRIMARY))
        canvas.setLineWidth(0.75)
        canvas.line(12 * mm, 13 * mm, doc.pagesize[0] - 12 * mm, 13 * mm)
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor('#64748b'))
        canvas.drawString(12 * mm, 8 * mm, 'Offset Printing ERP')
        canvas.drawRightString(doc.pagesize[0] - 12 * mm, 8 * mm, f'Page {canvas.getPageNumber()}')
        canvas.restoreState()

    buffer = io.BytesIO()
    page_w, page_h = landscape(A4)
    left_margin = right_margin = top_margin = 12 * mm
    bottom_margin = 16 * mm
    usable_width = page_w - left_margin - right_margin

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
    )

    title_style = ParagraphStyle(
        'reportTitle', fontName='Helvetica-Bold', fontSize=14, leading=17,
        textColor=colors.HexColor(BRAND_PRIMARY_STRONG),
    )
    meta_style = ParagraphStyle(
        'reportMeta', fontName='Helvetica', fontSize=8, leading=11, textColor=colors.HexColor('#64748b'),
    )

    if not rows:
        story = [
            Paragraph(f'Offset Printing ERP - {xml_escape(report_title)}', title_style),
            Paragraph(f'Generated at: {generated_at} — Rows: {row_count}', meta_style),
            Paragraph('No tabular rows available for export', meta_style),
        ]
        doc.build(story, onFirstPage=draw_page_chrome, onLaterPages=draw_page_chrome)
        return buffer.getvalue()

    headers = payload.get('headers') or (payload.get('data') or {}).get('headers')
    if not headers:
        headers = sorted({key for row in rows for key in row.keys()})
    labels = payload.get('header_labels') or (payload.get('data') or {}).get('header_labels') or {}
    headers, labels = _headers_with_row_numbers(headers, labels)

    header_style = ParagraphStyle(
        'header',
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=8,
        alignment=0,
        textColor=colors.white,
    )
    cell_style = ParagraphStyle(
        'cell',
        fontName='Helvetica',
        fontSize=6,
        leading=7,
        alignment=0,
    )

    def cell_paragraph(header, idx, row):
        value = _cell_value(header, idx, row)
        color = _status_color(header, value)
        text = xml_escape(value)
        if color:
            text = f'<font color="{color}"><b>{text}</b></font>'
        return Paragraph(text, cell_style)

    # Small enough that Machine Planning's 19 columns (with the narrow S#,
    # colour/pass/hour columns) all fit on one landscape A4 page instead of
    # spilling extra columns onto a second page - reduces paper usage.
    min_col_width = 9 * mm
    max_columns = max(1, min(len(headers), int(usable_width // min_col_width)))
    header_chunks = [headers] if len(headers) <= max_columns else [headers[i:i + max_columns] for i in range(0, len(headers), max_columns)]
    row_count_for_widths = min(row_count, 2000)  # cap the content-width scan; a sample is enough
    story = [
        Paragraph(f'Offset Printing ERP - {xml_escape(report_title)}', title_style),
        Paragraph(f'Generated at: {generated_at} — Rows: {row_count}', meta_style),
    ]
    totals_line = _report_totals_line(payload, rows)
    if totals_line:
        story.append(Paragraph(totals_line, meta_style))
    story.append(Paragraph('<br/>', meta_style))

    for chunk_index, chunk_headers in enumerate(header_chunks):
        table_data = [
            [Paragraph(xml_escape(str(labels.get(header, header))), header_style) for header in chunk_headers]
        ]
        for idx, row in enumerate(rows, start=1):
            table_data.append([cell_paragraph(header, idx, row) for header in chunk_headers])

        # Column widths follow actual content length (header label vs. widest
        # cell seen, sampled) instead of a hand-maintained per-report map, so
        # every report - including ones added later - gets sensible widths:
        # short numeric/status columns stay narrow, long free-text columns
        # (SKU, remarks) get the room they need without leaving other columns
        # padded with empty space. Weights are floored/capped so one very
        # long value can't starve the rest of the row, then normalized to
        # exactly fill the page width.
        MIN_CHARS, MAX_CHARS = 4, 38
        weights = []
        for header in chunk_headers:
            max_len = len(str(labels.get(header, header)))
            for row in rows[:row_count_for_widths]:
                cell = _cell_value(header, 1, row)
                if cell and len(cell) > max_len:
                    max_len = len(cell)
            weights.append(min(max(max_len, MIN_CHARS), MAX_CHARS))
        total_weight = sum(weights) or len(chunk_headers)
        col_widths = [w / total_weight * usable_width for w in weights]

        table_style_commands = [
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(BRAND_PRIMARY)),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]
        # Zebra striping on body rows makes a wide table scannable at a
        # glance instead of reading as an undifferentiated block of text.
        for row_index in range(2, len(table_data), 2):
            table_style_commands.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.HexColor(BRAND_PRIMARY_SOFT)))

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(table_style_commands))
        story.append(table)
        if chunk_index < len(header_chunks) - 1:
            story.append(PageBreak())

    doc.build(story, onFirstPage=draw_page_chrome, onLaterPages=draw_page_chrome)
    return buffer.getvalue()
