from django.shortcuts import render, get_object_or_404, redirect
from django.conf import settings
from django.http import HttpResponse, JsonResponse, Http404
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from django.db.models import Prefetch, Q, Sum, Min, Max, F
from django.db import transaction
from django.db.models.deletion import ProtectedError
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.urls import reverse
from functools import wraps
import csv
import json
import re
from datetime import datetime, date, timedelta

from .bulk_upload import process_jobcard_upload, get_template_headers, get_template_example
from .jc_numbering import allocate_next_jc_number
from .machine_routing import MM_PER_INCH
from .models import (
    ApplicationType,
    ChangeLog,
    DeliveryLocation,
    Department,
    Dispatch,
    EditOverrideRequest,
    JOB_CARD_DISPATCHABLE_STATUSES,
    JOB_CARD_PRODUCTION_START_STATUSES,
    JobCard,
    Machine,
    MachineWorkSchedule,
    Material,
    Notification,
    Operator,
    PrintColor,
    ProductType,
    Production,
    ProductionDowntime,
    ShiftConfig,
    Supervisor,
    UserProfile,
    Vendor,
)
from .jobcard_service import normalize_job_card_status
from workflow.services import start_production

from .constants import AUDIT_CONFIG
from .services import (
    add_unique_message,
    format_audit_value, normalize_colour_notation, extract_total_colors,
    compute_planned_minutes, get_remaining_planned_minutes, build_audit_snapshot,
    sync_departments_from_planning,
    sync_delivery_locations_from_planning,
    sync_product_types_from_sku_recipes,
    sync_materials_from_planning,
    count_active_planning_jobs_for_product_type,
    build_change_summary, log_change, user_has_entity_permission,
    user_can_archive_records, user_can_bypass_edit_lock, get_record_edit_lock_days,
    get_record_edit_lock_cutoff, record_is_time_locked, get_valid_override,
    ensure_edit_lock_allowed, get_accessible_entities, get_active_record_or_404,
    validate_delete_allowed,
    validate_restore_allowed, archive_record, restore_record_state,
    run_bulk_archive, run_bulk_permanent_delete
)


def _parse_optional_decimal(raw_value):
    if raw_value is None:
        return None
    raw_text = str(raw_value).strip().replace(',', '')
    if not raw_text:
        return None
    try:
        return Decimal(raw_text)
    except InvalidOperation:
        return None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ═══════════════════════════════════
# PERMISSION DECORATORS (RBAC)
# ═══════════════════════════════════


def require_role(*allowed_roles):
    """Decorator to check if user has required role"""
    def decorator(view_func):
        @wraps(view_func)
        def wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            try:
                profile = request.user.profile
                user_role = (profile.role or '').strip().lower()
                if user_role not in [role.lower() for role in allowed_roles] and not request.user.is_staff:
                    add_unique_message(request, messages.ERROR, '❌ You do not have permission to access this page.')
                    return redirect('home')
            except UserProfile.DoesNotExist:
                add_unique_message(request, messages.ERROR, '⚠️ Your user profile is not configured. Contact admin.')
                return redirect('login')
            return view_func(request, *args, **kwargs)
        return wrapped_view
    return decorator


def permission_required(permission_method):
    """Decorator to check specific permission method on UserProfile"""
    def decorator(view_func):
        @wraps(view_func)
        def wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            try:
                profile = request.user.profile
                if not getattr(profile, permission_method)():
                    add_unique_message(request, messages.ERROR, '❌ You do not have permission to access this feature.')
                    return redirect('home')
            except (UserProfile.DoesNotExist, AttributeError):
                add_unique_message(request, messages.ERROR, '⚠️ Permission check failed. Contact admin.')
                return redirect('login')
            return view_func(request, *args, **kwargs)
        return wrapped_view
    return decorator


@login_required
def home(request):
    from django.db import models
    from django.utils import timezone
    from planning.models import PlanningJob, SkuRecipe
    from printing_plates.models import PlateRequest
    from core.models import Notification, Production, Dispatch, Machine, Operator
    from supply_chain.models import RawMaterialSku

    pending_qc_jobs = PlanningJob.objects.filter(status='pending_qc').count()
    pending_sku_reviews = SkuRecipe.objects.filter(master_data_status='pending_review').count()
    active_plate_requests = PlateRequest.objects.filter(status__in=PlateRequest.OPEN_STATUSES).count()
    unread_notifications = Notification.objects.filter(user=request.user, is_read=False).count()

    # Global Operations KPIs
    total_in_production = PlanningJob.objects.filter(status='in_production').count()
    total_released_jobs = PlanningJob.objects.filter(status='released').count()
    total_produced_sheets = Production.objects.filter(is_active=True).aggregate(total=models.Sum('output_sheets'))['total'] or 0
    total_dispatched_qty = Dispatch.objects.filter(is_active=True).aggregate(total=models.Sum('dispatch_qty'))['total'] or 0
    total_inventory_items = RawMaterialSku.objects.filter(is_active=True).count()

    # Shop Floor Operations KPIs
    today = timezone.now().date()
    active_machines = Machine.objects.filter(is_active=True).count()
    active_operators = Operator.objects.filter(is_active=True).count()
    today_sorting_waste = Production.objects.filter(date=today, is_active=True).aggregate(total=models.Sum('sorting_waste_qty'))['total'] or 0
    today_downtime = Production.objects.filter(date=today, is_active=True).aggregate(total=models.Sum('downtime_minutes'))['total'] or 0

    context = {
        'kpis': {
            'pending_qc_jobs': pending_qc_jobs,
            'pending_sku_reviews': pending_sku_reviews,
            'active_plate_requests': active_plate_requests,
            'unread_notifications': unread_notifications,
            'total_in_production': total_in_production,
            'total_released_jobs': total_released_jobs,
            'total_produced_sheets': total_produced_sheets,
            'total_dispatched_qty': total_dispatched_qty,
            'total_inventory_items': total_inventory_items,
            
            # Shop Floor
            'active_machines': active_machines,
            'active_operators': active_operators,
            'today_sorting_waste': today_sorting_waste,
            'today_downtime': today_downtime,
        }
    }
    return render(request, 'home.html', context)


def erp_version(request):
    return JsonResponse({
        'erp_software_version': getattr(settings, 'ERP_SOFTWARE_VERSION', '0.0.0'),
        'erp_software_release_date': getattr(settings, 'ERP_SOFTWARE_RELEASE_DATE', ''),
        'server_time': timezone.now().isoformat(),
    })


@login_required
@permission_required('can_edit_jobcard')
def bulk_upload_jobcards(request):
    context = {}

    if request.method == "POST":
        file = request.FILES.get('file')

        if not file:
            context = {
                'success_count': 0,
                'error_count': 1,
                'errors': [{'row': 0, 'errors': 'Please choose a file to upload.'}],
            }
            return render(request, "upload.html", context)

        result = process_jobcard_upload(file, uploaded_by=request.user)
        context = result

        return render(request, "upload.html", context)

    return render(request, "upload.html", context)


@login_required
@require_POST
def quick_add_master(request):
    """Create master dropdown values for planner workflow without admin dependency."""
    master_type = (request.POST.get('type') or '').strip().lower()
    name = (request.POST.get('name') or '').strip()

    if master_type not in {'material', 'machine', 'department', 'delivery_location', 'operator', 'vendor', 'product_type', 'print_color', 'application'}:
        return JsonResponse({'ok': False, 'error': 'Invalid master type.'}, status=400)

    if not name:
        return JsonResponse({'ok': False, 'error': 'Name is required.'}, status=400)

    if master_type in {'material', 'application'}:
        profile = getattr(request.user, 'profile', None)
        role = getattr(profile, 'normalized_role', '') if profile else ''
        if not (
            request.user.is_superuser
            or role in {'admin', 'planner'}
        ):
            return JsonResponse({'ok': False, 'error': f'Only planner or admin can add {master_type.replace("_", " ")}s.'}, status=403)

    if master_type == 'print_color':
        profile = getattr(request.user, 'profile', None)
        if not (request.user.is_superuser or (profile and profile.role == 'admin')):
            return JsonResponse({'ok': False, 'error': 'Only admin can add print colors.'}, status=403)
        existing = PrintColor.objects.filter(name__iexact=name).first()
        if existing:
            return JsonResponse({
                'ok': True,
                'created': False,
                'id': existing.id,
                'name': existing.name,
                'type': master_type,
                'message': 'Already exists. Selected existing value.',
            })
        obj = PrintColor.objects.create(name=name, is_active=True)
        return JsonResponse({
            'ok': True,
            'created': True,
            'id': obj.id,
            'name': obj.name,
            'type': master_type,
            'message': 'Created successfully.',
        })

    if master_type == 'operator':
        employee_code = (request.POST.get('employee_code') or '').strip() or None

        existing = None
        if employee_code:
            existing = Operator.objects.filter(employee_code__iexact=employee_code).first()
        if existing is None:
            existing = Operator.objects.filter(
                name__iexact=name,
                employee_code__iexact=employee_code or ''
            ).first() if employee_code else Operator.objects.filter(name__iexact=name, employee_code__isnull=True).first()

        if existing:
            display_name = f"{existing.name} ({existing.employee_code})" if existing.employee_code else existing.name
            return JsonResponse({
                'ok': True,
                'created': False,
                'id': existing.id,
                'name': existing.name,
                'display_name': display_name,
                'employee_code': existing.employee_code,
                'type': master_type,
                'message': 'Already exists. Selected existing value.'
            })

        obj = Operator.objects.create(name=name, employee_code=employee_code)
        display_name = f"{obj.name} ({obj.employee_code})" if obj.employee_code else obj.name
        return JsonResponse({
            'ok': True,
            'created': True,
            'id': obj.id,
            'name': obj.name,
            'display_name': display_name,
            'employee_code': obj.employee_code,
            'type': master_type,
            'message': 'Created successfully.'
        })

    if master_type == 'machine':
        standard_speed_raw = (request.POST.get('standard_impressions_per_hour') or '').strip()
        setup_per_color_raw = (request.POST.get('standard_setup_minutes_per_color') or '').strip()
        standard_speed = 4000
        setup_per_color = 15
        if standard_speed_raw:
            try:
                standard_speed = float(standard_speed_raw)
            except ValueError:
                return JsonResponse({'ok': False, 'error': 'Ideal speed must be a number.'}, status=400)
            if standard_speed <= 0:
                return JsonResponse({'ok': False, 'error': 'Ideal speed must be greater than 0.'}, status=400)

        if setup_per_color_raw:
            try:
                setup_per_color = float(setup_per_color_raw)
            except ValueError:
                return JsonResponse({'ok': False, 'error': 'Setup minutes per color must be a number.'}, status=400)
            if setup_per_color < 0:
                return JsonResponse({'ok': False, 'error': 'Setup minutes per color cannot be negative.'}, status=400)

        existing = Machine.objects.filter(name__iexact=name).first()
        if existing:
            return JsonResponse({
                'ok': True,
                'created': False,
                'id': existing.id,
                'name': existing.name,
                'standard_impressions_per_hour': existing.standard_impressions_per_hour,
                'standard_setup_minutes_per_color': existing.standard_setup_minutes_per_color,
                'type': master_type,
                'message': 'Already exists. Selected existing value.'
            })

        obj = Machine.objects.create(
            name=name,
            standard_impressions_per_hour=standard_speed,
            standard_setup_minutes_per_color=setup_per_color,
        )
        return JsonResponse({
            'ok': True,
            'created': True,
            'id': obj.id,
            'name': obj.name,
            'standard_impressions_per_hour': obj.standard_impressions_per_hour,
            'standard_setup_minutes_per_color': obj.standard_setup_minutes_per_color,
            'type': master_type,
            'message': 'Created successfully.'
        })

    model_map = {
        'material': Material,
        'department': Department,
        'delivery_location': DeliveryLocation,
        'vendor': Vendor,
        'product_type': ProductType,
        'application': ApplicationType,
    }
    model = model_map[master_type]

    existing = model.objects.filter(name__iexact=name).first()
    if existing:
        return JsonResponse({
            'ok': True,
            'created': False,
            'id': existing.id,
            'name': existing.name,
            'type': master_type,
            'message': 'Already exists. Selected existing value.'
        })

    obj = model.objects.create(name=name)
    return JsonResponse({
        'ok': True,
        'created': True,
        'id': obj.id,
        'name': obj.name,
        'type': master_type,
        'message': 'Created successfully.'
    })


@login_required
@permission_required('can_edit_jobcard')
def job_card_entry(request):
    """Redirect legacy production job card entry to planning."""
    return redirect('planning:job_cards')


@login_required
@permission_required('can_edit_jobcard')
def job_card_records(request):
    """Legacy job card records redirect."""
    return redirect('planning:job_cards')


def _dispatchable_job_cards_queryset(edit_record=None):
    qs = JobCard.objects.filter(is_active=True, status__in=JOB_CARD_DISPATCHABLE_STATUSES)
    if edit_record:
        qs = JobCard.objects.filter(is_active=True).filter(
            Q(status__in=JOB_CARD_DISPATCHABLE_STATUSES) | Q(pk=edit_record.job_card_id)
        ).distinct()
    return qs.select_related('planning_job').prefetch_related(
        Prefetch(
            'dispatch_set',
            queryset=Dispatch.objects.filter(is_active=True).select_related('created_by').order_by('-dispatch_date', '-id'),
        ),
        Prefetch(
            'productions',
            queryset=Production.objects.filter(is_active=True),
        ),
    )


def _build_dispatch_job_card_info(job_card, edit_record_id=None):
    dispatch_rows = list(job_card.dispatch_set.all())
    history_data = []
    for dispatch_row in dispatch_rows:
        history_data.append({
            'id': dispatch_row.id,
            'date': dispatch_row.dispatch_date.strftime('%Y-%m-%d') if dispatch_row.dispatch_date else '',
            'dc_no': dispatch_row.dc_no or '-',
            'qty': dispatch_row.dispatch_qty,
            'qty_display': f'{dispatch_row.dispatch_qty:,}',
            'added_by': dispatch_row.created_by.username if dispatch_row.created_by else '-',
            'is_current': dispatch_row.id == edit_record_id,
        })

    dispatched_total = sum(row.dispatch_qty for row in dispatch_rows)
    if edit_record_id and any(row.id == edit_record_id for row in dispatch_rows):
        edit_qty = next(row.dispatch_qty for row in dispatch_rows if row.id == edit_record_id)
        dispatched_before = dispatched_total - (edit_qty or 0)
    else:
        dispatched_before = dispatched_total

    remaining_qty = max(0, job_card.order_qty - dispatched_total)
    if edit_record_id and any(row.id == edit_record_id for row in dispatch_rows):
        remaining_qty = max(0, job_card.order_qty - dispatched_before)

    dispatch_pct = round((dispatched_total / job_card.order_qty) * 100, 2) if job_card.order_qty > 0 else 0

    return {
        'job_card_no': job_card.job_card_no,
        'sku': job_card.SKU or '-',
        'customer': job_card.destination or '-',
        'po_no': job_card.PO_No or '-',
        'order_qty': job_card.order_qty,
        'order_qty_display': f'{job_card.order_qty:,}',
        'produced_qty': int(job_card.total_packed_pcs or 0) if job_card.is_print_job else int(job_card.total_packed_pcs or 0),
        'produced_qty_display': f'{int(job_card.total_packed_pcs or 0):,}',
        'printed_pcs_display': f'{int(job_card.total_printed_pcs or 0):,}' if job_card.is_print_job else 'N/A',
        'dispatched_before': dispatched_before,
        'dispatched_before_display': f'{dispatched_before:,}',
        'dispatched_total': dispatched_total,
        'dispatched_total_display': f'{dispatched_total:,}',
        'remaining': remaining_qty,
        'remaining_display': f'{remaining_qty:,}',
        'dispatch_pct': dispatch_pct,
        'is_print_job': job_card.is_print_job,
        'history': history_data,
        'history_count': len(history_data),
    }


def _dispatch_remaining_badge(order_qty, total_dispatch):
    remaining = order_qty - total_dispatch
    if remaining <= 0:
        return {
            'label': 'Complete',
            'badge_class': 'erp-badge-completed',
            'remaining': 0,
            'remaining_display': '0',
        }
    if total_dispatch <= 0:
        return {
            'label': 'Not dispatched',
            'badge_class': 'erp-badge-draft',
            'remaining': remaining,
            'remaining_display': f'{remaining:,}',
        }
    return {
        'label': f'{remaining:,} left',
        'badge_class': 'erp-badge-pending',
        'remaining': remaining,
        'remaining_display': f'{remaining:,}',
    }


@login_required
@permission_required('can_approve_dispatch')
def dispatch_entry(request):
    """Dispatch entry form"""
    view_id = (request.GET.get('view') or '').strip()
    is_view_mode = bool(view_id)
    edit_id = '' if is_view_mode else (request.POST.get('edit_id') or request.GET.get('edit') or '').strip()
    edit_record = get_active_record_or_404(Dispatch, view_id) if is_view_mode else (get_active_record_or_404(Dispatch, edit_id) if edit_id else None)
    if edit_record and not is_view_mode and not ensure_edit_lock_allowed(request, 'dispatch', edit_record):
        return redirect('dispatch_records')

    if request.method == 'POST' and not is_view_mode:
        try:
            change_reason = (request.POST.get('change_reason') or '').strip()
            job_card_id = request.POST.get('job_card')
            dc_no = (request.POST.get('dc_no') or '').strip()
            dispatch_date_raw = request.POST.get('dispatch_date')
            dispatch_qty = int(request.POST.get('dispatch_qty') or 0)

            if edit_record and not change_reason:
                raise ValueError('Change reason is required when editing dispatch')
            if not job_card_id:
                raise ValueError('Job card is required')
            if not dc_no:
                raise ValueError('DC / DR number is required')
            if dispatch_qty <= 0:
                raise ValueError('Dispatch quantity must be greater than 0')

            job_card = get_active_record_or_404(JobCard, job_card_id)
            if not (job_card.PO_No or '').strip():
                raise ValueError('PO number is required before dispatch. Link PO to this job first.')
            dispatch_date = datetime.strptime(dispatch_date_raw, "%Y-%m-%d").date() if dispatch_date_raw else timezone.now().date()

            payload = {
                'job_card': job_card,
                'dc_no': dc_no,
                'dispatch_date': dispatch_date,
                'dispatch_qty': dispatch_qty,
            }

            if edit_record:
                before_snapshot = build_audit_snapshot('dispatch', edit_record)
                for field_name, value in payload.items():
                    setattr(edit_record, field_name, value)
                edit_record.save()

                if log_change('dispatch', edit_record, before_snapshot, request.user, 'update', change_reason):
                    messages.success(request, f'Dispatch updated for {job_card.job_card_no}')
                else:
                    messages.success(request, f'No changes detected for {job_card.job_card_no}')
                return redirect('dispatch_records')

            record = Dispatch.objects.create(**payload)
            record.created_by = request.user
            record.save(update_fields=['created_by'])
            log_change('dispatch', record, {}, request.user, 'create', 'Initial entry created')

            messages.success(request, f'Dispatch saved for {job_card.job_card_no}')
            return redirect('dispatch_entry')
        except Exception as e:
            messages.error(request, f'Error saving dispatch: {str(e)}')

    dispatch_jobs = _dispatchable_job_cards_queryset(edit_record).order_by('-created_at')[:200]

    job_card_info_map = {}
    edit_record_id = edit_record.id if edit_record else None
    for job_card in dispatch_jobs:
        job_card_info_map[str(job_card.id)] = _build_dispatch_job_card_info(job_card, edit_record_id)

    prefill_job_card_id = (request.GET.get('job_card') or '').strip()
    if edit_record:
        prefill_job_card_id = str(edit_record.job_card_id)
    prefill_dc_no = (request.GET.get('dc_no') or '').strip()
    if edit_record and not prefill_dc_no:
        prefill_dc_no = edit_record.dc_no or ''

    context = {
        'job_cards': dispatch_jobs,
        'job_card_info_json': json.dumps(job_card_info_map),
        'today': edit_record.dispatch_date if edit_record else timezone.now().date(),
        'edit_record': edit_record,
        'edit_lock_days': get_record_edit_lock_days(),
        'edit_lock_applies': bool(edit_record and record_is_time_locked('dispatch', edit_record)),
        'is_view_mode': is_view_mode,
        'prefill_dc_no': prefill_dc_no,
        'prefill_job_card_id': prefill_job_card_id,
    }
    return render(request, 'dispatch_entry.html', context)


@login_required
@permission_required('can_approve_dispatch')
@require_GET
def dispatch_job_card_search(request):
    """Server-side job card search for dispatch entry (beyond initial 200 preload)."""
    query = (request.GET.get('q') or '').strip()
    edit_id = (request.GET.get('edit_id') or '').strip()
    edit_record = get_active_record_or_404(Dispatch, edit_id) if edit_id else None

    if len(query) < 2:
        return JsonResponse({'results': []})

    qs = _dispatchable_job_cards_queryset(edit_record).filter(
        Q(job_card_no__icontains=query) |
        Q(SKU__icontains=query) |
        Q(PO_No__icontains=query) |
        Q(destination__icontains=query)
    ).order_by('-created_at')[:60]

    edit_record_id = edit_record.id if edit_record else None
    results = []
    for job_card in qs:
        info = _build_dispatch_job_card_info(job_card, edit_record_id)
        results.append({
            'id': job_card.id,
            'label': f'{job_card.job_card_no} - {job_card.SKU}',
            'job_card_no': job_card.job_card_no,
            'sku': info['sku'],
            'customer': info['customer'],
            'remaining': info['remaining'],
            'remaining_display': info['remaining_display'],
            'info': info,
        })

    completed_matches = []
    if not results:
        from core.services import find_completed_job_card_matches
        completed_matches = find_completed_job_card_matches(query)

    return JsonResponse({'results': results, 'completed_matches': completed_matches})


@login_required
@permission_required('can_approve_dispatch')
@require_GET
def dispatch_dc_duplicate_check(request):
    """Check DC usage: block same DC on same JC; list other SKUs/lines on shared DC."""
    dc_no = (request.GET.get('dc_no') or '').strip()
    job_card_id = (request.GET.get('job_card_id') or '').strip()
    exclude_id = (request.GET.get('exclude_id') or '').strip()

    if not dc_no:
        return JsonResponse({
            'same_jc_duplicate': False,
            'same_dc_entries': [],
            'same_dc_sku_count': 0,
            'same_dc_line_count': 0,
            'blocking': False,
        })

    matches_qs = Dispatch.objects.filter(is_active=True, dc_no__iexact=dc_no).select_related('job_card')
    if exclude_id:
        matches_qs = matches_qs.exclude(pk=exclude_id)

    same_jc_duplicate = False
    same_dc_entries = []

    for match in matches_qs.order_by('job_card__job_card_no', '-dispatch_date', '-id'):
        if job_card_id and str(match.job_card_id) == job_card_id:
            same_jc_duplicate = True
            continue
        same_dc_entries.append({
            'id': match.id,
            'job_card_no': match.job_card.job_card_no,
            'sku': match.job_card.SKU or '-',
            'customer': match.job_card.destination or '-',
            'dispatch_date': match.dispatch_date.strftime('%d %b %Y') if match.dispatch_date else '-',
            'dispatch_qty': match.dispatch_qty,
        })

    sku_values = {entry['sku'] for entry in same_dc_entries if entry['sku'] and entry['sku'] != '-'}

    return JsonResponse({
        'same_jc_duplicate': same_jc_duplicate,
        'same_dc_entries': same_dc_entries,
        'same_dc_sku_count': len(sku_values),
        'same_dc_line_count': len(same_dc_entries),
        'blocking': same_jc_duplicate,
    })


@login_required
@permission_required('can_view_dispatch_records')
def dispatch_records(request):
    """Dispatch records list page"""
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        if action == 'bulk_delete':
            if not request.user.is_superuser:
                add_unique_message(request, messages.ERROR, '❌ Only a superuser can run bulk delete.')
                return redirect('dispatch_records')

            selected_ids = request.POST.getlist('selected_ids')
            deleted_count, failures = run_bulk_permanent_delete(request, 'dispatch', selected_ids)
            if deleted_count:
                add_unique_message(request, messages.SUCCESS, f'Deleted {deleted_count} dispatch record(s) permanently.')
            if failures:
                add_unique_message(request, messages.ERROR, f'Bulk delete completed with issues: {"; ".join(failures[:5])}')
            return redirect('dispatch_records')

    query = (request.GET.get('q') or '').strip()
    dc_filter = (request.GET.get('dc_no') or '').strip()
    balance_status = (request.GET.get('balance_status') or '').strip().lower()
    date_from_raw = (request.GET.get('date_from') or '').strip()
    date_to_raw = (request.GET.get('date_to') or '').strip()
    sort = (request.GET.get('sort') or 'dispatch_date').strip()
    direction = (request.GET.get('dir') or 'desc').strip().lower()
    per_page = request.GET.get('per_page') or '50'
    try:
        per_page = int(per_page)
    except (TypeError, ValueError):
        per_page = 50
    if per_page not in (50, 100):
        per_page = 50

    records = Dispatch.objects.filter(is_active=True, job_card__is_active=True).select_related('job_card', 'created_by').order_by('-dispatch_date', '-id')
    if query:
        records = records.filter(
            Q(job_card__job_card_no__icontains=query) |
            Q(job_card__SKU__icontains=query) |
            Q(job_card__destination__icontains=query) |
            Q(dc_no__icontains=query)
        )

    if dc_filter:
        records = records.filter(dc_no__iexact=dc_filter)

    if balance_status in {'complete', 'partial', 'not_dispatched'}:
        jc_qs = JobCard.objects.filter(is_active=True).annotate(
            dispatched_total=Coalesce(
                Sum('dispatch__dispatch_qty', filter=Q(dispatch__is_active=True)),
                0,
            )
        )
        if balance_status == 'complete':
            jc_qs = jc_qs.filter(dispatched_total__gte=F('order_qty'))
        elif balance_status == 'not_dispatched':
            jc_qs = jc_qs.filter(dispatched_total=0)
        else:
            jc_qs = jc_qs.filter(dispatched_total__gt=0, dispatched_total__lt=F('order_qty'))
        records = records.filter(job_card_id__in=jc_qs.values('id'))

    date_from = None
    date_to = None
    if date_from_raw:
        try:
            date_from = datetime.strptime(date_from_raw, '%Y-%m-%d').date()
            records = records.filter(dispatch_date__gte=date_from)
        except ValueError:
            add_unique_message(request, messages.ERROR, 'Invalid From date format. Use YYYY-MM-DD.')
            date_from_raw = ''
    if date_to_raw:
        try:
            date_to = datetime.strptime(date_to_raw, '%Y-%m-%d').date()
            records = records.filter(dispatch_date__lte=date_to)
        except ValueError:
            add_unique_message(request, messages.ERROR, 'Invalid To date format. Use YYYY-MM-DD.')
            date_to_raw = ''

    if date_from and date_to and date_from > date_to:
        add_unique_message(request, messages.ERROR, 'From date cannot be later than To date.')
        records = records.none()

    sortable_fields = {
        'date': 'dispatch_date',
        'job_card': 'job_card__job_card_no',
        'sku': 'job_card__SKU',
        'dc_no': 'dc_no',
        'qty': 'dispatch_qty',
        'added_by': 'created_by__username',
    }
    order_field = sortable_fields.get(sort, 'dispatch_date')
    if direction not in ('asc', 'desc'):
        direction = 'desc'
    ordering = order_field if direction == 'asc' else f'-{order_field}'
    records = records.order_by(ordering)

    total_count = records.count()
    filtered_qty_total = records.aggregate(total=Sum('dispatch_qty'))['total'] or 0
    paginator = Paginator(records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    records = list(page_obj.object_list)
    page_qty_total = sum(row.dispatch_qty or 0 for row in records)

    job_card_ids = {row.job_card_id for row in records}
    dispatch_totals = {
        row['job_card_id']: row['total'] or 0
        for row in Dispatch.objects.filter(is_active=True, job_card_id__in=job_card_ids)
        .values('job_card_id')
        .annotate(total=Sum('dispatch_qty'))
    }
    for row in records:
        total_dispatch = dispatch_totals.get(row.job_card_id, 0)
        row.remaining_badge = _dispatch_remaining_badge(row.job_card.order_qty, total_dispatch)

    # Batched into one query instead of 2 queries per distinct DC number on the
    # page (previously up to ~200 queries for a 100-row page). Still preserves
    # the original per-value case-insensitive (iexact) grouping semantics.
    dc_values = list({row.dc_no for row in records if row.dc_no})
    dc_bundle_map = {}
    if dc_values:
        dc_match_filter = Q()
        for dc_value in dc_values:
            dc_match_filter |= Q(dc_no__iexact=dc_value)
        grouped = {}
        for dc_row in Dispatch.objects.filter(is_active=True).filter(dc_match_filter).values('dc_no', 'job_card__SKU'):
            key = (dc_row['dc_no'] or '').strip().lower()
            bucket = grouped.setdefault(key, {'line_count': 0, 'skus': set()})
            bucket['line_count'] += 1
            if dc_row['job_card__SKU']:
                bucket['skus'].add(dc_row['job_card__SKU'])
        for dc_value in dc_values:
            key = dc_value.strip().lower()
            bucket = grouped.get(key, {'line_count': 0, 'skus': set()})
            dc_bundle_map[key] = {
                'line_count': bucket['line_count'],
                'sku_count': len(bucket['skus']),
                'dc_no': dc_value,
            }

    for row in records:
        row.dc_bundle = dc_bundle_map.get(
            (row.dc_no or '').strip().lower(),
            {'line_count': 1, 'sku_count': 1, 'dc_no': row.dc_no},
        )

    page_complete_count = sum(1 for row in records if row.remaining_badge['badge_class'] == 'erp-badge-completed')
    page_open_count = len(records) - page_complete_count

    cutoff = get_record_edit_lock_cutoff()
    pending_ids: set = set()
    approved_ids: set = set()
    if cutoff and not user_can_bypass_edit_lock(request.user):
        user_overrides = EditOverrideRequest.objects.filter(
            entity_type='dispatch',
            requested_by=request.user,
        ).values('record_id', 'status', 'expires_at')
        for ov in user_overrides:
            if ov['status'] == 'pending':
                pending_ids.add(ov['record_id'])
            elif ov['status'] == 'approved' and ov['expires_at'] and ov['expires_at'] > timezone.now():
                approved_ids.add(ov['record_id'])

    context = {
        'records': records,
        'page_obj': page_obj,
        'total_count': total_count,
        'filtered_qty_total': filtered_qty_total,
        'page_qty_total': page_qty_total,
        'page_complete_count': page_complete_count,
        'page_open_count': page_open_count,
        'has_active_filters': bool(query or dc_filter or balance_status or date_from_raw or date_to_raw),
        'q': query,
        'dc_no': dc_filter,
        'balance_status': balance_status,
        'today': timezone.now().date().isoformat(),
        'week_start': (timezone.now().date() - timedelta(days=timezone.now().weekday())).isoformat(),
        'month_start': timezone.now().date().replace(day=1).isoformat(),
        'date_from': date_from_raw,
        'date_to': date_to_raw,
        'sort': sort,
        'dir': direction,
        'per_page': per_page,
        'edit_lock_days': get_record_edit_lock_days(),
        'edit_lock_cutoff': cutoff,
        'can_bypass_edit_lock': user_can_bypass_edit_lock(request.user),
        'pending_override_ids': pending_ids,
        'approved_override_ids': approved_ids,
    }
    return render(request, 'dispatch_records.html', context)


@login_required
def change_history(request, entity_type, record_id):
    config = AUDIT_CONFIG.get(entity_type)
    if not config:
        messages.error(request, 'Unsupported history request.')
        return redirect('home')

    if not user_has_entity_permission(request.user, entity_type):
        add_unique_message(request, messages.ERROR, '❌ You do not have permission to access this feature.')
        return redirect('home')

    record = get_object_or_404(config['model'], pk=record_id)
    history_entries = ChangeLog.objects.filter(entity_type=entity_type, record_id=record_id).select_related('changed_by')

    context = {
        'entity_type': entity_type,
        'entity_label': config['model']._meta.verbose_name.title(),
        'record': record,
        'history_entries': history_entries,
        'back_view_name': config['list_view'],
    }
    return render(request, 'change_history.html', context)


@login_required
def archived_records(request):
    if not user_can_archive_records(request.user):
        add_unique_message(request, messages.ERROR, '❌ You do not have permission to access this feature.')
        return redirect('home')

    accessible_entities = get_accessible_entities(request.user)

    requested_entity = (request.GET.get('entity') or '').strip().lower()
    if requested_entity == 'all':
        entity_type = 'all'
    elif requested_entity in accessible_entities:
        entity_type = requested_entity
    else:
        entity_type = 'all'

    query = (request.GET.get('q') or '').strip()

    def build_archived_queryset(entity):
        config = AUDIT_CONFIG[entity]
        records = config['model'].objects.filter(is_active=False)

        if entity == 'job_card':
            records = records.select_related('material', 'machine_name', 'department').order_by('-created_at')
            if query:
                records = records.filter(
                    Q(job_card_no__icontains=query) |
                    Q(SKU__icontains=query) |
                    Q(PO_No__icontains=query)
                )
        elif entity == 'production':
            records = records.select_related('job_card', 'machine', 'operator').order_by('-date', '-id')
            if query:
                records = records.filter(
                    Q(job_card__job_card_no__icontains=query) |
                    Q(machine__name__icontains=query) |
                    Q(operator__name__icontains=query)
                )
        else:
            records = records.select_related('job_card').order_by('-dispatch_date', '-id')
            if query:
                records = records.filter(
                    Q(job_card__job_card_no__icontains=query) |
                    Q(dc_no__icontains=query)
                )

        return records

    records = None
    records_by_entity = None
    if entity_type == 'all':
        records_by_entity = {entity: build_archived_queryset(entity) for entity in accessible_entities}
    else:
        records = build_archived_queryset(entity_type)

    context = {
        'accessible_entities': accessible_entities,
        'entity_type': entity_type,
        'query': query,
        'records': records,
        'records_by_entity': records_by_entity,
        'all_mode': entity_type == 'all',
    }
    return render(request, 'archived_records.html', context)


@login_required
def delete_record(request, entity_type, record_id):
    config = AUDIT_CONFIG.get(entity_type)
    if not config:
        raise Http404('Unsupported record type')

    if not user_can_archive_records(request.user):
        add_unique_message(request, messages.ERROR, '❌ You do not have permission to access this feature.')
        return redirect('home')

    record = get_active_record_or_404(config['model'], record_id)

    if request.method == 'POST':
        reason = (request.POST.get('delete_reason') or '').strip()
        if not reason:
            messages.error(request, 'Delete reason is required.')
        else:
            try:
                validate_delete_allowed(entity_type, record)
                archive_record(entity_type, record, request.user, reason)
                messages.success(request, f'{config["model"]._meta.verbose_name.title()} archived successfully.')
                return redirect(config['list_view'])
            except Exception as exc:
                messages.error(request, str(exc))

    context = {
        'entity_type': entity_type,
        'entity_label': config['model']._meta.verbose_name.title(),
        'record': record,
        'back_view_name': config['list_view'],
    }
    return render(request, 'confirm_delete.html', context)


@login_required
def restore_record(request, entity_type, record_id):
    config = AUDIT_CONFIG.get(entity_type)
    if not config:
        raise Http404('Unsupported record type')

    if not user_can_archive_records(request.user):
        add_unique_message(request, messages.ERROR, '❌ You do not have permission to access this feature.')
        return redirect('home')

    record = get_inactive_record_or_404(config['model'], record_id)

    if request.method == 'POST':
        reason = (request.POST.get('restore_reason') or '').strip()
        if not reason:
            messages.error(request, 'Restore reason is required.')
        else:
            try:
                validate_restore_allowed(entity_type, record)
                restore_record_state(entity_type, record, request.user, reason)
                messages.success(request, f'{config["model"]._meta.verbose_name.title()} restored successfully.')
                return redirect(f"{reverse('archived_records')}?entity={entity_type}")
            except Exception as exc:
                messages.error(request, str(exc))

    context = {
        'entity_type': entity_type,
        'entity_label': config['model']._meta.verbose_name.title(),
        'record': record,
    }
    return render(request, 'confirm_restore.html', context)


OVERRIDE_EDIT_WINDOW_HOURS = 2


@login_required
def request_edit_override(request, entity_type, record_id):
    """Operational user submits a reason-based request to edit a locked record."""
    config = AUDIT_CONFIG.get(entity_type)
    if config is None or entity_type not in ('job_card', 'production', 'dispatch'):
        messages.error(request, 'Override requests are not supported for this record type.')
        return redirect('home')

    if not user_has_entity_permission(request.user, entity_type):
        add_unique_message(request, messages.ERROR, '❌ You do not have permission to access this feature.')
        return redirect('home')

    record = get_active_record_or_404(config['model'], record_id)
    entry_view_name = config['list_view'].replace('_records', '_entry')

    if not record_is_time_locked(entity_type, record):
        messages.info(request, 'This record is not locked — you can edit it directly.')
        return redirect(f"{reverse(entry_view_name)}?edit={record_id}")

    if user_can_bypass_edit_lock(request.user):
        return redirect(f"{reverse(entry_view_name)}?edit={record_id}")

    existing = EditOverrideRequest.objects.filter(
        entity_type=entity_type,
        record_id=record_id,
        requested_by=request.user,
        status='pending',
    ).first()
    if existing:
        messages.info(request, 'You already have a pending override request for this record.')
        return redirect(config['list_view'])

    if request.method == 'POST':
        reason = (request.POST.get('reason') or '').strip()
        if not reason:
            messages.error(request, 'A reason is required for the override request.')
        else:
            EditOverrideRequest.objects.create(
                entity_type=entity_type,
                record_id=record_id,
                record_label=str(record),
                requested_by=request.user,
                reason=reason,
            )
            messages.success(request, 'Override request submitted. You will be able to edit once a manager approves it.')
            return redirect(config['list_view'])

    context = {
        'entity_type': entity_type,
        'entity_label': config['model']._meta.verbose_name.title(),
        'record': record,
        'back_view_name': config['list_view'],
        'override_hours': OVERRIDE_EDIT_WINDOW_HOURS,
    }
    return render(request, 'request_edit_override.html', context)


@login_required
def override_requests(request):
    """Manager/admin inbox of all override requests."""
    if not user_can_archive_records(request.user):
        add_unique_message(request, messages.ERROR, '❌ You do not have permission to access this feature.')
        return redirect('home')

    status_filter = (request.GET.get('status') or 'pending').strip()
    qs_all = EditOverrideRequest.objects.select_related('requested_by', 'reviewed_by').all()
    if status_filter in ('pending', 'approved', 'rejected'):
        qs = qs_all.filter(status=status_filter)
    else:
        qs = qs_all
        status_filter = 'all'

    context = {
        'override_list': qs,
        'status_filter': status_filter,
        'pending_count': qs_all.filter(status='pending').count(),
    }
    return render(request, 'override_requests.html', context)


@login_required
def review_override_request(request, override_id):
    """Manager/admin approves or rejects an override request."""
    if not user_can_archive_records(request.user):
        add_unique_message(request, messages.ERROR, '❌ You do not have permission to access this feature.')
        return redirect('home')

    override = get_object_or_404(EditOverrideRequest, pk=override_id, status='pending')

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        note = (request.POST.get('review_note') or '').strip()

        if action not in ('approve', 'reject'):
            messages.error(request, 'Invalid action.')
        else:
            override.reviewed_by = request.user
            override.review_note = note
            override.reviewed_at = timezone.now()
            entry_view_name = AUDIT_CONFIG.get(override.entity_type, {}).get(
                'list_view', 'production_records'
            ).replace('_records', '_entry')
            try:
                edit_link = f"{reverse(entry_view_name)}?edit={override.record_id}"
            except Exception:
                edit_link = reverse('my_override_requests')
            if action == 'approve':
                override.status = 'approved'
                override.expires_at = timezone.now() + timedelta(hours=OVERRIDE_EDIT_WINDOW_HOURS)
                messages.success(
                    request,
                    f'Override approved. {override.requested_by.get_full_name() or override.requested_by.username}'
                    f' can now edit the record for {OVERRIDE_EDIT_WINDOW_HOURS} hour(s).'
                )
                Notification.objects.create(
                    user=override.requested_by,
                    event_type='override.approved',
                    title='Edit override approved',
                    message=(
                        f'Your request to edit {override.record_label} was approved. '
                        f'You can edit it for the next {OVERRIDE_EDIT_WINDOW_HOURS} hour(s).'
                        + (f' Note: {note}' if note else '')
                    ),
                    link=edit_link,
                    entity_type=override.entity_type,
                    entity_id=override.record_id,
                    created_by=request.user,
                )
            else:
                override.status = 'rejected'
                messages.success(request, 'Override request rejected.')
                Notification.objects.create(
                    user=override.requested_by,
                    event_type='override.rejected',
                    title='Edit override rejected',
                    message=(
                        f'Your request to edit {override.record_label} was rejected.'
                        + (f' Reason: {note}' if note else '')
                    ),
                    link=reverse('my_override_requests'),
                    entity_type=override.entity_type,
                    entity_id=override.record_id,
                    created_by=request.user,
                )
            override.save()
            return redirect('override_requests')

    context = {
        'override': override,
        'override_hours': OVERRIDE_EDIT_WINDOW_HOURS,
    }
    return render(request, 'review_override_request.html', context)


@login_required
def my_override_requests(request):
    """Requester-facing list of the user's edit-override requests and status.

    Shows what needs their action (approved & ready to edit), what is still
    waiting on a manager, and a short history of resolved requests. Approved
    items drop out of the actionable list once the edit is done (consumed_at)
    or the window expires.
    """
    now = timezone.now()
    qs = EditOverrideRequest.objects.filter(requested_by=request.user).order_by('-created_at')

    def _edit_link(ov):
        entry_view = AUDIT_CONFIG.get(ov.entity_type, {}).get(
            'list_view', 'production_records'
        ).replace('_records', '_entry')
        try:
            return f"{reverse(entry_view)}?edit={ov.record_id}"
        except Exception:
            return ''

    actionable, pending, history = [], [], []
    for ov in qs:
        if ov.status == 'pending':
            pending.append(ov)
        elif ov.is_valid_for_edit:
            ov.edit_link = _edit_link(ov)
            actionable.append(ov)
        else:
            # rejected, consumed, or expired
            if ov.status == 'approved' and ov.consumed_at is None and ov.expires_at and ov.expires_at <= now:
                ov.is_expired = True
            history.append(ov)

    context = {
        'actionable': actionable,
        'pending': pending,
        'history': history[:25],
        'override_hours': OVERRIDE_EDIT_WINDOW_HOURS,
    }
    return render(request, 'my_override_requests.html', context)


def build_erp_readme_text():
    return """Offset Printing ERP - Easy User Guide

Last Updated: 2026-04-17

=============================
1) JOB CARD ENTRY (Planning)
=============================
Purpose:
- This is the planning sheet for one customer order.

Important fields (simple meaning):
- Job Card No: unique ID of the order.
- SKU: product code/name.
- PO Number / PO Date: customer purchase order details.
- Material: paper/board type.
- Colours: print colors (example 4 or 1+1).
- Order Qty (pcs): final quantity customer needs.
- UPS: how many pieces fit on one sheet.
- Wastage (sheets): planned extra sheets for setup/loss.
- Machine: planned machine for this job.
- Department: process area.
- Production Tolerance %: allowed extra production beyond plan.

Auto calculations:
- Required Sheets = Order Qty / UPS
- Total Planned Sheets = Required Sheets + Wastage
- Tolerance Sheets = Total Planned Sheets * Tolerance %
- Allowed Sheets = Planned Sheets + Tolerance Sheets

Time planning (auto):
- Run Time (min) = (Total Impressions Required / Machine Speed per hour) * 60
- Setup Time (min) = Total Colors * Machine Setup Minutes per Color
- Total Planned Time = Run Time + Setup Time

================================
2) PRODUCTION ENTRY (Execution)
================================
Purpose:
- Operator/supervisor logs actual production done in each shift.

Important fields:
- Job Card: which planned job is running.
- Machine: auto from Job Card (override allowed if needed).
- Operator, Shift, Date: who ran, when, which shift.
- Impressions: total machine impressions done.
- Output Sheets: good sheets produced.
- Waste Sheets + Waste Reason: scrap and reason.
- Planned Time: auto remaining planned minutes.
- Run Time, Setup Time, Downtime: actual consumed minutes.
- Downtime Category: reason bucket for downtime.

Validations:
- Output + Waste cannot exceed Allowed Sheets.
- Total impressions are cumulative across repeated production entries for the same job card.
- Total impressions are validated against Total Impressions Required plus Production Tolerance %.
- If planned allocation exceeds remaining planned minutes, overrun reason is mandatory.
- Overrun Minutes = (Run + Setup + Downtime) - Planned Time

================================
3) DISPATCH ENTRY (Delivery)
================================
Purpose:
- Record quantity sent to customer.

Important fields:
- Job Card
- Dispatch Date
- Dispatch Qty (pcs)
- DC No (optional)

Completion logic:
- Job is treated as Completed at 95%+ dispatch ratio.
- Remaining below 100% is Short Close (not auto waste).
- Manager/dispatch can close pending short-close with reason.
- Closed short-close is moved to Closed as Wastage.

================================
4) SHIFT & MACHINE SCHEDULE
================================
Purpose:
- Define realistic available capacity by shift and machine.

Shift hours fields:
- Effective From / Effective To: date range for this schedule version.
- Day + Shift A/B net hours: productive hours after breaks.

Machine work schedule fields:
- Check box ON = machine runs in that day+shift.
- Check box OFF = machine is not planned to run.

This is used in dashboard:
- Schedule Utilization % = Actual used minutes / Scheduled available minutes * 100

=============================
5) DASHBOARD (What it means)
=============================
Top KPIs:
- OEE: overall productivity quality score.
- Availability: uptime after unplanned downtime impact.
- Performance: speed efficiency vs machine standard speed.
- Quality: good output ratio.

Planning KPIs:
- Planned Time vs Actual Time
- Planned Variance
- Overrun split (setup, downtime, run-performance gap)

Dispatch and closure KPIs:
- Dispatch qty and fulfillment
- Pending Short Close
- Short Close Closed as Wastage

=============================
6) IF DROPDOWN VALUE IS WRONG
=============================
Example: wrong machine, operator, material, or department name added by mistake.

Use "Master Corrections" page:
1. Open the correction page from home/nav.
2. Rename wrong text to correct name.
3. For machine/operator, you can deactivate so it no longer appears in dropdowns.
4. Existing historical records remain safe and auditable.

=============================
7) QUICK RULES FOR USERS
=============================
- Always select correct Job Card first.
- Do not use waste to hide dispatch short-close.
- Total impressions are tracked across production entries and must stay within allowance.
- Give a reason when overriding machine or closing short-close.
- Keep shift schedule dates current for accurate utilization.

=============================
8) ARCHIVED RECORDS
=============================
- Use the Archived Records page to view deleted job cards, production, and dispatch entries.
- You can filter by entity type or use the All view to see every archived record type together.
- Restored records come back active with audit history intact.

=============================
9) MAINTENANCE NOTE
=============================
- Update this guide whenever fields, formulas, rules, or dashboard KPIs change.
"""


@login_required
def erp_readme(request):
    context = {
        'generated_on': timezone.now(),
    }
    return render(request, 'erp_readme.html', context)


@login_required
def download_erp_readme(request):
    content = build_erp_readme_text()
    response = HttpResponse(content, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="offset_erp_calculation_guide.txt"'
    return response


@login_required
@permission_required('can_manage_masters')
def master_data(request):
    """Manager/admin screen to manage dropdown master values across ERP."""

    from planning.models import PlanningJob, SkuRecipe

    User = get_user_model()
    model_map = {
        'machine': Machine,
        'operator': Operator,
        'material': Material,
        'department': Department,
        'delivery_location': DeliveryLocation,
        'product_type': ProductType,
        'application': ApplicationType,
        'supervisor': Supervisor,
        'vendor': Vendor,
        'print_color': PrintColor,
    }

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        entity_type = (request.POST.get('entity_type') or '').strip().lower()
        machine_id = (request.POST.get('machine_id') or request.POST.get('record_id') or '').strip()
        model = model_map.get(entity_type)
        record = get_object_or_404(model, pk=machine_id) if (model and machine_id) else None
        is_admin_user = bool(getattr(request.user, 'profile', None) and request.user.profile.role == 'admin')

        if action == 'delete_master' and not request.user.is_superuser:
            messages.error(request, 'Only a superuser can delete master values.')
            return redirect('master_data')

        if action in {'rename_machine', 'edit_master', 'create_master'} and not is_admin_user:
            messages.error(request, 'Only admin can edit master values.')
            return redirect('master_data')

        if action == 'create_master' and entity_type in model_map:
            new_name = (request.POST.get('new_name') or '').strip()
            label = entity_type.replace('_', ' ').title()
            if not new_name:
                messages.error(request, f'{label} name is required.')
                return redirect('master_data')
            if model.objects.filter(name__iexact=new_name).exists():
                messages.error(request, f'{label} "{new_name}" already exists.')
                return redirect('master_data')

            create_kwargs = {'name': new_name}
            if entity_type == 'print_color':
                create_kwargs['is_active'] = True
            elif entity_type in {'operator', 'supervisor'}:
                employee_code = (request.POST.get('employee_code') or '').strip() or None
                create_kwargs['employee_code'] = employee_code

            model.objects.create(**create_kwargs)
            messages.success(request, f'{label} "{new_name}" added.')
            return redirect('master_data')

        if action in {'rename_machine', 'edit_master'} and record:
            new_name = (request.POST.get('new_name') or '').strip()
            if not new_name:
                messages.error(request, 'Name is required.')
            else:
                duplicate = model.objects.exclude(pk=record.pk).filter(name__iexact=new_name).first()
                if duplicate:
                    messages.error(request, f'Name already exists as #{duplicate.id} ({duplicate.name}).')
                    return redirect('master_data')
                old_name = record.name

                changed_fields = []

                if record.name != new_name:
                    record.name = new_name
                    changed_fields.append('name')

                if entity_type in {'operator', 'supervisor'}:
                    new_employee_code = (request.POST.get('employee_code') or '').strip() or None
                    if record.employee_code != new_employee_code:
                        record.employee_code = new_employee_code
                        changed_fields.append('employee_code')

                if entity_type == 'machine':
                    machine_type_raw = (request.POST.get('machine_type') or '').strip()
                    speed_raw = (request.POST.get('standard_impressions_per_hour') or '').strip()
                    setup_raw = (request.POST.get('standard_setup_minutes_per_color') or '').strip()
                    plate_life_raw = (request.POST.get('plate_life_impressions') or '').strip()
                    group_code_raw = (request.POST.get('machine_group_code') or '').strip()
                    # Colour/size fields are optional and only meaningful for offset
                    # printing machines - a cutting machine or digital printer can be
                    # created without them; blank means "not applicable", not "1".
                    default_colors_raw = (request.POST.get('default_colors') or '').strip()
                    operational_colors_raw = (request.POST.get('operational_colors') or '').strip()
                    # Planners always work in inches (matches print_sheet_size data entry);
                    # the form collects inches and we convert to mm for storage so the DB
                    # has one unambiguous unit and there's no mm/inch guesswork on entry.
                    min_l_in_raw = (request.POST.get('min_print_length_in') or '').strip()
                    min_w_in_raw = (request.POST.get('min_print_width_in') or '').strip()
                    max_l_in_raw = (request.POST.get('max_print_length_in') or '').strip()
                    max_w_in_raw = (request.POST.get('max_print_width_in') or '').strip()

                    valid_machine_types = {code for code, _ in Machine.MACHINE_TYPE_CHOICES}
                    new_machine_type = machine_type_raw if machine_type_raw in valid_machine_types else (record.machine_type or 'other')

                    try:
                        new_speed = float(speed_raw) if speed_raw else float(record.standard_impressions_per_hour or 0)
                        new_setup = float(setup_raw) if setup_raw else float(record.standard_setup_minutes_per_color or 0)
                        new_plate_life = int(plate_life_raw) if plate_life_raw else int(record.plate_life_impressions or 25000)
                        new_default_colors = int(default_colors_raw) if default_colors_raw else record.default_colors
                        new_operational_colors = int(operational_colors_raw) if operational_colors_raw != '' else record.operational_colors
                        new_min_l = round(float(min_l_in_raw) * MM_PER_INCH, 2) if min_l_in_raw else None
                        new_min_w = round(float(min_w_in_raw) * MM_PER_INCH, 2) if min_w_in_raw else None
                        new_max_l = round(float(max_l_in_raw) * MM_PER_INCH, 2) if max_l_in_raw else None
                        new_max_w = round(float(max_w_in_raw) * MM_PER_INCH, 2) if max_w_in_raw else None
                    except ValueError:
                        messages.error(request, 'Machine speed, setup minutes, plate life, colours, and print size fields must be numeric values.')
                        return redirect('master_data')

                    if new_speed <= 0 or new_setup <= 0:
                        messages.error(request, 'Machine speed and setup minutes per color must be greater than 0.')
                        return redirect('master_data')
                    if new_plate_life < 1:
                        messages.error(request, 'Plate life impressions must be at least 1.')
                        return redirect('master_data')
                    if new_default_colors is not None and new_default_colors < 1:
                        messages.error(request, 'Default colors must be at least 1.')
                        return redirect('master_data')
                    if new_operational_colors is not None and new_operational_colors < 0:
                        messages.error(request, 'Operational colors cannot be negative.')
                        return redirect('master_data')

                    if float(record.standard_impressions_per_hour) != float(new_speed):
                        record.standard_impressions_per_hour = new_speed
                        changed_fields.append('standard_impressions_per_hour')
                    if float(record.standard_setup_minutes_per_color) != float(new_setup):
                        record.standard_setup_minutes_per_color = new_setup
                        changed_fields.append('standard_setup_minutes_per_color')
                    if int(record.plate_life_impressions or 0) != int(new_plate_life):
                        record.plate_life_impressions = new_plate_life
                        changed_fields.append('plate_life_impressions')
                    if (record.machine_type or '') != new_machine_type:
                        record.machine_type = new_machine_type
                        changed_fields.append('machine_type')
                    if (record.machine_group_code or '') != group_code_raw:
                        record.machine_group_code = group_code_raw
                        changed_fields.append('machine_group_code')
                    if record.default_colors != new_default_colors:
                        record.default_colors = new_default_colors
                        changed_fields.append('default_colors')
                    if record.operational_colors != new_operational_colors:
                        record.operational_colors = new_operational_colors
                        changed_fields.append('operational_colors')
                    if record.min_print_length_mm != new_min_l:
                        record.min_print_length_mm = new_min_l
                        changed_fields.append('min_print_length_mm')
                    if record.min_print_width_mm != new_min_w:
                        record.min_print_width_mm = new_min_w
                        changed_fields.append('min_print_width_mm')
                    if record.max_print_length_mm != new_max_l:
                        record.max_print_length_mm = new_max_l
                        changed_fields.append('max_print_length_mm')
                    if record.max_print_width_mm != new_max_w:
                        record.max_print_width_mm = new_max_w
                        changed_fields.append('max_print_width_mm')

                if changed_fields:
                    record.save(update_fields=changed_fields)
                    if entity_type == 'product_type' and old_name != record.name:
                        SkuRecipe.objects.filter(product_type__iexact=old_name).update(product_type=record.name)
                    if entity_type == 'delivery_location' and old_name != record.name:
                        PlanningJob.objects.filter(destination__iexact=old_name).update(destination=record.name)
                        JobCard.objects.filter(destination__iexact=old_name).update(destination=record.name)
                    if old_name != record.name:
                        messages.success(request, f'{entity_type.replace("_", " ").title()} updated: {old_name} -> {record.name}')
                    else:
                        messages.success(request, f'{entity_type.replace("_", " ").title()} details updated successfully.')
                elif entity_type == 'machine':
                    messages.success(request, f'No changes detected for {entity_type.title()} {record.name}.')

        elif action == 'toggle_machine' and record and hasattr(record, 'is_active'):
            record.is_active = not record.is_active
            record.save(update_fields=['is_active'])
            state = 'active' if record.is_active else 'inactive'
            display_name = record.name
            messages.success(request, f'{entity_type.title()} {display_name} marked {state}.')

        elif action == 'delete_master' and record:
            record_name = record.name

            if entity_type == 'machine':
                linked_jobcards = JobCard.objects.filter(machine_name=record).count()
                linked_productions = Production.objects.filter(machine=record).count()
                if linked_jobcards or linked_productions:
                    messages.error(
                        request,
                        f'Cannot delete Machine {record_name}. Linked records found (Job Cards: {linked_jobcards}, Production: {linked_productions}).'
                    )
                    return redirect('master_data')

            elif entity_type == 'operator':
                linked_productions = Production.objects.filter(operator=record).count()
                if linked_productions:
                    messages.error(
                        request,
                        f'Cannot delete Operator {record_name}. Linked production records: {linked_productions}.'
                    )
                    return redirect('master_data')

            elif entity_type == 'supervisor':
                linked_productions = Production.objects.filter(supervisor=record).count()
                if linked_productions:
                    messages.error(
                        request,
                        f'Cannot delete Supervisor {record_name}. Linked production records: {linked_productions}.'
                    )
                    return redirect('master_data')

            elif entity_type == 'material':
                linked_jobcards = JobCard.objects.filter(material=record).count()
                if linked_jobcards:
                    messages.error(
                        request,
                        f'Cannot delete Material {record_name}. Linked job cards: {linked_jobcards}.'
                    )
                    return redirect('master_data')

            elif entity_type == 'department':
                linked_jobcards = JobCard.objects.filter(department=record).count()
                linked_planning_jobs = PlanningJob.objects.filter(department__iexact=record.name).count()
                if linked_jobcards or linked_planning_jobs:
                    messages.error(
                        request,
                        f'Cannot delete Department {record_name}. Linked records found (Job Cards: {linked_jobcards}, Planning Jobs: {linked_planning_jobs}).'
                    )
                    return redirect('master_data')

            elif entity_type == 'delivery_location':
                linked_jobcards = JobCard.objects.filter(destination__iexact=record.name).count()
                linked_planning_jobs = PlanningJob.objects.filter(destination__iexact=record.name).count()
                if linked_jobcards or linked_planning_jobs:
                    messages.error(
                        request,
                        f'Cannot delete Delivery Location {record_name}. Linked records found (Job Cards: {linked_jobcards}, Planning Jobs: {linked_planning_jobs}).'
                    )
                    return redirect('master_data')

            elif entity_type == 'product_type':
                linked_recipes = SkuRecipe.objects.filter(product_type__iexact=record.name).count()
                linked_planning_jobs = count_active_planning_jobs_for_product_type(record.name)
                if linked_recipes or linked_planning_jobs:
                    messages.error(
                        request,
                        f'Cannot delete Product Type {record_name}. Linked records found (SKU Recipes: {linked_recipes}, Planning Jobs: {linked_planning_jobs}).'
                    )
                    return redirect('master_data')

            try:
                record.delete()
                messages.success(request, f'{entity_type.title()} {record_name} deleted successfully.')
            except ProtectedError:
                messages.error(request, f'Cannot delete {entity_type.title()} {record_name} because it is referenced by other records.')

        return redirect('master_data')

    departments_created = sync_departments_from_planning()
    if departments_created:
        messages.success(
            request,
            f'Added {departments_created} department{"s" if departments_created != 1 else ""} from planning data.',
        )

    delivery_locations_created = sync_delivery_locations_from_planning()
    if delivery_locations_created:
        messages.success(
            request,
            f'Added {delivery_locations_created} delivery location{"s" if delivery_locations_created != 1 else ""} from planning data.',
        )

    product_types_created = sync_product_types_from_sku_recipes()
    if product_types_created:
        messages.success(
            request,
            f'Added {product_types_created} product type{"s" if product_types_created != 1 else ""} from SKU master recipes.',
        )

    materials_created = sync_materials_from_planning()
    if materials_created:
        messages.success(
            request,
            f'Added {materials_created} material{"s" if materials_created != 1 else ""} from planning data.',
        )

    machine_rows = []
    from collections import Counter
    from printing_plates.models import PlateRequest

    # Fetch values for in-memory mapping in 6 single queries
    jc_values = list(JobCard.objects.filter(is_active=True).values('machine_name_id', 'material_id', 'department_id', 'destination'))
    jc_machine_counter = Counter(x['machine_name_id'] for x in jc_values if x['machine_name_id'])
    jc_material_counter = Counter(x['material_id'] for x in jc_values if x['material_id'])
    jc_department_counter = Counter(x['department_id'] for x in jc_values if x['department_id'])
    jc_destination_counter = Counter((x['destination'] or '').strip().lower() for x in jc_values if x['destination'])

    prod_values = list(Production.objects.filter(is_active=True).values('machine_id', 'operator_id', 'supervisor_id'))
    prod_machine_counter = Counter(x['machine_id'] for x in prod_values if x['machine_id'])
    prod_operator_counter = Counter(x['operator_id'] for x in prod_values if x['operator_id'])
    prod_supervisor_counter = Counter(x['supervisor_id'] for x in prod_values if x['supervisor_id'])

    pj_values = list(PlanningJob.objects.filter(is_active=True).values('department', 'destination', 'color_spec', 'sku'))
    pj_department_counter = Counter((x['department'] or '').strip().lower() for x in pj_values if x['department'])
    pj_destination_counter = Counter((x['destination'] or '').strip().lower() for x in pj_values if x['destination'])
    pj_color_counter = Counter((x['color_spec'] or '').strip().lower() for x in pj_values if x['color_spec'])

    recipe_values = list(SkuRecipe.objects.filter(is_active=True).values('product_type', 'color_spec', 'sku'))
    recipe_product_type_counter = Counter((x['product_type'] or '').strip().lower() for x in recipe_values if x['product_type'])
    recipe_color_counter = Counter((x['color_spec'] or '').strip().lower() for x in recipe_values if x['color_spec'])

    plate_req_values = list(PlateRequest.objects.values_list('vendor', flat=True))
    plate_req_vendor_counter = Counter(v.strip().lower() for v in plate_req_values if v)

    # Pre-map SKU -> Product Type for planning job product type counts
    sku_to_pt = {
        x['sku'].strip().lower(): x['product_type'].strip().lower()
        for x in recipe_values
        if x['sku'] and x['product_type']
    }
    pj_pt_counter = Counter()
    for pj in pj_values:
        sku_key = (pj['sku'] or '').strip().lower()
        pt = sku_to_pt.get(sku_key)
        if pt:
            pj_pt_counter[pt] += 1

    # Build rows
    machine_rows = []
    for item in Machine.objects.all().order_by('name', 'id'):
        def _mm_to_in(value):
            return round(float(value) / MM_PER_INCH, 3) if value is not None else None

        machine_rows.append({
            'record': item,
            'job_card_count': jc_machine_counter[item.id],
            'production_count': prod_machine_counter[item.id],
            'min_print_length_in': _mm_to_in(item.min_print_length_mm),
            'min_print_width_in': _mm_to_in(item.min_print_width_mm),
            'max_print_length_in': _mm_to_in(item.max_print_length_mm),
            'max_print_width_in': _mm_to_in(item.max_print_width_mm),
        })

    operator_rows = []
    for item in Operator.objects.all().order_by('name', 'id'):
        operator_rows.append({
            'record': item,
            'production_count': prod_operator_counter[item.id],
        })

    material_rows = []
    for item in Material.objects.all().order_by('name', 'id'):
        material_rows.append({
            'record': item,
            'job_card_count': jc_material_counter[item.id],
        })

    department_rows = []
    for item in Department.objects.all().order_by('name', 'id'):
        name_key = (item.name or '').strip().lower()
        department_rows.append({
            'record': item,
            'job_card_count': jc_department_counter[item.id],
            'planning_job_count': pj_department_counter[name_key],
        })

    delivery_location_rows = []
    for item in DeliveryLocation.objects.all().order_by('name', 'id'):
        name_key = (item.name or '').strip().lower()
        delivery_location_rows.append({
            'record': item,
            'job_card_count': jc_destination_counter[name_key],
            'planning_job_count': pj_destination_counter[name_key],
        })

    product_type_rows = []
    for item in ProductType.objects.all().order_by('name', 'id'):
        name_key = (item.name or '').strip().lower()
        product_type_rows.append({
            'record': item,
            'planning_job_count': pj_pt_counter[name_key],
            'sku_recipe_count': recipe_product_type_counter[name_key],
        })

    supervisor_rows = []
    for item in Supervisor.objects.all().order_by('name', 'id'):
        supervisor_rows.append({
            'record': item,
            'production_count': prod_supervisor_counter[item.id],
        })

    vendor_rows = []
    for item in Vendor.objects.all().order_by('name', 'id'):
        name_key = (item.name or '').strip().lower()
        vendor_rows.append({
            'record': item,
            'plate_request_count': plate_req_vendor_counter[name_key],
        })

    print_color_rows = []
    for item in PrintColor.objects.all().order_by('sort_order', 'name', 'id'):
        name_key = (item.name or '').strip().lower()
        print_color_rows.append({
            'record': item,
            'sku_recipe_count': recipe_color_counter[name_key],
            'planning_job_count': pj_color_counter[name_key],
        })

    application_type_rows = [{'record': item} for item in ApplicationType.objects.all().order_by('name', 'id')]

    context = {
        'machine_rows': machine_rows,
        'machine_type_choices': Machine.MACHINE_TYPE_CHOICES,
        'operator_rows': operator_rows,
        'supervisor_rows': supervisor_rows,
        'material_rows': material_rows,
        'department_rows': department_rows,
        'delivery_location_rows': delivery_location_rows,
        'product_type_rows': product_type_rows,
        'application_type_rows': application_type_rows,
        'vendor_rows': vendor_rows,
        'print_color_rows': print_color_rows,
        'is_admin_user': bool(getattr(request.user, 'profile', None) and request.user.profile.role == 'admin'),
        'is_superuser_user': bool(request.user.is_superuser),
    }
    return render(request, 'master_data.html', context)


@login_required
@permission_required('can_manage_masters')
def machine_master_tools(request):
    """Backward-compatible redirect for the renamed master data route."""
    return redirect('master_data')


@login_required
@require_role('admin')
def manage_user_roles(request):
    """Admin interface for managing user roles"""
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    users = User.objects.select_related('profile').all()
    
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        new_role = request.POST.get('role')
        
        try:
            user = User.objects.get(pk=user_id)
            profile = user.profile
            profile.role = new_role
            profile.save()
            messages.success(request, f'✅ {user.username} role updated to {profile.get_role_display()}')
        except (User.DoesNotExist, UserProfile.DoesNotExist) as e:
            messages.error(request, f'❌ Error updating role: {str(e)}')
        return redirect('manage_user_roles')
    
    from core.models import Role
    role_choices = list(Role.objects.order_by('display_name').values_list('slug', 'display_name'))

    context = {
        'users': users,
        'role_choices': role_choices,
    }
    return render(request, 'manage_user_roles.html', context)


@login_required
def user_create(request):
    """Superuser-only account creation. No password is ever entered here —
    the new user gets an emailed set-password link (same token flow as
    self-service password reset)."""
    if not request.user.is_superuser:
        add_unique_message(request, messages.ERROR, '❌ Only a superuser can create user accounts.')
        return redirect('notification_settings_home')

    from django.contrib.auth import get_user_model
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.encoding import force_bytes
    from django.utils.http import urlsafe_base64_encode
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from core.models import Role, Department

    User = get_user_model()
    role_choices = list(Role.objects.order_by('display_name').values_list('slug', 'display_name'))
    departments = Department.objects.all().order_by('name')

    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        email = (request.POST.get('email') or '').strip()
        first_name = (request.POST.get('first_name') or '').strip()
        last_name = (request.POST.get('last_name') or '').strip()
        role = (request.POST.get('role') or '').strip()
        department_id = request.POST.get('department') or None

        if not username or not email or not role:
            messages.error(request, 'Username, email, and role are required.')
        elif User.objects.filter(username__iexact=username).exists():
            messages.error(request, f"A user named '{username}' already exists.")
        else:
            user = User.objects.create(
                username=username, email=email,
                first_name=first_name, last_name=last_name,
                is_active=True,
            )
            user.set_unusable_password()
            user.save()

            profile = user.profile
            profile.role = role
            profile.department_id = department_id
            profile.save()

            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            reset_link = request.build_absolute_uri(
                reverse('password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
            )
            email_body = render_to_string('user_welcome_email.html', {
                'user': user, 'reset_link': reset_link,
            })
            send_mail(
                subject='Welcome to Offset ERP - Set Your Password',
                message=email_body,
                from_email=None,
                recipient_list=[email],
                fail_silently=True,
            )

            messages.success(request, f"User '{username}' created. A set-password link was emailed to {email}.")
            return redirect('notification_settings_home')

    context = {
        'role_choices': role_choices,
        'departments': departments,
    }
    return render(request, 'user_create.html', context)


@login_required
@require_POST
def email_settings_edit(request):
    """Superuser-only: store the Gmail sender used for password-reset and
    new-user emails. Read at send-time by core.email_backend, so this takes
    effect immediately with no server restart."""
    if not request.user.is_superuser:
        add_unique_message(request, messages.ERROR, '❌ Only a superuser can edit email settings.')
        return redirect('notification_settings_home')

    from core.models import EmailSettings

    gmail_address = (request.POST.get('gmail_address') or '').strip()
    gmail_app_password = (request.POST.get('gmail_app_password') or '').strip().replace(' ', '')

    settings_obj = EmailSettings.get_solo()
    settings_obj.gmail_address = gmail_address
    # Blank password field on submit means "keep the existing one" — the
    # value is never echoed back into the form, so an empty submit isn't a
    # deliberate clear.
    if gmail_app_password:
        settings_obj.gmail_app_password = gmail_app_password
    settings_obj.updated_by = request.user
    settings_obj.save()

    messages.success(request, 'Email settings saved.')
    return redirect(f"/settings/#access-control")


@login_required
def download_template(request):
    """Download template in CSV or Excel format"""
    file_format = request.GET.get('format', 'csv').lower()
    
    headers = get_template_headers()
    example = get_template_example()
    
    if file_format == 'excel' and EXCEL_AVAILABLE:
        try:
            # Generate Excel file
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Job Cards"
            
            # Add headers with styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Add example row
            for col_num, value in enumerate(example, 1):
                worksheet.cell(row=2, column=col_num, value=value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Send file
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="jobcard_template.xlsx"'
            workbook.save(response)
            return response
        except Exception as e:
            # Fallback to CSV if Excel generation fails
            print(f"Excel generation failed: {e}")
            import traceback
            traceback.print_exc()
    
    # Generate CSV file (default/fallback)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="jobcard_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerow(example)
    
    return response


@login_required
@require_role('admin', 'manager')
def shift_config(request):
    """Manage weekly shift net hours and machine work schedules."""
    days = ShiftConfig.DAY_CHOICES
    shifts = ['A', 'B']
    machines = Machine.objects.filter(
        Q(is_active=True) |
        Q(id__in=JobCard.objects.exclude(machine_name__isnull=True).values_list('machine_name_id', flat=True)) |
        Q(production__isnull=False)
    ).distinct().order_by('name', 'id')

    active_date_raw = (request.GET.get('effective_date') or '').strip()
    if active_date_raw:
        try:
            active_date = datetime.strptime(active_date_raw, '%Y-%m-%d').date()
        except ValueError:
            active_date = timezone.now().date()
    else:
        active_date = timezone.now().date()

    def parse_effective_range(post_data):
        start_raw = (post_data.get('effective_from') or '').strip()
        end_raw = (post_data.get('effective_to') or '').strip()

        if not start_raw and not end_raw:
            return (None, None, None)
        if not start_raw or not end_raw:
            return (None, None, 'Both Effective From and Effective To are required.')

        try:
            start_date = datetime.strptime(start_raw, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_raw, '%Y-%m-%d').date()
        except ValueError:
            return (None, None, 'Invalid effective date format.')

        if start_date > end_date:
            return (None, None, 'Effective From cannot be after Effective To.')

        return (start_date, end_date, None)

    def range_filter_for(target_date):
        return (
            Q(effective_from__isnull=True, effective_to__isnull=True) |
            Q(effective_from__lte=target_date, effective_to__gte=target_date)
        )

    if request.method == 'POST':
        action = request.POST.get('action')
        effective_from, effective_to, range_error = parse_effective_range(request.POST)

        if range_error:
            messages.error(request, range_error)
            return redirect('shift_config')

        if action == 'save_hours':
            for day_val, _ in days:
                for shift_val in shifts:
                    key = f'hours_{day_val}_{shift_val}'
                    raw = (request.POST.get(key) or '').strip()
                    if not raw:
                        continue
                    try:
                        net_h = float(raw)
                    except ValueError:
                        continue
                    ShiftConfig.objects.update_or_create(
                        day_of_week=day_val,
                        shift=shift_val,
                        effective_from=effective_from,
                        effective_to=effective_to,
                        defaults={'net_hours': net_h},
                    )
            messages.success(request, 'Shift hours saved.')

        elif action == 'save_schedule':
            for machine in machines:
                for day_val, _ in days:
                    for shift_val in shifts:
                        key = f'work_{machine.id}_{day_val}_{shift_val}'
                        is_working = request.POST.get(key) == 'on'
                        MachineWorkSchedule.objects.update_or_create(
                            machine=machine,
                            day_of_week=day_val,
                            shift=shift_val,
                            effective_from=effective_from,
                            effective_to=effective_to,
                            defaults={'is_working': is_working},
                        )
            messages.success(request, 'Machine schedule saved.')

        return redirect('shift_config')

    raw_hours = {}
    for sc in ShiftConfig.objects.filter(range_filter_for(active_date)).order_by('day_of_week', 'shift', 'effective_from', 'id'):
        raw_hours[(sc.day_of_week, sc.shift)] = sc.net_hours
    shift_hours_rows = []
    for day_val, day_name in days:
        shift_hours_rows.append({
            'day_val': day_val,
            'day_name': day_name,
            'A': raw_hours.get((day_val, 'A'), ''),
            'B': raw_hours.get((day_val, 'B'), ''),
        })

    working_map = {}
    for ms in MachineWorkSchedule.objects.filter(range_filter_for(active_date)).order_by('effective_from', 'id'):
        working_map[f'{ms.machine_id}_{ms.day_of_week}_{ms.shift}'] = bool(ms.is_working)

    machine_rows = []
    for machine in machines:
        cells = []
        for day_val, _ in days:
            for shift_val in shifts:
                key = f'{machine.id}_{day_val}_{shift_val}'
                cells.append({
                    'name': f'work_{machine.id}_{day_val}_{shift_val}',
                    'is_working': working_map.get(key, True),
                })
        machine_rows.append({'machine': machine, 'cells': cells})

    context = {
        'days': days,
        'shifts': shifts,
        'machines': machines,
        'shift_hours_rows': shift_hours_rows,
        'machine_rows': machine_rows,
        'active_effective_date': active_date.isoformat(),
        'active_effective_from': '',
        'active_effective_to': '',
    }
    return render(request, 'shift_config.html', context)


# ==========================================
# RULE-BASED NOTIFICATION & SETTINGS VIEWS
# ==========================================

from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import Http404, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages

def forgot_password(request):
    """
    Locally mocked forgot password view appropriate for a local-server setup.
    """
    if request.method == 'POST':
        username_or_email = request.POST.get('username_or_email', '').strip()
        if username_or_email:
            from core.models import PasswordResetRequest
            PasswordResetRequest.objects.create(username_or_email=username_or_email)
            messages.success(
                request,
                f"Password reset request logged for '{username_or_email}'. "
                "Since this ERP runs on a local server, please contact your local IT Administrator "
                "or HR Manager to approve this reset and retrieve a temporary password."
            )
        else:
            messages.error(request, "Please enter your username or email address.")
        return render(request, 'forgot_password.html', {'username_or_email': username_or_email, 'success': True})
    return render(request, 'forgot_password.html')


@login_required
def notification_settings_home(request):
    """
    Main workspace settings dashboard for rules, events, transitions, and audits.
    """
    if request.user.profile.role not in ('admin', 'manager'):
        messages.error(request, "You are not authorized to view settings.")
        return redirect('home')

    from core.models import (
        NotificationEvent, NotificationRule, WorkflowTransition, NotificationRuleAuditLog, Department,
        Role, Permission, UserPermissionOverride, AccessControlAuditLog,
    )
    from django.contrib.auth import get_user_model
    User = get_user_model()

    events = NotificationEvent.objects.all().order_by('name')
    rules = NotificationRule.objects.all().select_related('event', 'user', 'department').order_by('event__name', 'id')
    transitions = WorkflowTransition.objects.all().order_by('module', 'current_stage')
    audits_raw = NotificationRuleAuditLog.objects.all().select_related('changed_by', 'rule').order_by('-timestamp')[:50]
    audits = []
    for audit in audits_raw:
        changes = []
        if audit.action == 'update':
            for k, v in audit.new_values.items():
                old_val = audit.old_values.get(k, '')
                changes.append({
                    'field': k,
                    'old': old_val,
                    'new': v
                })
        audit.changes_list = changes
        audits.append(audit)

    departments = Department.objects.all().order_by('name')
    users = User.objects.filter(is_active=True).order_by('username')
    roles = UserProfile.ROLE_CHOICES
    all_users = User.objects.select_related('profile', 'profile__department').order_by('-is_active', 'username') if request.user.is_superuser else User.objects.none()

    # --- Access control tab data ---
    access_roles = Role.objects.all().prefetch_related('permissions').order_by('display_name')
    all_permissions = Permission.objects.filter(is_active=True).order_by('category', 'name')
    permission_categories = {}
    for perm in all_permissions:
        permission_categories.setdefault(perm.category, []).append(perm)

    selected_role_id = request.GET.get('role_id')
    selected_role = None
    selected_role_codes = set()
    if selected_role_id:
        selected_role = access_roles.filter(id=selected_role_id).first()
    if selected_role is None:
        selected_role = access_roles.first()
    if selected_role is not None:
        selected_role_codes = set(selected_role.permissions.values_list('code', flat=True))

    role_category_open = {
        category: any(perm.code in selected_role_codes for perm in perms)
        for category, perms in permission_categories.items()
    }

    selected_user_id = request.GET.get('user_id')
    selected_override_user = None
    override_rows = []
    override_categories = {}
    if selected_user_id:
        selected_override_user = users.filter(id=selected_user_id).first()
    if selected_override_user is not None:
        target_profile = getattr(selected_override_user, 'profile', None)
        target_role = access_roles.filter(slug=(getattr(target_profile, 'role', '') or '').strip().lower()).first()
        role_codes = set(target_role.permissions.values_list('code', flat=True)) if target_role else set()
        override_by_code = {
            o.permission.code: o.granted
            for o in UserPermissionOverride.objects.filter(user=selected_override_user).select_related('permission')
        }
        for perm in all_permissions:
            from_role = perm.code in role_codes
            override = override_by_code.get(perm.code)
            effective = override if override is not None else from_role
            row = {
                'permission': perm,
                'from_role': from_role,
                'is_override': override is not None,
                'effective': effective,
            }
            override_rows.append(row)
            override_categories.setdefault(perm.category, []).append(row)

    override_category_open = {
        category: any(row['effective'] for row in rows)
        for category, rows in override_categories.items()
    }

    access_audits_raw = AccessControlAuditLog.objects.all().select_related('changed_by').order_by('-timestamp')[:50]

    from core.models import EmailSettings
    email_settings = EmailSettings.get_solo() if request.user.is_superuser else None

    context = {
        'events': events,
        'rules': rules,
        'transitions': transitions,
        'audits': audits,
        'departments': departments,
        'users': users,
        'all_users': all_users,
        'roles': roles,
        'access_roles': access_roles,
        'permission_categories': permission_categories,
        'selected_role': selected_role,
        'selected_role_codes': selected_role_codes,
        'role_category_open': role_category_open,
        'selected_override_user': selected_override_user,
        'override_rows': override_rows,
        'override_categories': override_categories,
        'override_category_open': override_category_open,
        'access_audits': access_audits_raw,
        'email_settings': email_settings,
    }
    return render(request, 'notification_settings.html', context)


@require_role('admin')
@require_POST
def access_role_create(request):
    from django.utils.text import slugify
    from core.models import Role, AccessControlAuditLog

    slug = slugify((request.POST.get('slug') or '').strip())[:30]
    display_name = (request.POST.get('display_name') or '').strip()
    description = (request.POST.get('description') or '').strip()

    if not slug or not display_name:
        messages.error(request, "Role slug and display name are required.")
        return redirect('notification_settings_home')

    role, created = Role.objects.get_or_create(
        slug=slug,
        defaults={'display_name': display_name, 'description': description},
    )
    if not created:
        messages.error(request, f"A role with slug '{slug}' already exists.")
    else:
        AccessControlAuditLog.objects.create(
            changed_by=request.user, action='create', target_type='role',
            target_label=role.display_name,
            new_values={'slug': slug, 'display_name': display_name, 'description': description},
        )
        messages.success(request, f"Role '{role.display_name}' created. Tick its permissions below.")
    return redirect(f"/settings/?role_id={role.id}#access-control")


@require_role('admin')
@require_POST
def access_role_delete(request, role_id):
    from core.models import Role, AccessControlAuditLog

    role = get_object_or_404(Role, id=role_id)
    if role.is_system:
        messages.error(request, "Built-in roles cannot be deleted.")
        return redirect('notification_settings_home')

    AccessControlAuditLog.objects.create(
        changed_by=request.user, action='delete', target_type='role',
        target_label=role.display_name,
        old_values={'slug': role.slug, 'display_name': role.display_name},
    )
    role.delete()
    messages.success(request, "Role deleted.")
    return redirect('notification_settings_home')


@require_role('admin')
@require_POST
def access_role_permissions_edit(request):
    from core.models import Role, Permission, AccessControlAuditLog

    role = get_object_or_404(Role, id=request.POST.get('role_id'))
    old_codes = sorted(role.permissions.values_list('code', flat=True))
    new_codes = sorted(set(request.POST.getlist('permission_codes')))

    role.permissions.set(Permission.objects.filter(code__in=new_codes))

    if old_codes != new_codes:
        AccessControlAuditLog.objects.create(
            changed_by=request.user, action='update', target_type='role',
            target_label=role.display_name,
            old_values={'permissions': old_codes},
            new_values={'permissions': new_codes},
        )
    messages.success(request, f"Permissions updated for role '{role.display_name}'.")
    return redirect(f"/settings/?role_id={role.id}#access-control")


@require_role('admin')
@require_POST
def access_user_overrides_edit(request):
    from django.contrib.auth import get_user_model
    from core.models import Role, Permission, UserPermissionOverride, AccessControlAuditLog

    User = get_user_model()
    target_user = get_object_or_404(User, id=request.POST.get('user_id'))

    profile = getattr(target_user, 'profile', None)
    role = Role.objects.filter(slug=(getattr(profile, 'role', '') or '').strip().lower()).first()
    role_codes = set(role.permissions.values_list('code', flat=True)) if role else set()

    checked_codes = set(request.POST.getlist('permission_codes'))
    existing_overrides = {
        o.permission.code: o
        for o in UserPermissionOverride.objects.filter(user=target_user).select_related('permission')
    }

    old_snapshot = {code: o.granted for code, o in existing_overrides.items()}
    new_snapshot = {}

    for perm in Permission.objects.filter(is_active=True):
        checked = perm.code in checked_codes
        in_role = perm.code in role_codes
        existing = existing_overrides.get(perm.code)

        if checked == in_role:
            if existing is not None:
                existing.delete()
            continue

        granted = checked
        if existing is None or existing.granted != granted:
            UserPermissionOverride.objects.update_or_create(
                user=target_user, permission=perm,
                defaults={'granted': granted, 'created_by': request.user},
            )
        new_snapshot[perm.code] = granted

    if old_snapshot != new_snapshot or new_snapshot:
        AccessControlAuditLog.objects.create(
            changed_by=request.user, action='update', target_type='user_override',
            target_label=target_user.username,
            old_values=old_snapshot,
            new_values=new_snapshot,
        )
    messages.success(request, f"Access overrides updated for {target_user.username}.")
    return redirect(f"/settings/?user_id={target_user.id}#access-control")


@login_required
@require_POST
def access_user_role_update(request):
    """Superuser-only: change an existing user's role from the Roles &
    Access Control user list (separate from manage_user_roles, which is the
    older admin-role-gated screen)."""
    if not request.user.is_superuser:
        add_unique_message(request, messages.ERROR, '❌ Only a superuser can change user roles.')
        return redirect('notification_settings_home')

    from django.contrib.auth import get_user_model
    from core.models import AccessControlAuditLog

    User = get_user_model()
    target_user = get_object_or_404(User, id=request.POST.get('user_id'))
    new_role = (request.POST.get('role') or '').strip()
    if not new_role:
        messages.error(request, 'A role is required.')
        return redirect('/settings/#access-control')

    profile = target_user.profile
    old_role = profile.role
    if old_role != new_role:
        profile.role = new_role
        profile.save()
        AccessControlAuditLog.objects.create(
            changed_by=request.user, action='update', target_type='user_role',
            target_label=target_user.username,
            old_values={'role': old_role},
            new_values={'role': new_role},
        )
        messages.success(request, f"{target_user.username}'s role updated to {profile.get_role_display()}.")
    return redirect('/settings/#access-control')


@login_required
@require_POST
def access_user_password_reset(request):
    """Superuser-only: send a password-reset link to an existing user's
    email, reusing the same self-service reset flow/templates."""
    if not request.user.is_superuser:
        add_unique_message(request, messages.ERROR, '❌ Only a superuser can trigger password resets.')
        return redirect('notification_settings_home')

    from django.contrib.auth import get_user_model
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.encoding import force_bytes
    from django.utils.http import urlsafe_base64_encode
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    from django.urls import reverse

    User = get_user_model()
    target_user = get_object_or_404(User, id=request.POST.get('user_id'))

    if not target_user.email:
        messages.error(request, f"{target_user.username} has no email on file — add one before sending a reset link.")
        return redirect('/settings/#access-control')

    # Build the link manually (rather than PasswordResetForm) because that
    # form silently skips accounts with an unusable password — exactly the
    # state a brand-new user is in before they've ever set one.
    uid = urlsafe_base64_encode(force_bytes(target_user.pk))
    token = default_token_generator.make_token(target_user)
    reset_link = request.build_absolute_uri(
        reverse('password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
    )
    email_body = render_to_string('user_welcome_email.html', {
        'user': target_user, 'reset_link': reset_link,
    })
    send_mail(
        subject='Offset ERP - Set/Reset Your Password',
        message=email_body,
        from_email=None,
        recipient_list=[target_user.email],
        fail_silently=True,
    )
    messages.success(request, f"Password reset link sent to {target_user.email}.")
    return redirect('/settings/#access-control')


@login_required
@require_POST
def access_user_toggle_active(request):
    """Superuser-only: activate/deactivate a login. Deactivated users can't
    log in but their records/history are kept intact."""
    if not request.user.is_superuser:
        add_unique_message(request, messages.ERROR, '❌ Only a superuser can activate/deactivate accounts.')
        return redirect('notification_settings_home')

    from django.contrib.auth import get_user_model
    from core.models import AccessControlAuditLog

    User = get_user_model()
    target_user = get_object_or_404(User, id=request.POST.get('user_id'))

    if target_user.id == request.user.id:
        messages.error(request, "You can't deactivate your own account.")
        return redirect('/settings/#access-control')

    target_user.is_active = not target_user.is_active
    target_user.save(update_fields=['is_active'])
    AccessControlAuditLog.objects.create(
        changed_by=request.user, action='update', target_type='user_status',
        target_label=target_user.username,
        old_values={'is_active': not target_user.is_active},
        new_values={'is_active': target_user.is_active},
    )
    messages.success(request, f"{target_user.username} is now {'active' if target_user.is_active else 'deactivated'}.")
    return redirect('/settings/#access-control')


@login_required
@require_POST
def notification_rule_add(request):
    """
    Add a new rule dynamically from the settings page.
    """
    if request.user.profile.role not in ('admin', 'manager'):
        messages.error(request, "Not authorized.")
        return redirect('home')

    from core.models import NotificationEvent, NotificationRule, Department
    from django.contrib.auth import get_user_model
    from core.notifications import log_rule_change
    User = get_user_model()

    event_id = request.POST.get('event')
    recipient_type = request.POST.get('recipient_type')
    enabled = request.POST.get('enabled') == 'on'
    exclude_actor = request.POST.get('exclude_actor') == 'on'
    priority = request.POST.get('priority', 'medium')
    in_app_enabled = request.POST.get('in_app_enabled') == 'on'

    role = request.POST.get('role')
    user_id = request.POST.get('user')
    department_id = request.POST.get('department')

    send_to_creator = request.POST.get('send_to_creator') == 'on'
    send_to_manager = request.POST.get('send_to_manager') == 'on'
    send_to_supervisor = request.POST.get('send_to_supervisor') == 'on'
    send_to_next_stage = request.POST.get('send_to_next_stage') == 'on'

    try:
        event = NotificationEvent.objects.get(id=int(event_id))
    except (TypeError, ValueError, NotificationEvent.DoesNotExist):
        messages.error(request, "Invalid event selected.")
        return redirect('notification_settings_home')

    rule = NotificationRule(
        event=event,
        enabled=enabled,
        recipient_type=recipient_type,
        exclude_actor=exclude_actor,
        priority=priority,
        in_app_enabled=in_app_enabled,
        send_to_creator=send_to_creator,
        send_to_manager=send_to_manager,
        send_to_supervisor=send_to_supervisor,
        send_to_next_stage=send_to_next_stage,
    )

    if recipient_type == 'role':
        rule.role = role
    elif recipient_type == 'user' and user_id:
        try:
            rule.user = User.objects.get(id=int(user_id))
        except (ValueError, User.DoesNotExist):
            pass
    elif recipient_type == 'department' and department_id:
        try:
            rule.department = Department.objects.get(id=int(department_id))
        except (ValueError, Department.DoesNotExist):
            pass

    rule.save()
    log_rule_change(request.user, rule, 'create')
    messages.success(request, f"New notification rule created for event: {event.name}.")
    return redirect('notification_settings_home')


@login_required
@require_POST
def notification_rule_delete(request, rule_id):
    """
    Remove a notification rule and log the change.
    """
    if request.user.profile.role not in ('admin', 'manager'):
        messages.error(request, "Not authorized.")
        return redirect('home')

    from core.models import NotificationRule
    from core.notifications import log_rule_change

    rule = get_object_or_404(NotificationRule, id=rule_id)
    log_rule_change(request.user, rule, 'delete')
    rule.delete()
    messages.success(request, "Notification rule deleted successfully.")
    return redirect('notification_settings_home')


@login_required
@require_POST
def workflow_transition_add(request):
    """
    Add a workflow transition rule from settings.
    """
    if request.user.profile.role not in ('admin', 'manager'):
        messages.error(request, "Not authorized.")
        return redirect('home')

    from core.models import WorkflowTransition

    module = request.POST.get('module', '').strip()
    current_stage = request.POST.get('current_stage', '').strip()
    action = request.POST.get('action', '').strip()
    next_stage = request.POST.get('next_stage', '').strip()
    notify_role = request.POST.get('notify_role', '').strip()

    if not (module and current_stage and action and next_stage):
        messages.error(request, "Please fill all workflow transition fields.")
        return redirect('notification_settings_home')

    try:
        WorkflowTransition.objects.update_or_create(
            module=module,
            current_stage=current_stage,
            action=action,
            next_stage=next_stage,
            defaults={'notify_role': notify_role}
        )
        messages.success(request, f"Workflow transition for {module} added successfully.")
    except Exception as e:
        messages.error(request, f"Error creating workflow transition: {e}")

    return redirect('notification_settings_home')


@login_required
@require_POST
def workflow_transition_delete(request, transition_id):
    """
    Delete a workflow transition rule.
    """
    if request.user.profile.role not in ('admin', 'manager'):
        messages.error(request, "Not authorized.")
        return redirect('home')

    from core.models import WorkflowTransition

    transition = get_object_or_404(WorkflowTransition, id=transition_id)
    transition.delete()
    messages.success(request, "Workflow transition configuration deleted.")
    return redirect('notification_settings_home')