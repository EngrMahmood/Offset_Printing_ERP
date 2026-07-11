from datetime import date

from django.contrib.auth import get_user_model
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from django.urls import reverse

from core.models import UserProfile
from manual_working.services import PAGE_SIZE, build_manual_working_rows, get_manual_working_queryset
from planning.models import PlanningJob


class ManualWorkingPerformanceTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username='manual_board', password='pass')
        UserProfile.objects.get_or_create(user=self.user)
        self.client.force_login(self.user)

        for i in range(PAGE_SIZE + 8):
            PlanningJob.objects.create(
                jc_number=f'JC-MW-{i:03d}',
                po_number=f'PO-MW-{i:03d}',
                sku=f'SKU-MW-{i:03d}',
                job_name=f'Job {i}',
                status='qc_approved',
                planning_stage='jc_ready',
                plan_month='July',
                po_approval_date=date(2026, 7, 1),
                is_active=True,
            )

        PlanningJob.objects.create(
            jc_number='JC-MW-ARCHIVED',
            po_number='PO-MW-ARCHIVED',
            sku='SKU-MW-ARCHIVED',
            status='draft',
            is_active=False,
        )

    def test_defaults_to_active_jobs_only(self):
        qs = get_manual_working_queryset({})
        self.assertFalse(qs.filter(is_active=False).exists())
        self.assertFalse(qs.filter(jc_number='JC-MW-ARCHIVED').exists())

    def test_server_pagination(self):
        response = self.client.get(reverse('manual_working:manual_working_list'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['rows']), PAGE_SIZE)
        self.assertTrue(response.context['page_obj'].has_next())
        self.assertEqual(response.context['page_obj'].paginator.count, PAGE_SIZE + 8)

    def test_pagination_preserves_filters(self):
        response = self.client.get(
            reverse('manual_working:manual_working_list'),
            {'sku': 'SKU-MW-001', 'page': '1'},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn('sku=SKU-MW-001', response.context['filter_query'])
        self.assertEqual(len(response.context['rows']), 1)
        self.assertEqual(response.context['rows'][0]['sku'], 'SKU-MW-001')

    def test_po_approval_uses_job_field_without_extra_queries(self):
        job = PlanningJob.objects.get(jc_number='JC-MW-000')
        jobs = list(
            PlanningJob.objects.filter(pk=job.pk)
            .select_related('job_card', 'job_card__production_wip_status__status')
            .prefetch_related('po_documents', 'print_runs', 'dispatch_runs')
        )
        # Warm prefetch caches, then rebuild rows and ensure no PoDocument SELECT.
        with CaptureQueriesContext(connection) as ctx:
            rows = build_manual_working_rows(jobs)
        self.assertEqual(rows[0]['date'], '01-07-2026')
        podoc_queries = [q['sql'] for q in ctx.captured_queries if 'planning_podocument' in q['sql'].lower() or 'po_document' in q['sql'].lower()]
        self.assertEqual(podoc_queries, [])

    def test_list_query_count_stays_bounded(self):
        with CaptureQueriesContext(connection) as ctx:
            response = self.client.get(reverse('manual_working:manual_working_list'))
        self.assertEqual(response.status_code, 200)
        self.assertLess(len(ctx), 40)

    def test_export_xlsx_includes_all_filtered_rows(self):
        from io import BytesIO

        from openpyxl import load_workbook

        response = self.client.get(
            reverse('manual_working:manual_working_list'),
            {'export': 'xlsx'},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn('manual-working.xlsx', response['Content-Disposition'])
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        # Title, generated_at, blank, headers, then data rows
        data_rows = list(sheet.iter_rows(min_row=5, values_only=True))
        self.assertEqual(len(data_rows), PAGE_SIZE + 8)

    def test_export_respects_sku_filter(self):
        from io import BytesIO

        from openpyxl import load_workbook

        response = self.client.get(
            reverse('manual_working:manual_working_list'),
            {'sku': 'SKU-MW-001', 'export': 'xlsx'},
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        data_rows = list(sheet.iter_rows(min_row=5, values_only=True))
        self.assertEqual(len(data_rows), 1)
