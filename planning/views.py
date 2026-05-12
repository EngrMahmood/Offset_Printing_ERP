import csv
import logging
import base64
import io
import json
import re
from difflib import SequenceMatcher
from datetime import datetime


def _build_qr_image_base64(data):
    try:
        import qrcode
    except ImportError:
        return None

    buffer = io.BytesIO()
    qr = qrcode.QRCode(border=1, box_size=3)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    img.save(buffer, format='PNG')
    return base64.b64encode(buffer.getvalue()).decode('utf-8')
from decimal import Decimal, InvalidOperation
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required

logger = logging.getLogger(__name__)
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db import OperationalError, transaction
from django.db.models import Count, Q, Sum
from django.http import Http404, HttpResponse
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from core.jc_numbering import allocate_next_jc_number
from core.models import ChangeLog
from core.views import permission_required
from .forms import JobCardLayoutForm, PlanningJobEditForm, PlanningJobFinalizationForm, SkuRecipeForm
from .services import (
    _user_is_admin, _planning_status_filter_values, _parse_date_filter,
    _build_job_card_pdf_bytes, _sku_key, _missing_required_master_fields,
    _sync_new_sku_requirement, _build_recipe_map, _to_optional_positive_int,
    _to_optional_decimal, _sanitize_po_payload_items, _po_payload_items,
    _annotate_items_with_recipe, _deduplicate_po_items_by_sku,
    _history_repeat_new_counts, _sync_repeat_jobs_from_po,
    _sync_new_jobs_for_approved_sku, _merge_po_items_for_existing_po,
    _collect_pending_sku_rows
)
from .models import (
    PLANNING_QC_GATE_STATUSES,
    PLANNING_STATUS_ALIASES,
    PLANNING_STATUS_CHOICES,
    JobCardLayout,
    PlanningDispatchRun,
    PlanningJob,
    PlanningPrintRun,
    PoDocument,
    SkuRecipe,
)
from workflow.services import (
    _annotate_items_with_recipe,
    _build_cost_mismatch_note,
    _build_recipe_map,
    _collect_pending_sku_rows,
    _format_decimal_string,
    _format_display_qty,
    _missing_required_master_fields,
    _normalize_application_input,
    _normalize_color_spec_input,
    _normalize_status,
    _parse_iso_date,
    _po_payload_items,
    _sanitize_po_payload_items,
    _sku_key,
    sync_job_card_for_planning_status,
    _sync_new_jobs_for_approved_sku,
    _to_decimal,
    _to_int,
    _to_optional_decimal,
    _to_optional_positive_int,
    _user_is_admin,
)
from .po_extractor import extract_po_from_pdf


PLANNING_STATUSES = PLANNING_STATUS_CHOICES
PLANNING_STATUS_SET = {value for value, _ in PLANNING_STATUSES}
PLANNING_ACTIVE_STATUS_SET = {'released', 'in_production', 'completed', 'closed'}
PLANNING_QUEUE_STATUS_SET = {'draft', 'pending_qc', 'qc_approved'}
PLANNING_STATUS_LABELS = dict(PLANNING_STATUSES)
PLANNING_STATUS_FILTER_ALIASES = {
    'draft': {'draft', 'open', 'pending'},
    'pending_qc': {'pending_qc', 'reviewed'},
    'qc_approved': {'qc_approved', 'approved'},
    'released': {'released'},
    'in_production': {'in_production'},
    'completed': {'completed', 'closed'},
}
NEW_SKU_REQUIREMENT_NOTE = 'NEW SKU: Shade matching and setup verification required before production run.'
COST_MISMATCH_NOTE_PREFIX = 'COST ALERT:'
SKU_MASTER_APPROVAL_REQUIRED_FIELDS = [
    ('job_name', 'Job Name'),
    ('material', 'Material'),
    ('color_spec', 'Color'),
    ('application', 'Application'),
    ('print_sheet_size', 'Print Sheet'),
    ('purchase_sheet_size', 'Purchase Sheet'),
    ('ups', 'UPS'),
]

_COLOR_PLUS_RE = re.compile(r'^(\d+)\s*\+\s*(\d+)$')
_COLOR_SINGLE_RE = re.compile(r'^(\d+)\s*(?:colou?r(?:s)?)?$', re.IGNORECASE)


def _normalize_purchase_material_origin(raw_value):
    value = (raw_value or '').strip().lower()
    if value in {'local'}:
        return 'local'
    if value in {'import', 'imported'}:
        return 'import'
    return ''


def _get_active_job_card_layout():
    try:
        return JobCardLayout.get_active_layout()
    except OperationalError:
        return None


def _effective_planning_status(job):
    status_rank = {
        'draft': 0,
        'pending_qc': 1,
        'qc_approved': 2,
        'released': 3,
        'in_production': 4,
        'completed': 5,
    }
    planning_status = _normalize_status(job.status)
    job_card_status = None
    card_status = None
    if hasattr(job, 'job_card'):
        try:
            card_status = (job.job_card.workflow_status or '').strip().lower()
        except Exception:
            card_status = None
        if card_status in {'planning_approved', 'pending_qc'}:
            job_card_status = 'pending_qc'
        elif card_status == 'qc_approved':
            job_card_status = 'qc_approved'
        elif card_status in {'pending_pm_approval', 'production_approved', 'released', 'in_production', 'completed', 'closed'}:
            job_card_status = 'released'
        elif card_status in status_rank:
            job_card_status = card_status

    if planning_status == 'draft':
        return 'draft'
    if card_status in {'qc_rejected', 'pm_rejected'}:
        return 'draft'
    if planning_status not in status_rank:
        return job_card_status or planning_status or 'draft'
    if not job_card_status:
        return planning_status

    return planning_status if status_rank[planning_status] >= status_rank[job_card_status] else job_card_status


def _effective_planning_status_label(job, effective_status):
    if effective_status == 'released' and hasattr(job, 'job_card'):
        try:
            card_status = (job.job_card.workflow_status or '').strip().lower()
        except Exception:
            card_status = ''
        if card_status == 'production_approved':
            return 'Production Approved'
        if card_status == 'pending_pm_approval':
            return 'Pending PM Approval'
        if card_status == 'qc_approved':
            return 'QC Approved'
        if card_status in {'released', 'in_production', 'completed', 'closed'}:
            return 'Released'
    return PLANNING_STATUS_LABELS.get(effective_status, effective_status.replace('_', ' ').title())


def _repair_rejected_job_status(job):
    if _normalize_status(job.status) != 'pending_qc':
        return False
    job_card = getattr(job, 'job_card', None)
    if not job_card:
        return False
    try:
        card_status = (job_card.workflow_status or '').strip().lower()
    except Exception:
        return False
    if card_status in {'draft', 'pending_data', 'qc_rejected', 'pm_rejected'}:
        job.status = 'draft'
        job.save(update_fields=['status', 'updated_at'])
        return True
    return False


def _job_card_layout_field_labels():
    return {
        'po_number': 'PO Number',
        'pr_reference': 'PR Reference',
        'order_qty': 'Order Qty',
        'po_received_date': 'PO Received',
        'delivery_date': 'Delivery Date',
        'destination': 'Delivery Location',
        'department': 'Department',
        'job_name': 'Job Name',
        'material_display': 'Material',
        'color_spec_display': 'Color',
        'number_of_colors': 'Color Count',
        'application_display': 'Application',
        'repeat_flag': 'Repeat Flag',
        'print_sheet_size_display': 'Print Sheet',
        'purchase_sheet_size_display': 'Purchase Sheet',
        'purchase_sheet_ups_display': 'Purchase Sheet UPS',
        'ups_display': 'UPS',
        'total_sheet_quantity': 'Total Sheets',
        'purchase_sheet_required_display': 'Purchase Sheet Qty',
        'actual_sheet_required_display': 'Actual Sheet Req.',
        'wastage_sheets': 'Wastage Sheets',
        'machine_name': 'Machine',
        'plate_set_no': 'Plate Set No.',
        'awc_no_display': 'AWC No.',
        'die_cutting_display': 'Die Cutting',
        'purchase_material_origin': 'Purchase Origin',
        'stock_qty': 'Stock Qty',
        'mi_quantity': 'Issued Qty',
        'mi_balance': 'Balance',
        'remaining_sheet': 'Remaining Sheet',
    }


def _format_layout_field_value(value):
    if value is None:
        return None
    if hasattr(value, 'strftime'):
        try:
            return value.strftime('%d %b %Y')
        except Exception:
            return str(value)
    return value


def _build_job_card_layout_context(job):
    active_layout = _get_active_job_card_layout()
    if not active_layout or not active_layout.layout:
        return {'layout_enabled': False}

    field_labels = _job_card_layout_field_labels()
    field_values = {
        'po_number': _format_layout_field_value(job.po_number),
        'pr_reference': _format_layout_field_value(job.pr_reference),
        'order_qty': _format_layout_field_value(job.order_qty),
        'po_received_date': _format_layout_field_value(job.po_received_date),
        'delivery_date': _format_layout_field_value(job.delivery_date),
        'destination': job.destination,
        'department': job.department,
        'job_name': job.job_name,
        'material_display': job.material_display,
        'color_spec_display': job.color_spec_display,
        'number_of_colors': job.number_of_colors,
        'application_display': job.application_display,
        'repeat_flag': job.repeat_flag,
        'print_sheet_size_display': job.print_sheet_size_display,
        'purchase_sheet_size_display': job.purchase_sheet_size_display,
        'purchase_sheet_ups_display': job.purchase_sheet_ups_display,
        'ups_display': job.ups_display,
        'total_sheet_quantity': job.total_sheet_quantity,
        'purchase_sheet_required_display': job.purchase_sheet_required_display,
        'actual_sheet_required_display': job.actual_sheet_required_display,
        'wastage_sheets': job.wastage_sheets,
        'machine_name': job.machine_name,
        'plate_set_no': job.plate_set_no,
        'awc_no_display': job.awc_no_display,
        'die_cutting_display': job.die_cutting_display,
        'purchase_material_origin': job.purchase_material_origin,
        'stock_qty': job.stock_qty,
        'mi_quantity': job.mi_quantity,
        'mi_balance': job.mi_balance,
        'remaining_sheet': job.remaining_sheet,
    }
    normalized_sections = []
    for section in active_layout.layout:
        normalized_fields = []
        for field_id in section.get('fields', []):
            normalized_fields.append({
                'id': field_id,
                'label': field_labels.get(field_id) or field_id.replace('_', ' ').title(),
                'value': field_values.get(field_id, '-') or '-',
            })
        normalized_sections.append({
            'title': section.get('title', ''),
            'fields': normalized_fields,
        })
    return {
        'layout_enabled': True,
        'layout_sections': normalized_sections,
    }


def build_planning_readme_text():
    return """Offset ERP - Planning Module Easy Guide

Last Updated: 2026-04-19

=============================
1) MASTER SKU (STEP 1)
=============================
Purpose:
- Keep approved SKU master data ready before routing PO jobs.

How to use:
- Create or bulk upload SKU master data.
- Save as Draft first.
- Move Draft -> Reviewed -> Approved.
- Only approved recipes are used as final master data.

Required fields for approval:
- Job Name, Material, Color, Application, Machine
- Print Sheet, Purchase Sheet, UPS, Purchase Material Origin

=============================
2) PO INTAKE (STEP 2)
=============================
Purpose:
- Upload PO and split lines into Repeat and New.

Routing rule:
- Repeat lines -> Planning Jobs
- New lines -> Pending SKU Master Data

Important notes:
- Duplicate SKU lines in one PO are merged.
- Qty display is normalized (trailing decimals removed).

=============================
3) PENDING NEW SKU (STEP 3)
=============================
Purpose:
- Complete missing master data for new SKU lines from PO.

Rules:
- Job Name comes from PO and is not manual.
- Department and unit cost can prefill from PO when available.
- Application must be one of: UV, Lamination Gloss, Lamination Matt, NO.
- Purchase Material Origin must be Local or Imported.

Approval path:
- Save Draft -> Send For Approval -> Approved
- Approved new SKU records refresh matching draft planning jobs.

=============================
4) PLANNING JOBS (STEP 4)
=============================
Purpose:
- Create/update, review, and manage production planning jobs.

Input source:
- Repeat jobs from PO intake
- Approved new SKUs after master approval

Operational controls:
- Filter by PO/SKU/status/department/machine/date.
- Bulk status update for selected rows.
- Open detail, edit, print A4 job card.

=============================
5) APPROVAL QUEUE (STEP 5)
=============================
Purpose:
- Release jobs through QC and Production Manager checkpoints.

Status transitions:
- Draft -> Pending Review
- Pending Review -> Pending Approval (Manager)

After approval:
- Print job card and run shop-floor execution flow.

=============================
6) SHOP FLOOR EXECUTION (STEP 6)
=============================
Purpose:
- Use QR scan and A4 card for execution traceability.

Use:
- Open job via scan.
- Track run/dispatch logs from planning-linked records.

=============================
7) CHANGE & MISMATCH ALERTS
=============================
Repeat route behavior:
- If PO cost differs from master cost, a COST ALERT note is attached.
- If PO department differs from master department, a DEPARTMENT ALERT note is attached.

New route behavior:
- New SKU records must complete master approval before planning sync.

=============================
8) DAILY DISCIPLINE
=============================
- Keep master records clean and approved.
- Route every PO through intake queue.
- Resolve pending SKUs same day.
- Complete approvals before production release.
"""


@login_required
@permission_required('can_edit_jobcard')
def planning_readme(request):
    return render(
        request,
        'planning/planning_readme.html',
        {'generated_on': timezone.now()},
    )


@login_required
@permission_required('can_edit_jobcard')
def planning_job_card_layout_builder(request):
    active_layout = _get_active_job_card_layout()
    if request.method == 'POST':
        layout_data = request.POST.get('layout_data', '[]').strip()
        name = (request.POST.get('name') or 'Job Card Layout').strip() or 'Job Card Layout'
        try:
            layout = json.loads(layout_data)
        except (ValueError, TypeError, json.JSONDecodeError):
            messages.error(request, 'Invalid layout JSON. Please save again from the layout editor.')
            return redirect('planning:job_card_layout_builder')

        try:
            if active_layout:
                active_layout.name = name
                active_layout.layout = layout
                active_layout.is_active = True
                active_layout.save(update_fields=['name', 'layout', 'is_active', 'updated_at'])
            else:
                JobCardLayout.objects.create(name=name, layout=layout, is_active=True)
            messages.success(request, 'Job card layout saved successfully.')
        except OperationalError:
            messages.error(request, 'Layout builder is not yet initialized. Run migrations before saving the layout.')
        return redirect('planning:job_card_layout_builder')

    available_fields = _job_card_layout_field_labels()
    default_sections = [
        {
            'title': 'Software Data',
            'fields': [
                'po_number', 'pr_reference', 'order_qty', 'po_received_date',
                'delivery_date', 'destination', 'department',
                'job_name', 'material_display', 'color_spec_display',
                'number_of_colors', 'application_display', 'repeat_flag',
                'print_sheet_size_display', 'purchase_sheet_size_display',
                'purchase_sheet_ups_display', 'ups_display', 'total_sheet_quantity',
                'purchase_sheet_required_display', 'actual_sheet_required_display',
                'wastage_sheets', 'machine_name', 'plate_set_no', 'awc_no_display',
                'die_cutting_display', 'purchase_material_origin', 'stock_qty',
            ],
        },
        {
            'title': 'Operator Production Entry',
            'fields': [
                'mi_quantity', 'mi_balance', 'remaining_sheet',
            ],
        },
    ]
    layout_sections = active_layout.layout if active_layout else default_sections
    # Normalize saved layout fields into objects with id and label for template rendering.
    normalized_sections = []
    for section in layout_sections:
        normalized_fields = []
        for field_item in section.get('fields', []):
            if isinstance(field_item, dict):
                field_id = field_item.get('id')
            else:
                field_id = field_item
            if not field_id:
                continue
            normalized_fields.append({
                'id': field_id,
                'label': available_fields.get(field_id, field_id),
            })
        normalized_sections.append({
            'title': section.get('title', 'Section'),
            'fields': normalized_fields,
        })
    layout_sections = normalized_sections
    return render(
        request,
        'planning/job_card_layout_builder.html',
        {
            'active_layout': active_layout,
            'available_fields': available_fields,
            'layout_sections': layout_sections,
        },
    )


@login_required
@permission_required('can_edit_jobcard')
def download_planning_readme(request):
    content = build_planning_readme_text()
    response = HttpResponse(content, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="planning_workflow_guide.txt"'
    return response


@login_required
def planning_welcome(request):
    profile = getattr(request.user, 'profile', None)
    user_role = 'unassigned'
    can_edit_jobcard = False
    can_view_reports = False
    can_manage_masters = False

    if profile is not None:
        user_role = (profile.role or 'unassigned').strip().lower()
        can_edit_jobcard = bool(profile.can_edit_jobcard())
        can_view_reports = bool(profile.can_view_reports())
        can_manage_masters = bool(profile.can_manage_masters())

    context = {
        'user_role': user_role,
        'can_edit_jobcard': can_edit_jobcard,
        'can_view_reports': can_view_reports,
        'can_manage_masters': can_manage_masters,
    }
    return render(request, 'planning/planning_welcome.html', context)


@login_required
@permission_required('can_edit_jobcard')
def planning_po_root(request):
    return redirect('planning:po_inbox')


@login_required
@permission_required('can_edit_jobcard')
def planning_pending_actions(request):
    return redirect(f"{reverse('planning:jobs')}?status=draft")


@login_required
@permission_required('can_edit_jobcard')
def planning_jobs_drafts(request):
    return redirect(f"{reverse('planning:jobs')}?status=draft")


@login_required
@permission_required('can_edit_jobcard')
def planning_jobs_locked(request):
    return redirect(f"{reverse('planning:jobs')}?status=qc_approved")


@login_required
@permission_required('can_edit_jobcard')
def planning_sku_queue(request):
    return redirect('planning:pending_skus')


@login_required
@permission_required('can_edit_jobcard')
def planning_sku_recipes_list(request):
    return redirect('qc:sku_recipes')


@login_required
def planning_home(request):
    _user_can_plan = getattr(getattr(request.user, 'profile', None), 'can_plan', lambda: False)()
    queryset = PlanningJob.objects.select_related('job_card').prefetch_related('print_runs', 'dispatch_runs').filter(
        is_active=True,
    )

    if request.method == 'POST':
        action = request.POST.get('action')
        if not _user_can_plan:
            messages.error(request, 'You do not have permission to modify planning jobs.')
            return redirect('planning:jobs')

        if action == 'bulk_update_status':
            selected_ids = []
            for raw_id in request.POST.getlist('selected_ids'):
                try:
                    selected_ids.append(int(raw_id))
                except (TypeError, ValueError):
                    continue

            target_status = _normalize_status(request.POST.get('target_status'), default='')
            if target_status not in {'released', 'in_production', 'completed'}:
                messages.error(request, 'Please select a valid target status for bulk update.')
                return redirect('planning:jobs')

            if not selected_ids:
                messages.error(request, 'Select at least one planning row for bulk update.')
                return redirect('planning:jobs')

            updated = 0
            skipped_locked = 0
            status_rank = {'released': 1, 'in_production': 2, 'completed': 3}
            for job in PlanningJob.objects.filter(id__in=selected_ids, is_active=True):
                current_status = _normalize_status(job.status)
                if status_rank.get(current_status, 0) > status_rank.get(target_status, 0):
                    skipped_locked += 1
                    continue
                if current_status == target_status:
                    continue

                job.status = target_status
                job.issued_to_production = True
                try:
                    job.save(update_fields=['status', 'issued_to_production', 'updated_at'])
                except ValidationError as exc:
                    skipped_locked += 1
                    error_dict = getattr(exc, 'message_dict', None)
                    if error_dict:
                        for field_errors in error_dict.values():
                            for error_message in field_errors:
                                messages.error(request, f'{job.jc_number}: {error_message}')
                    else:
                        messages.error(request, f'{job.jc_number}: {exc}')
                    continue
                updated += 1

            messages.success(
                request,
                f'Bulk status update complete. Updated {updated}, locked-skip {skipped_locked}.',
            )
            return redirect('planning:jobs')

        if action == 'bulk_archive':
            selected_ids = []
            for raw_id in request.POST.getlist('selected_ids'):
                try:
                    selected_ids.append(int(raw_id))
                except (TypeError, ValueError):
                    continue

            reason = (request.POST.get('archive_reason') or '').strip()
            if not selected_ids:
                messages.error(request, 'Select at least one planning row to archive.')
                return redirect('planning:jobs')

            archived_count = 0
            for job in PlanningJob.objects.filter(id__in=selected_ids, is_active=True):
                job.is_active = False
                job.archive_reason = reason
                job.archived_by = request.user
                job.archived_at = timezone.now()
                job.save(update_fields=['is_active', 'archive_reason', 'archived_by', 'archived_at', 'updated_at'])
                archived_count += 1

            messages.success(request, f'Bulk archive complete. Archived {archived_count} jobs.')
            return redirect('planning:jobs')

        if action == 'bulk_delete':
            if not _user_is_admin(request.user):
                messages.error(request, 'Only administrators can permanently delete planning jobs.')
                return redirect('planning:jobs')

            selected_ids = []
            for raw_id in request.POST.getlist('selected_ids'):
                try:
                    selected_ids.append(int(raw_id))
                except (TypeError, ValueError):
                    continue

            if not selected_ids:
                messages.error(request, 'Select at least one planning row to delete.')
                return redirect('planning:jobs')

            deleted_count = PlanningJob.objects.filter(id__in=selected_ids, is_active=True).delete()[0]
            messages.success(request, f'Bulk delete complete. Deleted {deleted_count} jobs.')
            return redirect('planning:jobs')

        if action in {'hold', 'release_hold', 'archive', 'delete'}:
            job_id = request.POST.get('job_id')
            try:
                job_id = int(job_id)
            except (TypeError, ValueError):
                messages.error(request, 'Invalid planning job selected.')
                return redirect('planning:jobs')

            job = get_object_or_404(PlanningJob, id=job_id)
            if action == 'delete' and not _user_is_admin(request.user):
                messages.error(request, 'Only administrators can delete planning jobs.')
                return redirect('planning:jobs')

            if action == 'hold':
                reason = (request.POST.get('reason') or '').strip()
                if not reason:
                    messages.error(request, 'A hold reason is required to place a job on hold.')
                    return redirect('planning:jobs')
                job.is_on_hold = True
                job.hold_reason = reason
                job.hold_by = request.user
                job.hold_at = timezone.now()
                job.save(update_fields=['is_on_hold', 'hold_reason', 'hold_by', 'hold_at', 'updated_at'])
                messages.success(request, f'Planning job {job.jc_number} was placed on hold.')
                return redirect('planning:jobs')

            if action == 'release_hold':
                job.is_on_hold = False
                job.hold_reason = ''
                job.hold_by = None
                job.hold_at = None
                job.save(update_fields=['is_on_hold', 'hold_reason', 'hold_by', 'hold_at', 'updated_at'])
                messages.success(request, f'Planning job {job.jc_number} hold was released.')
                return redirect('planning:jobs')

            if action == 'archive':
                reason = (request.POST.get('reason') or '').strip()
                job.is_active = False
                job.archive_reason = reason
                job.archived_by = request.user
                job.archived_at = timezone.now()
                job.save(update_fields=['is_active', 'archive_reason', 'archived_by', 'archived_at', 'updated_at'])
                messages.success(request, f'Planning job {job.jc_number} was archived.')
                return redirect('planning:jobs')

            if action == 'delete':
                job.delete()
                messages.success(request, f'Planning job {job.jc_number} was permanently deleted.')
                return redirect('planning:jobs')

    q = (request.GET.get('q') or '').strip()
    status_filter = _normalize_status(request.GET.get('status'), default='')
    department_filter = (request.GET.get('department') or '').strip()
    machine_filter = (request.GET.get('machine') or '').strip()
    from_date = _parse_date_filter(request.GET.get('from_date'))
    to_date = _parse_date_filter(request.GET.get('to_date'))

    if q:
        queryset = queryset.filter(
            Q(jc_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(sku__icontains=q)
            | Q(job_name__icontains=q)
        )
    if status_filter:
        queryset = queryset.filter(status__in=_planning_status_filter_values(status_filter))
    if department_filter:
        queryset = queryset.filter(department__icontains=department_filter)
    if machine_filter:
        queryset = queryset.filter(machine_name__icontains=machine_filter)
    if from_date:
        queryset = queryset.filter(plan_date__gte=from_date)
    if to_date:
        queryset = queryset.filter(plan_date__lte=to_date)

    status_rows = queryset.values('status').annotate(total=Count('id')).order_by('status')
    status_counts = {}
    for row in status_rows:
        normalized_status = _normalize_status(row['status'])
        status_counts[normalized_status] = status_counts.get(normalized_status, 0) + (row['total'] or 0)

    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    jobs = paginator.get_page(page_number)
    approved_sku_keys = {
        _sku_key(sku)
        for sku in SkuRecipe.objects.filter(
            is_active=True,
            master_data_status='approved',
        ).values_list('sku', flat=True)
        if sku
    }
    for job in jobs:
        _repair_rejected_job_status(job)
        job.effective_status = _effective_planning_status(job)
        job.effective_status_label = _effective_planning_status_label(job, job.effective_status)
        job.can_submit_qc = True
        job.submit_qc_block_reason = ''
        if job.effective_status == 'draft':
            has_approved_recipe = _sku_key(job.sku) in approved_sku_keys
            if not has_approved_recipe:
                job.can_submit_qc = False
                job.submit_qc_block_reason = 'SKU master is pending review/approval in QC.'

    # --- Dashboard section counts (display only, no logic change) ---
    po_docs_qs = PoDocument.objects.all()
    po_queue_count = po_docs_qs.count()
    po_pending_review_count = po_docs_qs.filter(extraction_status='pending').count()
    po_processed_count = po_docs_qs.filter(extraction_status='processed').count()
    po_rejected_count = po_docs_qs.filter(extraction_status='failed').count()

    po_docs_with_payload = po_docs_qs.exclude(extracted_payload__isnull=True)
    sku_pending_count = len(_collect_pending_sku_rows(po_docs_with_payload))
    sku_pending_master_entries_count = SkuRecipe.objects.filter(
        is_active=True,
        master_data_status__in=['draft', 'pending_review'],
    ).count()
    sku_approved_count = SkuRecipe.objects.filter(is_active=True, master_data_status='approved').count()

    ignored_sku_keys = set()
    for payload in po_docs_with_payload.values_list('extracted_payload', flat=True):
        payload = payload or {}
        ignored_values = payload.get('new_skus_ignored') or []
        for raw_sku in ignored_values:
            sku_key = _sku_key(raw_sku)
            if sku_key:
                ignored_sku_keys.add(sku_key)
    sku_ignored_count = len(ignored_sku_keys)

    planning_status_counts = {}
    for job in PlanningJob.objects.select_related('job_card').filter(is_active=True):
        status_key = _effective_planning_status(job)
        planning_status_counts[status_key] = planning_status_counts.get(status_key, 0) + 1

    planning_draft_count = planning_status_counts.get('draft', 0)
    planning_pending_qc_count = planning_status_counts.get('pending_qc', 0)
    planning_approved_count = planning_status_counts.get('qc_approved', 0)
    planning_released_count = planning_status_counts.get('released', 0)

    return render(
        request,
        'planning/planning_home.html',
        {
            'jobs': jobs,
            'status_counts': status_counts,
            'status_choices': [
                (value, label)
                for value, label in PLANNING_STATUSES
            ],
            'can_admin_actions': _user_is_admin(request.user),
            'user_can_plan': _user_can_plan,
            'filters': {
                'q': q,
                'status': status_filter,
                'department': department_filter,
                'machine': machine_filter,
                'from_date': request.GET.get('from_date', ''),
                'to_date': request.GET.get('to_date', ''),
            },
            # PO Intake section (required names)
            'po_queue_count': po_queue_count,
            'po_pending_review_count': po_pending_review_count,
            'po_processed_count': po_processed_count,
            'po_rejected_count': po_rejected_count,
            # SKU Queue section (required names)
            'sku_pending_count': sku_pending_count,
            'sku_pending_master_entries_count': sku_pending_master_entries_count,
            'sku_approved_count': sku_approved_count,
            'sku_ignored_count': sku_ignored_count,
            # Planning Queue section (required names)
            'planning_draft_count': planning_draft_count,
            'planning_pending_qc_count': planning_pending_qc_count,
            'planning_approved_count': planning_approved_count,
            'planning_released_count': planning_released_count,
        },
    )


@login_required
@permission_required('can_edit_jobcard')
def planning_jobs_archived(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action in {'bulk_restore', 'bulk_delete'}:
            selected_ids = []
            for raw_id in request.POST.getlist('selected_ids'):
                try:
                    selected_ids.append(int(raw_id))
                except (TypeError, ValueError):
                    continue

            if not selected_ids:
                messages.error(request, 'Select at least one archived planning job.')
                return redirect('planning:jobs_archived')

            if action == 'bulk_restore':
                reason = (request.POST.get('reason') or '').strip()
                restored_count = 0
                for job in PlanningJob.objects.filter(id__in=selected_ids, is_active=False):
                    job.is_active = True
                    job.restored_by = request.user
                    job.restored_at = timezone.now()
                    job.restore_reason = reason
                    job.save(update_fields=['is_active', 'restored_by', 'restored_at', 'restore_reason', 'updated_at'])
                    restored_count += 1
                messages.success(request, f'Bulk restore complete. Restored {restored_count} jobs.')
                return redirect('planning:jobs_archived')

            if action == 'bulk_delete':
                if not _user_is_admin(request.user):
                    messages.error(request, 'Only administrators can permanently delete archived planning jobs.')
                    return redirect('planning:jobs_archived')
                deleted_count = PlanningJob.objects.filter(id__in=selected_ids, is_active=False).delete()[0]
                messages.success(request, f'Bulk delete complete. Deleted {deleted_count} jobs.')
                return redirect('planning:jobs_archived')

        job_id = request.POST.get('job_id')
        try:
            job_id = int(job_id)
        except (TypeError, ValueError):
            messages.error(request, 'Invalid archived planning job selected.')
            return redirect('planning:jobs_archived')

        job = get_object_or_404(PlanningJob, id=job_id, is_active=False)

        if action == 'restore':
            reason = (request.POST.get('reason') or '').strip()
            job.is_active = True
            job.restored_by = request.user
            job.restored_at = timezone.now()
            job.restore_reason = reason
            job.save(update_fields=['is_active', 'restored_by', 'restored_at', 'restore_reason', 'updated_at'])
            messages.success(request, f'Planning job {job.jc_number} was restored from archive.')
            return redirect('planning:jobs_archived')

        if action == 'delete':
            if not _user_is_admin(request.user):
                messages.error(request, 'Only administrators can permanently delete archived planning jobs.')
                return redirect('planning:jobs_archived')
            job.delete()
            messages.success(request, f'Planning job {job.jc_number} was permanently deleted.')
            return redirect('planning:jobs_archived')

        messages.error(request, 'Unknown action for archived planning jobs.')
        return redirect('planning:jobs_archived')

    queryset = PlanningJob.objects.prefetch_related('print_runs', 'dispatch_runs').filter(is_active=False)

    q = (request.GET.get('q') or '').strip()
    status_filter = _normalize_status(request.GET.get('status'), default='')
    department_filter = (request.GET.get('department') or '').strip()
    machine_filter = (request.GET.get('machine') or '').strip()
    from_date = _parse_date_filter(request.GET.get('from_date'))
    to_date = _parse_date_filter(request.GET.get('to_date'))

    if q:
        queryset = queryset.filter(
            Q(jc_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(sku__icontains=q)
            | Q(job_name__icontains=q)
        )
    if status_filter:
        queryset = queryset.filter(status__in=_planning_status_filter_values(status_filter))
    if department_filter:
        queryset = queryset.filter(department__icontains=department_filter)
    if machine_filter:
        queryset = queryset.filter(machine_name__icontains=machine_filter)
    if from_date:
        queryset = queryset.filter(plan_date__gte=from_date)
    if to_date:
        queryset = queryset.filter(plan_date__lte=to_date)

    status_rows = queryset.values('status').annotate(total=Count('id')).order_by('status')
    status_counts = {}
    for row in status_rows:
        normalized_status = _normalize_status(row['status'])
        status_counts[normalized_status] = status_counts.get(normalized_status, 0) + (row['total'] or 0)

    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    jobs = paginator.get_page(page_number)
    return render(
        request,
        'planning/planning_archived_jobs.html',
        {
            'jobs': jobs,
            'status_counts': status_counts,
            'status_choices': PLANNING_STATUSES,
            'can_admin_actions': _user_is_admin(request.user),
            'filters': {
                'q': q,
                'status': status_filter,
                'department': department_filter,
                'machine': machine_filter,
                'from_date': request.GET.get('from_date', ''),
                'to_date': request.GET.get('to_date', ''),
            },
        },
    )


@login_required
@permission_required('can_edit_jobcard')
@transaction.atomic
def import_planning_sheet(request):
    if request.method == 'POST':
        upload = request.FILES.get('sheet_file')
        if not upload:
            messages.error(request, 'Please choose a CSV file first.')
            return redirect('planning:import_sheet')

        if not upload.name.lower().endswith('.csv'):
            messages.error(request, 'Only CSV file is supported in this first import phase.')
            return redirect('planning:import_sheet')

        decoded = upload.read().decode('utf-8-sig', errors='ignore')
        rows = list(csv.reader(io.StringIO(decoded)))
        header_index = None
        for idx, candidate in enumerate(rows[:8]):
            normalized = {str(col).strip().lower() for col in candidate}
            if 'jc' in normalized and 'job name' in normalized:
                header_index = idx
                break

        if header_index is None:
            messages.error(request, 'Could not detect a valid header row (expected JC and Job Name columns).')
            return redirect('planning:import_sheet')

        header = rows[header_index]
        data_rows = rows[header_index + 1 :]

        imported_count = 0
        updated_count = 0

        for raw_row in data_rows:
            row = {
                header[i]: raw_row[i] if i < len(raw_row) else ''
                for i in range(len(header))
            }
            jc_number = (row.get('JC') or '').strip()
            if not jc_number:
                continue

            defaults = {
                'plan_month': (row.get('Month') or '').strip(),
                'plan_date': _to_date(row.get('Date')),
                'delivery_date': _to_date(row.get('Delivery Date')),
                'po_number': (row.get('Po') or '').strip(),
                'sku': (row.get('SKU') or '').strip(),
                'job_name': (row.get('Job Name') or '').strip(),
                'repeat_flag': (row.get('Repeat') or '').strip(),
                'material': (row.get('Material') or '').strip(),
                'color_spec': (row.get('Color') or '').strip(),
                'application': (row.get('Application') or '').strip(),
                'size_w_mm': _to_decimal(row.get('Size W mm')),
                'size_h_mm': _to_decimal(row.get('Size H mm')),
                'size_w_inch': _to_decimal(row.get('Size W Inch')),
                'size_h_inch': _to_decimal(row.get('Size H Inch')),
                'order_qty': _to_int(row.get('Order Qty')),
                'print_pcs': _to_int(row.get('Print Pcs')),
                'ups': _to_int(row.get('Ups')),
                'print_sheet_size': (row.get('Print Sheet Size') or '').strip(),
                'print_sheets': _to_int(row.get('Print Sheets')),
                'wastage_sheets': _to_int(row.get('Wastage')),
                'actual_sheet_required': _to_int(row.get('Actual Sheet require')),
                'purchase_sheet_size': (row.get('Purchase Sheet Size') or '').strip(),
                'purchase_sheet_ups': _to_int(row.get('Purchase Sheet ups')),
                'purchase_sheet_required': _to_int(row.get('Purchase Sheet require')),
                'pkt_value': _to_decimal(row.get('PKT')),
                'remarks': (row.get('Remarks  ') or '').strip(),
                'requirement': (row.get('Requirement') or '').strip(),
                'front_colors': _to_int(row.get('No. of Clrs Front')),
                'back_colors': _to_int(row.get('No. Of Clrs Back')),
                'total_colors': _to_int(row.get('Total Crls')),
                'total_mr_time_minutes': _to_int(row.get('Total M/R Time (15m/clr)')),
                'front_pass': _to_int(row.get('Front Pass')),
                'back_pass': _to_int(row.get('Back Pass')),
                'planned_total_impressions': _to_int(row.get('Total Impressions')),
                'mi_quantity': _to_int(row.get('MI Quantity 5')),
                'mi_balance': _to_int(row.get('MI Balance')),
                'remaining_sheet': _to_int(row.get('Remaining sheet')),
                'status': (row.get('status') or '').strip(),
                'pr_reference': (row.get('PR') or '').strip(),
                'rejected_qty': _to_int(row.get('Rejected')),
                'balance_qty': _to_int(row.get('Balance')),
                'destination': (row.get('Destination') or '').strip(),
                'unit_cost': _to_decimal(row.get('Cost')),
                'stock_bag': _to_decimal(row.get('Stock Bag')),
                'machine_name': (row.get('Machine Name') or '').strip(),
                'purchase_material_origin': _normalize_purchase_material_origin((row.get('Purchase Material') or '').strip()),
                'stock_qty': _to_decimal(row.get('Stock')),
                'daily_demand': _to_decimal(row.get('Daily Demand')),
                'department': (row.get('Department') or '').strip(),
                'plate_set_no': (row.get('Plate Set No') or '').strip(),
                'aging_days': _to_int(row.get('Aging')),
            }

            job, created = PlanningJob.objects.update_or_create(
                jc_number=jc_number,
                defaults=defaults,
            )

            if created:
                imported_count += 1
            else:
                updated_count += 1

            job.print_runs.all().delete()
            print_rows = []
            for i in range(1, 6):
                print_date = _to_date(row.get(f'Print Date {i}'))
                print_qty = _to_int(row.get(f'Print Qty {i}'))
                wastage_qty = _to_int(row.get(f'Wastage {i}'))
                if print_date or print_qty or wastage_qty:
                    print_rows.append(
                        PlanningPrintRun(
                            planning_job=job,
                            run_index=i,
                            print_date=print_date,
                            print_qty=print_qty,
                            wastage_qty=wastage_qty,
                        )
                    )
            if print_rows:
                PlanningPrintRun.objects.bulk_create(print_rows)

            job.dispatch_runs.all().delete()
            dispatch_rows = []
            for i in range(1, 7):
                idx = f'{i:02d}'
                delivery_date = _to_date(row.get(f'Date Delivery {idx}'))
                dc_no = (row.get(f'DC {idx}') or '').strip()
                delivered_qty = _to_int(row.get(f'Delivered Quantity {idx}'))
                if delivery_date or dc_no or delivered_qty:
                    dispatch_rows.append(
                        PlanningDispatchRun(
                            planning_job=job,
                            dispatch_index=i,
                            delivery_date=delivery_date,
                            dc_no=dc_no,
                            delivered_qty=delivered_qty,
                        )
                    )
            if dispatch_rows:
                PlanningDispatchRun.objects.bulk_create(dispatch_rows)

        messages.success(
            request,
            f'Import completed. New jobs: {imported_count}, updated jobs: {updated_count}.',
        )
        return redirect('planning:jobs')

    return render(request, 'planning/planning_import.html')


@login_required
def planning_job_detail(request, job_id):
    job = get_object_or_404(
        PlanningJob.objects.prefetch_related('print_runs', 'dispatch_runs'),
        id=job_id,
    )
    is_repeat_with_changes = (
        (job.repeat_flag or '').lower() == 'repeat'
        and job.has_edits_since_creation
        and job.edited_fields_list
    )
    status_now = _normalize_status(job.status)
    job_card = getattr(job, 'job_card', None)
    can_print_from_job = status_now in {'qc_approved', 'released', 'in_production', 'completed'}
    can_print_from_card = job_card.can_print_job_card() if job_card else False

    return render(
        request,
        'planning/planning_job_detail.html',
        {
            'job': job,
            'status_now': status_now,
            'can_edit_job': status_now in {'draft', 'pending_qc'},
            'is_repeat_with_changes': is_repeat_with_changes,
            'changed_fields': job.edited_fields_list or [],
            'last_edited_by': job.last_edited_by,
            'last_edited_at': job.last_edited_at,
            'qc_missing_fields': job.qc_missing_fields(),
            'can_print_job_card': can_print_from_job or can_print_from_card,
            'can_admin_delete': _user_is_admin(request.user),
            'user_can_plan': getattr(getattr(request.user, 'profile', None), 'can_plan', lambda: False)(),
        },
    )


@login_required
@permission_required('can_plan')
def planning_job_edit(request, job_id):
    job = get_object_or_404(PlanningJob, id=job_id)
    _repair_rejected_job_status(job)
    current_status = _normalize_status(job.status)

    if current_status in {'qc_approved', 'released', 'in_production', 'completed'}:
        messages.error(request, 'QC approved and released records are locked. Reopen the job before editing.')
        return redirect('planning:job_detail', job_id=job.id)

    if request.method == 'POST':
        form = PlanningJobEditForm(request.POST, instance=job)
        if form.is_valid():
            edited = form.save(commit=False)
            edited.status = _normalize_status(edited.status)
            edited.job_card_version = (job.job_card_version or 1) + 1
            
            # Detect changes for repeat jobs
            if (job.repeat_flag or '').lower() == 'repeat':
                changed_fields = []
                edit_fields = [
                    'delivery_date', 'wastage_sheets', 'plate_set_no', 'machine_name',
                    'planned_total_impressions', 'purchase_material_origin', 'destination',
                    'remarks', 'requirement', 'status',
                ]
                for field in edit_fields:
                    old_val = getattr(job, field, None)
                    new_val = getattr(edited, field, None)
                    if str(old_val) != str(new_val):
                        changed_fields.append(field)
                
                if changed_fields:
                    edited.has_edits_since_creation = True
                    edited.edited_fields_list = changed_fields
                    edited.last_edited_by = request.user
                    edited.last_edited_at = timezone.now()
            
            try:
                edited.save()
            except ValidationError as exc:
                error_dict = getattr(exc, 'message_dict', None)
                if error_dict:
                    for field_name, field_errors in error_dict.items():
                        for error_message in field_errors:
                            if field_name == '__all__' or field_name not in form.fields:
                                form.add_error(None, error_message)
                            else:
                                form.add_error(field_name, error_message)
                else:
                    form.add_error(None, str(exc))
            else:
                messages.success(request, f'Planning job {edited.jc_number} updated.')
                if edited.has_edits_since_creation and (edited.repeat_flag or '').lower() == 'repeat':
                    messages.info(request, f"Changes detected and flagged for production team: {', '.join(edited.edited_fields_list)}")
                return redirect('planning:job_detail', job_id=edited.id)
    else:
        form = PlanningJobEditForm(instance=job)

    recipe = SkuRecipe.objects.filter(
        sku__iexact=job.sku, is_active=True, master_data_status='approved',
    ).first()

    return render(
        request,
        'planning/planning_job_edit.html',
        {
            'job': job,
            'form': form,
            'recipe': recipe,
            'can_admin_delete': _user_is_admin(request.user),
        },
    )


@login_required
@permission_required('can_plan')
@transaction.atomic
def planning_job_status_update(request, job_id):
    if request.method != 'POST':
        return redirect('planning:job_detail', job_id=job_id)

    job = get_object_or_404(PlanningJob, id=job_id)
    current_status = _normalize_status(job.status)
    transition = (request.POST.get('transition') or '').strip()
    next_url = (request.POST.get('next') or '').strip()

    transitions = {
        'submit_review': ('draft', 'pending_qc'),
        'submit_qc': ('draft', 'pending_qc'),
        'reopen': ('completed', 'draft'),
    }
    if transition not in transitions:
        messages.error(request, 'Unknown status transition request.')
        return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    required_from, target_status = transitions[transition]
    if current_status == target_status:
        messages.info(request, f'Job already in {target_status} status.')
        return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    if required_from and current_status != required_from:
        messages.error(request, f'Transition not allowed from {current_status} to {target_status}.')
        return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    if target_status == 'pending_qc':
        approved_recipe = SkuRecipe.objects.filter(
            sku__iexact=job.sku,
            is_active=True,
            master_data_status='approved',
        ).first()
        if not approved_recipe:
            messages.error(request, f'SKU recipe for {job.sku or "this job"} is missing or not approved.')
            return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

        planning_required_fields = [
            ('machine_name', 'Machine Name'),
            ('wastage_sheets', 'Wastage Sheets'),
            ('plate_set_no', 'Plate Set No'),
            ('purchase_material_origin', 'Purchase Material Origin'),
        ]
        missing_planning_fields = []
        for field_name, field_label in planning_required_fields:
            value = getattr(job, field_name, None)
            if value is None:
                missing_planning_fields.append(field_label)
                continue
            if isinstance(value, str) and not value.strip():
                missing_planning_fields.append(field_label)

        if missing_planning_fields:
            messages.error(
                request,
                f'Complete planning fields before Submit to QC: {", ".join(missing_planning_fields)}.',
            )
            return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

        qc_validation_errors = job.qc_validation_errors()
        if qc_validation_errors:
            for field_errors in qc_validation_errors.values():
                for error_message in field_errors if isinstance(field_errors, list) else [field_errors]:
                    messages.error(request, error_message)
            return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    job.status = target_status
    if target_status == 'pending_qc':
        job.issued_to_production = False
    if target_status == 'draft':
        job.issued_to_production = False

    try:
        job.save(update_fields=['status', 'issued_to_production', 'updated_at'])
        if target_status == 'pending_qc':
            try:
                sync_job_card_for_planning_status(job, target_status, request.user)
            except ValidationError as exc:
                transaction.set_rollback(True)
                error_dict = getattr(exc, 'message_dict', None)
                if error_dict:
                    for field_errors in error_dict.values():
                        for error_message in field_errors:
                            messages.error(request, error_message)
                else:
                    messages.error(request, str(exc))
                return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)
    except ValidationError as exc:
        error_dict = getattr(exc, 'message_dict', None)
        if error_dict:
            for field_errors in error_dict.values():
                for error_message in field_errors:
                    messages.error(request, error_message)
        else:
            messages.error(request, str(exc))
        return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    messages.success(request, f'Job status updated: {current_status} -> {target_status}.')
    return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)


@login_required
@permission_required('can_edit_jobcard')
def planning_job_card_print(request, job_id):
    job = get_object_or_404(
        PlanningJob.objects.prefetch_related('print_runs', 'dispatch_runs'),
        id=job_id,
    )
    status_now = _normalize_status(job.status)
    job_card = getattr(job, 'job_card', None)
    can_print_from_job = status_now in {'qc_approved', 'released', 'in_production', 'completed'}
    can_print_from_card = job_card.can_print_job_card() if job_card else False
    if not (can_print_from_job or can_print_from_card):
        messages.error(request, 'Job card print is available only after QC approval or production approval.')
        return redirect('planning:job_detail', job_id=job.id)

    missing_qc_fields = job.qc_missing_fields()
    if missing_qc_fields:
        messages.error(request, f'Job card cannot be printed until QC fields are completed: {", ".join(missing_qc_fields)}.')
        return redirect('planning:job_detail', job_id=job.id)

    recipe = SkuRecipe.objects.filter(sku=job.sku).first()

    def _display_user_identity(user):
        if not user:
            return ''
        full_name = (user.get_full_name() or '').strip()
        if full_name:
            return full_name
        return (user.username or '').strip()

    def _workflow_actor_for_status(target_status):
        if not job_card:
            return ''
        logs = ChangeLog.objects.filter(
            entity_type='job_card',
            record_id=job_card.pk,
        ).select_related('changed_by').order_by('-created_at')
        for log in logs:
            field_changes = log.field_changes if isinstance(log.field_changes, dict) else {}
            status_change = field_changes.get('status') if isinstance(field_changes, dict) else None
            if not isinstance(status_change, dict):
                continue
            to_status = str(status_change.get('to') or '').strip().lower()
            if to_status == target_status and log.changed_by:
                return _display_user_identity(log.changed_by)
        return ''

    def _workflow_date_for_status(target_status):
        if not job_card:
            return None
        logs = ChangeLog.objects.filter(
            entity_type='job_card',
            record_id=job_card.pk,
        ).order_by('-created_at')
        for log in logs:
            field_changes = log.field_changes if isinstance(log.field_changes, dict) else {}
            status_change = field_changes.get('status') if isinstance(field_changes, dict) else None
            if not isinstance(status_change, dict):
                continue
            to_status = str(status_change.get('to') or '').strip().lower()
            if to_status == target_status and log.created_at:
                return log.created_at.date()
        return None

    def _mm_int_string(value):
        if value is None:
            return None
        try:
            return str(int(Decimal(str(value))))
        except Exception:
            return str(value)

    repeat_flag = (job.repeat_flag or '').strip().lower()
    if 'repeat' in repeat_flag or repeat_flag in {'r', 'old', 'existing'}:
        production_type_tag = 'REPEAT'
    else:
        production_type_tag = 'NEW'

    recipe_material = (getattr(recipe, 'material', '') or '').strip() if recipe else ''
    material_type_clean = (job.material_display or '').strip() or recipe_material or '-'
    color_spec_clean = (job.color_spec_display or '').strip() or '-'

    special_notes = []
    if recipe:
        recipe_special = (getattr(recipe, 'special_instructions', '') or '').strip()
        recipe_notes = (getattr(recipe, 'notes', '') or '').strip()
        if recipe_special:
            special_notes.append(recipe_special)
        if recipe_notes and recipe_notes not in special_notes:
            special_notes.append(recipe_notes)
    requirement_note = (job.requirement or '').strip()
    if requirement_note and requirement_note not in special_notes:
        special_notes.append(requirement_note)
    special_instructions_text = ' | '.join(special_notes) if special_notes else '-'

    po_approval_date = (
        _workflow_date_for_status('production_approved')
        or _workflow_date_for_status('qc_approved')
        or job.delivery_date
    )

    prepared_by_display = _display_user_identity(job.last_edited_by or job.created_by)
    checked_by_display = _workflow_actor_for_status('qc_approved')
    approved_by_display = _workflow_actor_for_status('production_approved')

    def _pdf_filename(jc_number):
        if not jc_number:
            return 'JOB-CARD'
        normalized = str(jc_number).strip().upper()
        parts = [part for part in normalized.split('-') if part]
        if 'UPP' in parts:
            return '-'.join(parts)
        if len(parts) == 4 and parts[0] == 'JC':
            return '-'.join([parts[0], parts[1], parts[2], 'UPP', parts[3]])
        if len(parts) >= 4 and parts[0] == 'JC' and parts[-2] != 'UPP':
            return '-'.join(parts[:-1] + ['UPP', parts[-1]])
        return normalized

    job_scan_url = request.build_absolute_uri(reverse('planning:job_detail', kwargs={'job_id': job.id}))
    qr_base64 = _build_qr_image_base64(job_scan_url)
    job_qr_data_uri = f'data:image/png;base64,{qr_base64}' if qr_base64 else None
    pdf_filename = _pdf_filename(job.jc_number)

    return render(
        request,
        'Job Card.html',
        {
            'job': job,
            'recipe': recipe,
            'job_scan_url': job_scan_url,
            'job_qr_data_uri': job_qr_data_uri,
            'production_type_tag': production_type_tag,
            'po_received_date_display': job.po_received_date,
            'po_approval_date_display': po_approval_date,
            'material_type_clean': material_type_clean,
            'color_spec_clean': color_spec_clean,
            'size_w_mm_int': _mm_int_string(job.size_w_mm_display),
            'size_h_mm_int': _mm_int_string(job.size_h_mm_display),
            'special_instructions_text': special_instructions_text,
            'prepared_by_display': prepared_by_display,
            'checked_by_display': checked_by_display,
            'approved_by_display': approved_by_display,
            'pdf_filename': pdf_filename,
        },
    )


@login_required
@permission_required('can_edit_jobcard')
def planning_job_card_pdf(request, job_id):
    job = get_object_or_404(
        PlanningJob.objects.prefetch_related('print_runs', 'dispatch_runs'),
        id=job_id,
    )
    status_now = _normalize_status(job.status)
    job_card = getattr(job, 'job_card', None)
    can_print_from_job = status_now in {'qc_approved', 'released', 'in_production', 'completed'}
    can_print_from_card = job_card.can_print_job_card() if job_card else False
    if not (can_print_from_job or can_print_from_card):
        messages.error(request, 'Job card print is available only after QC approval or production approval.')
        return redirect('planning:job_detail', job_id=job.id)

    missing_qc_fields = job.qc_missing_fields()
    if missing_qc_fields:
        messages.error(request, f'Job card cannot be downloaded until QC fields are completed: {", ".join(missing_qc_fields)}.')
        return redirect('planning:job_detail', job_id=job.id)

    def _pdf_filename(jc_number):
        if not jc_number:
            return 'JOB-CARD'
        normalized = str(jc_number).strip().upper()
        parts = [part for part in normalized.split('-') if part]
        if 'UPP' in parts:
            return '-'.join(parts)
        if len(parts) == 4 and parts[0] == 'JC':
            return '-'.join([parts[0], parts[1], parts[2], 'UPP', parts[3]])
        if len(parts) >= 4 and parts[0] == 'JC' and parts[-2] != 'UPP':
            return '-'.join(parts[:-1] + ['UPP', parts[-1]])
        return normalized

    pdf_filename = _pdf_filename(job.jc_number)
    job_scan_url = request.build_absolute_uri(reverse('planning:job_detail', kwargs={'job_id': job.id}))

    try:
        pdf_bytes = _build_job_card_pdf_bytes(job, job_scan_url)
    except RuntimeError as exc:
        messages.error(request, str(exc))
        return redirect('planning:job_card_print', job_id=job.id)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{pdf_filename}.pdf"'
    return response


@login_required
@permission_required('can_view_approval_queue')
def approval_queue(request):
    return redirect('qc:approval_queue')


@login_required
@permission_required('can_edit_jobcard')
def planning_report(request):
    queryset = PlanningJob.objects.all()

    from_date = _parse_date_filter(request.GET.get('from_date'))
    to_date = _parse_date_filter(request.GET.get('to_date'))
    if from_date:
        queryset = queryset.filter(plan_date__gte=from_date)
    if to_date:
        queryset = queryset.filter(plan_date__lte=to_date)

    totals = queryset.aggregate(
        total_jobs=Count('id'),
        total_order_qty=Sum('order_qty'),
        released_jobs=Count('id', filter=Q(status__iexact='released')),
        completed_jobs=Count('id', filter=Q(status__in=['completed', 'closed'])),
    )

    by_status_map = {}
    for row in queryset.values('status').annotate(total=Count('id'), order_qty=Sum('order_qty')).order_by('status'):
        normalized_status = _normalize_status(row['status'])
        merged_row = by_status_map.setdefault(
            normalized_status,
            {
                'status': normalized_status,
                'status_label': PLANNING_STATUS_LABELS.get(normalized_status, normalized_status.replace('_', ' ').title()),
                'total': 0,
                'order_qty': 0,
            },
        )
        merged_row['total'] += row['total'] or 0
        merged_row['order_qty'] += row['order_qty'] or 0
    status_order = {value: index for index, (value, _) in enumerate(PLANNING_STATUS_CHOICES)}
    by_status = sorted(by_status_map.values(), key=lambda item: status_order.get(item['status'], len(status_order)))
    by_department = (
        queryset.values('department')
        .annotate(total=Count('id'), order_qty=Sum('order_qty'))
        .order_by('-total', 'department')[:20]
    )
    by_machine = (
        queryset.values('machine_name')
        .annotate(total=Count('id'), order_qty=Sum('order_qty'))
        .order_by('-total', 'machine_name')[:20]
    )

    context = {
        'totals': totals,
        'by_status': by_status,
        'by_department': by_department,
        'by_machine': by_machine,
        'filters': {
            'from_date': request.GET.get('from_date', ''),
            'to_date': request.GET.get('to_date', ''),
        },
    }
    return render(request, 'planning/planning_report.html', context)


@login_required
@permission_required('can_edit_jobcard')
def planning_scan(request):
    if request.method == 'POST':
        raw_code = (request.POST.get('scan_code') or '').strip()
        if not raw_code:
            messages.error(request, 'Scan code cannot be empty.')
            return redirect('planning:scan')

        # QR may contain full URL, plain JC number, or prefixed JC field.
        parsed = raw_code
        if '/scan/open/' in parsed:
            parsed = parsed.rsplit('/scan/open/', 1)[-1]
        if '?' in parsed:
            parsed = parsed.split('?', 1)[0]
        parsed = parsed.replace('JC:', '').strip().strip('/')

        job = PlanningJob.objects.filter(jc_number__iexact=parsed).order_by('-id').first()
        if not job:
            messages.error(request, f'No planning job found for code: {parsed}')
            return redirect('planning:scan')

        return redirect('planning:job_detail', job_id=job.id)

    return render(request, 'planning/planning_scan.html')


@login_required
@permission_required('can_edit_jobcard')
def planning_scan_open(request, jc_number):
    job = PlanningJob.objects.filter(jc_number__iexact=(jc_number or '').strip()).order_by('-id').first()
    if not job:
        messages.error(request, f'No planning job found for code: {jc_number}')
        return redirect('planning:scan')
    return redirect('planning:job_detail', job_id=job.id)


@login_required
def po_debug_extract(request):
    """Debug view: upload PDF and see raw text + table rows + per-strategy parse results."""
    import json as _json
    context = {}
    if request.method == 'POST':
        pdf_file = request.FILES.get('po_pdf')
        if pdf_file:
            try:
                import pdfplumber
                full_text = ''
                table_blobs = []
                table_rows = []
                pdf_file.seek(0)
                with pdfplumber.open(pdf_file) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text(x_tolerance=3, y_tolerance=3) or ''
                        full_text += page_text + '\n'
                        for table in (page.extract_tables() or []):
                            for row in table or []:
                                parts = [str(col).strip() for col in (row or []) if str(col).strip()]
                                if parts:
                                    table_blobs.append(' '.join(parts))
                                    table_rows.append(parts)

                from .po_extractor import (
                    _build_sku_jobname_map,
                    _detect_expected_line_count,
                    _extract_items_flexible,
                    _extract_items_from_table_blobs,
                    _extract_items_from_table_rows,
                    _extract_items_from_text_windows,
                    _extract_items_strict,
                )
                sku_map = _build_sku_jobname_map(full_text, table_blobs)
                expected = _detect_expected_line_count(full_text, table_rows)
                strict = _extract_items_strict(full_text, sku_map)
                flexible = _extract_items_flexible(full_text, sku_map)
                from_rows = _extract_items_from_table_rows(table_rows, sku_map)
                from_blobs = _extract_items_from_table_blobs(table_blobs, sku_map)
                from_windows = _extract_items_from_text_windows(full_text, sku_map)

                context = {
                    'full_text': full_text,
                    'table_rows': _json.dumps(table_rows, indent=2),
                    'table_blobs': _json.dumps(table_blobs, indent=2),
                    'expected': expected,
                    'strict': _json.dumps(strict, indent=2),
                    'flexible': _json.dumps(flexible, indent=2),
                    'from_rows': _json.dumps(from_rows, indent=2),
                    'from_blobs': _json.dumps(from_blobs, indent=2),
                    'from_windows': _json.dumps(from_windows, indent=2),
                    'strict_count': len(strict),
                    'flexible_count': len(flexible),
                    'from_rows_count': len(from_rows),
                    'from_blobs_count': len(from_blobs),
                    'from_windows_count': len(from_windows),
                }
            except Exception as exc:
                context = {'error': str(exc)}
    return render(request, 'planning/po_debug.html', context)


@login_required
@permission_required('can_edit_jobcard')
def sku_recipes_list(request):
    """List all SKU recipes with search; handles delete via POST."""
    is_admin_user = _user_is_admin(request.user)
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        recipe_id = request.POST.get('recipe_id')
        redirect_url = request.path
        if request.GET:
            redirect_url += '?' + request.GET.urlencode()

        if action == 'delete':
            try:
                recipe_obj = SkuRecipe.objects.get(id=int(recipe_id))
                if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                    messages.error(request, 'Approved records can only be deleted by admin users.')
                    return redirect(redirect_url)
                recipe_obj.delete()
                messages.success(request, 'SKU Recipe deleted.')
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
            return redirect(redirect_url)

        if action == 'archive':
            try:
                recipe_obj = SkuRecipe.objects.get(id=int(recipe_id))
                if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                    messages.error(request, 'Approved records can only be archived by admin users.')
                    return redirect(redirect_url)
                recipe_obj.is_active = False
                recipe_obj.archived_by = request.user
                recipe_obj.archived_at = timezone.now()
                recipe_obj.archive_reason = (request.POST.get('archive_reason') or '').strip()
                recipe_obj.save(update_fields=['is_active', 'archived_by', 'archived_at', 'archive_reason', 'updated_at'])
                messages.success(request, 'SKU Recipe archived.')
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
            return redirect(redirect_url)

        if action in {'bulk_archive', 'bulk_delete'}:
            selected_ids = []
            for raw_id in request.POST.getlist('selected_ids'):
                try:
                    selected_ids.append(int(raw_id))
                except (TypeError, ValueError):
                    continue

            if not selected_ids:
                messages.error(request, 'Select at least one SKU recipe first.')
                return redirect(redirect_url)

            processed = 0
            skipped_locked = 0
            failures = []
            for recipe_obj in SkuRecipe.objects.filter(id__in=selected_ids, is_active=True):
                if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                    skipped_locked += 1
                    continue

                if action == 'bulk_archive':
                    recipe_obj.is_active = False
                    recipe_obj.archived_by = request.user
                    recipe_obj.archived_at = timezone.now()
                    recipe_obj.archive_reason = ''
                    recipe_obj.save(update_fields=['is_active', 'archived_by', 'archived_at', 'archive_reason', 'updated_at'])
                    processed += 1
                else:
                    try:
                        recipe_obj.delete()
                        processed += 1
                    except Exception as exc:
                        failures.append(f'{recipe_obj.sku}: {str(exc)}')

            if action == 'bulk_archive':
                messages.success(request, f'Bulk archive complete. Archived {processed}, skipped {skipped_locked}.')
            else:
                messages.success(request, f'Bulk delete complete. Deleted {processed}, skipped {skipped_locked}.')
            if failures:
                messages.error(request, 'Some items could not be processed: ' + '; '.join(failures))
            return redirect(redirect_url)

        if action in {'submit_review', 'review', 'approve', 'back_to_draft'}:
            try:
                recipe = SkuRecipe.objects.get(id=int(recipe_id))
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
                return redirect(redirect_url)

            current_status = (recipe.master_data_status or 'draft').lower()

            if action == 'submit_review':
                if current_status != 'draft':
                    messages.error(request, f'SKU {recipe.sku} can only be submitted for review from Draft.')
                    return redirect(redirect_url)
                recipe.master_data_status = 'pending_review'
                recipe.reviewed_by = None
                recipe.reviewed_at = None
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.save(update_fields=['master_data_status', 'reviewed_by', 'reviewed_at', 'approved_by', 'approved_at', 'updated_at'])
                messages.success(request, f'SKU {recipe.sku} submitted for review.')
                return redirect(redirect_url)

            if action == 'review':
                if current_status != 'pending_review':
                    messages.error(request, f'SKU {recipe.sku} can only move to Reviewed from Pending Review.')
                    return redirect(redirect_url)
                recipe.master_data_status = 'reviewed'
                recipe.reviewed_by = request.user
                recipe.reviewed_at = timezone.now()
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.save(update_fields=['master_data_status', 'reviewed_by', 'reviewed_at', 'approved_by', 'approved_at', 'updated_at'])
                messages.success(request, f'SKU {recipe.sku} moved to Reviewed.')
                return redirect(redirect_url)

            if action == 'approve':
                if current_status != 'reviewed':
                    messages.error(request, f'SKU {recipe.sku} can only be Approved from Reviewed status.')
                    return redirect(redirect_url)
                missing_required = _missing_required_master_fields(recipe, recipe.job_name)
                if missing_required:
                    messages.error(
                        request,
                        f'SKU {recipe.sku} cannot be approved. Missing required master data: {", ".join(missing_required)}.',
                    )
                    return redirect(redirect_url)
                recipe.master_data_status = 'approved'
                recipe.approved_by = request.user
                recipe.approved_at = timezone.now()
                recipe.save(update_fields=['master_data_status', 'approved_by', 'approved_at', 'updated_at'])
                messages.success(request, f'SKU {recipe.sku} approved.')
                return redirect(redirect_url)

            if action == 'back_to_draft':
                if current_status == 'draft':
                    messages.info(request, f'SKU {recipe.sku} is already in Draft.')
                    return redirect(redirect_url)
                if current_status == 'approved' and not is_admin_user:
                    messages.error(request, 'Approved records can only be reverted by admin users.')
                    return redirect(redirect_url)
                comment = (request.POST.get('rejection_comment') or '').strip()
                if not comment:
                    messages.error(request, 'Please provide a reason when sending a record back to Draft.')
                    return redirect(redirect_url)
                recipe.master_data_status = 'draft'
                recipe.reviewed_by = None
                recipe.reviewed_at = None
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.rejection_comment = comment
                recipe.last_rejected_by = request.user
                recipe.last_rejected_at = timezone.now()
                recipe.save(update_fields=[
                    'master_data_status', 'reviewed_by', 'reviewed_at',
                    'approved_by', 'approved_at', 'rejection_comment',
                    'last_rejected_by', 'last_rejected_at', 'updated_at',
                ])
                if comment:
                    messages.warning(request, f'SKU {recipe.sku} sent back to Draft. Reason: {comment}')
                else:
                    messages.success(request, f'SKU {recipe.sku} moved back to Draft.')
                return redirect(redirect_url)

    q = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    qs = SkuRecipe.objects.filter(is_active=True)
    if q:
        qs = qs.filter(
            Q(sku__icontains=q)
            | Q(job_name__icontains=q)
            | Q(material__icontains=q)
            | Q(machine_name__icontains=q)
        )
    if status_filter in ('draft', 'pending_review', 'reviewed', 'approved'):
        qs = qs.filter(master_data_status=status_filter)
    paginator = Paginator(qs, 50)
    recipes = paginator.get_page(request.GET.get('page'))

    draft_count = SkuRecipe.objects.filter(is_active=True, master_data_status='draft').count()
    pending_review_count = SkuRecipe.objects.filter(is_active=True, master_data_status='pending_review').count()
    reviewed_count = SkuRecipe.objects.filter(is_active=True, master_data_status='reviewed').count()
    approved_count = SkuRecipe.objects.filter(is_active=True, master_data_status='approved').count()

    bulk_highlights = request.session.pop('sku_recipe_bulk_highlights', {})
    for recipe in recipes:
        meta = bulk_highlights.get(str(recipe.id), {})
        recipe.bulk_highlight_type = meta.get('type', '')
        recipe.bulk_highlight_fields = meta.get('fields', [])

    return render(request, 'planning/sku_recipes.html', {
        'recipes': recipes,
        'q': q,
        'status_filter': status_filter,
        'draft_count': draft_count,
        'pending_review_count': pending_review_count,
        'reviewed_count': reviewed_count,
        'approved_count': approved_count,
        'can_edit_approved': True,
        'can_admin_actions': is_admin_user,
    })


@login_required
def sku_recipes_status(request, status=None):
    """List SKU recipes filtered by a fixed status for role-specific views."""
    if status not in {'draft', 'pending_review', 'reviewed', 'approved'}:
        raise Http404('Unknown SKU recipe status view.')
    request.GET = request.GET.copy()
    request.GET['status'] = status
    return sku_recipes_list(request)


@login_required
@permission_required('can_edit_jobcard')
def sku_recipes_archived(request):
    """List archived SKU recipes."""
    is_admin_user = _user_is_admin(request.user)
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        redirect_url = request.path
        if request.GET:
            redirect_url += '?' + request.GET.urlencode()

        if action == 'restore':
            recipe_id = request.POST.get('recipe_id')
            try:
                recipe_obj = SkuRecipe.objects.get(id=int(recipe_id), is_active=False)
                if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                    messages.error(request, 'Approved recipes can only be restored by admin users.')
                    return redirect(redirect_url)
                recipe_obj.is_active = True
                recipe_obj.save(update_fields=['is_active', 'updated_at'])
                messages.success(request, 'SKU Recipe restored to active master list.')
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
            return redirect(redirect_url)

        if action == 'delete':
            if not is_admin_user:
                messages.error(request, 'Only admin users can permanently delete archived recipes.')
                return redirect(redirect_url)

            recipe_id = request.POST.get('recipe_id')
            try:
                recipe_obj = SkuRecipe.objects.get(id=int(recipe_id), is_active=False)
                recipe_obj.delete()
                messages.success(request, 'Archived SKU Recipe permanently deleted.')
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
            return redirect(redirect_url)

        if action == 'bulk_restore':
            selected_ids = []
            for raw_id in request.POST.getlist('selected_ids'):
                try:
                    selected_ids.append(int(raw_id))
                except (TypeError, ValueError):
                    continue

            if not selected_ids:
                messages.error(request, 'Select at least one archived SKU recipe first.')
                return redirect(redirect_url)

            restored = 0
            skipped_locked = 0
            for recipe_obj in SkuRecipe.objects.filter(id__in=selected_ids, is_active=False):
                if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                    skipped_locked += 1
                    continue
                recipe_obj.is_active = True
                recipe_obj.save(update_fields=['is_active', 'updated_at'])
                restored += 1

            messages.success(request, f'Bulk restore complete. Restored {restored}, skipped {skipped_locked}.')
            return redirect(redirect_url)

        if action == 'bulk_delete':
            if not is_admin_user:
                messages.error(request, 'Only admin users can permanently delete archived recipes.')
                return redirect(redirect_url)

            selected_ids = []
            for raw_id in request.POST.getlist('selected_ids'):
                try:
                    selected_ids.append(int(raw_id))
                except (TypeError, ValueError):
                    continue

            if not selected_ids:
                messages.error(request, 'Select at least one archived SKU recipe first.')
                return redirect(redirect_url)

            deleted = 0
            failures = []
            for recipe_obj in SkuRecipe.objects.filter(id__in=selected_ids, is_active=False):
                try:
                    recipe_obj.delete()
                    deleted += 1
                except Exception as exc:
                    failures.append(f'{recipe_obj.sku}: {str(exc)}')

            messages.success(request, f'Bulk delete complete. Deleted {deleted}.')
            if failures:
                messages.error(request, 'Some items could not be deleted: ' + '; '.join(failures))
            return redirect(redirect_url)

    q = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    qs = SkuRecipe.objects.filter(is_active=False)
    if q:
        qs = qs.filter(
            Q(sku__icontains=q)
            | Q(job_name__icontains=q)
            | Q(material__icontains=q)
            | Q(machine_name__icontains=q)
        )
    if status_filter in ('draft', 'pending_review', 'reviewed', 'approved'):
        qs = qs.filter(master_data_status=status_filter)
    paginator = Paginator(qs, 50)
    recipes = paginator.get_page(request.GET.get('page'))
    return render(request, 'planning/sku_recipes_archived.html', {
        'recipes': recipes,
        'q': q,
        'status_filter': status_filter,
        'can_restore_approved': is_admin_user,
        'can_delete_archived': is_admin_user,
    })


@login_required
@permission_required('can_edit_jobcard')
def sku_recipe_edit(request, recipe_id=None):
    """Create or edit a single SKU recipe."""
    if recipe_id:
        recipe = get_object_or_404(SkuRecipe, id=recipe_id, is_active=True)
        page_title = f'Edit SKU Recipe — {recipe.sku}'
    else:
        recipe = None
        page_title = 'Add New SKU Recipe'

    is_admin_user = _user_is_admin(request.user)
    can_edit_approved = not (recipe and recipe.master_data_status == 'approved')
    can_admin_actions = is_admin_user

    if request.method == 'POST':
        action = request.POST.get('action', '').strip()
        _do_sync_on_approve = False
        if recipe and action == 'delete':
            if recipe.master_data_status == 'approved' and not is_admin_user:
                messages.error(request, 'Approved records can only be deleted by admin users.')
            else:
                recipe.delete()
                messages.success(request, f'SKU Recipe "{recipe.sku}" deleted.')
            return redirect('qc:sku_recipes')

        if recipe and action == 'archive':
            if recipe.master_data_status == 'approved' and not is_admin_user:
                messages.error(request, 'Approved records can only be archived by admin users.')
                return redirect('qc:sku_recipes')
            recipe.is_active = False
            recipe.archived_by = request.user
            recipe.archived_at = timezone.now()
            recipe.archive_reason = (request.POST.get('archive_reason') or '').strip()
            recipe.save(update_fields=['is_active', 'archived_by', 'archived_at', 'archive_reason', 'updated_at'])
            messages.success(request, f'SKU Recipe "{recipe.sku}" archived.')
            return redirect('qc:sku_recipes')

        if recipe and recipe.master_data_status == 'approved' and action != 'reopen_sku':
            messages.error(request, 'Approved SKU is locked. Use Reopen SKU before making edits.')
            return render(request, 'planning/sku_recipe_edit.html', {'form': SkuRecipeForm(instance=recipe), 'recipe': recipe, 'page_title': page_title, 'can_edit_approved': can_edit_approved, 'can_admin_actions': can_admin_actions})

        form = SkuRecipeForm(request.POST, instance=recipe)
        if form.is_valid():
            obj = form.save(commit=False)
            if not recipe_id:
                obj.created_by = request.user

            current_status = (recipe.master_data_status if recipe else 'draft')
            obj.master_data_status = current_status
            obj.reviewed_by = recipe.reviewed_by if recipe else None
            obj.reviewed_at = recipe.reviewed_at if recipe else None
            obj.approved_by = recipe.approved_by if recipe else None
            obj.approved_at = recipe.approved_at if recipe else None

            if recipe_id and action:
                if action == 'submit_review' and current_status == 'draft':
                    obj.master_data_status = 'pending_review'
                    obj.reviewed_by = None
                    obj.reviewed_at = None
                    obj.approved_by = None
                    obj.approved_at = None
                    messages.success(request, f'SKU Recipe "{obj.sku}" submitted for review. Status: Pending Review.')
                elif action == 'review' and current_status == 'pending_review':
                    obj.master_data_status = 'reviewed'
                    obj.reviewed_by = request.user
                    obj.reviewed_at = timezone.now()
                    obj.approved_by = None
                    obj.approved_at = None
                    messages.success(request, f'SKU Recipe "{obj.sku}" reviewed and submitted for approval.')
                elif action == 'approve' and current_status == 'reviewed':
                    missing = _missing_required_master_fields(obj)
                    if missing:
                        messages.error(request, f'Cannot approve. Missing required fields: {", ".join(missing)}.')
                        return render(request, 'planning/sku_recipe_edit.html', {'form': form, 'recipe': obj, 'page_title': page_title, 'can_edit_approved': can_edit_approved, 'can_admin_actions': can_admin_actions})
                    obj.master_data_status = 'approved'
                    obj.approved_by = request.user
                    obj.approved_at = timezone.now()
                    # Will sync to planning after save
                    _do_sync_on_approve = True
                    messages.success(request, f'SKU Recipe "{obj.sku}" approved for master data usage.')
                elif action == 'back_to_draft' and current_status in ('pending_review', 'reviewed'):
                    comment = (request.POST.get('rejection_comment') or '').strip()
                    if not comment:
                        messages.error(request, 'Please provide a reason when sending a record back to Draft.')
                        return render(request, 'planning/sku_recipe_edit.html', {'form': form, 'recipe': obj, 'page_title': page_title, 'can_edit_approved': can_edit_approved})
                    obj.master_data_status = 'draft'
                    obj.reviewed_by = None
                    obj.reviewed_at = None
                    obj.approved_by = None
                    obj.approved_at = None
                    obj.rejection_comment = comment
                    obj.last_rejected_by = request.user
                    obj.last_rejected_at = timezone.now()
                    messages.success(request, f'SKU Recipe "{obj.sku}" moved back to Draft.')
                elif action == 'reopen_sku' and current_status == 'approved':
                    comment = (request.POST.get('rejection_comment') or '').strip()
                    if not comment:
                        messages.error(request, 'Please provide a reopen reason before moving approved SKU back to Draft.')
                        return render(request, 'planning/sku_recipe_edit.html', {'form': form, 'recipe': obj, 'page_title': page_title, 'can_edit_approved': can_edit_approved})
                    obj.master_data_status = 'draft'
                    obj.reviewed_by = None
                    obj.reviewed_at = None
                    obj.approved_by = None
                    obj.approved_at = None
                    obj.rejection_comment = comment
                    obj.last_rejected_by = request.user
                    obj.last_rejected_at = timezone.now()
                    messages.success(request, f'SKU Recipe "{obj.sku}" reopened to Draft.')
                else:
                    messages.info(request, f'SKU Recipe "{obj.sku}" saved without changing workflow status.')
            else:
                if recipe_id:
                    messages.success(request, f'SKU Recipe "{obj.sku}" saved.')
                else:
                    obj.master_data_status = 'draft'
                    obj.reviewed_by = None
                    obj.reviewed_at = None
                    obj.approved_by = None
                    obj.approved_at = None
                    messages.success(request, f'SKU Recipe "{obj.sku}" saved as Draft. Submit for approval from SKU Recipe Master.')

            obj.save()
            if _do_sync_on_approve:
                try:
                    sync_result = _sync_new_jobs_for_approved_sku(obj.sku, actor=request.user)
                    messages.success(
                        request,
                        f'Planning jobs refreshed for approved SKU: updated {sync_result["updated"]}, locked {sync_result["locked"]}, missing draft jobs {sync_result.get("missing_jobs", 0)}.',
                    )
                except Exception:
                    logger.exception('Approved SKU sync to planning failed for %s', obj.sku)
                    messages.error(request, 'Error while sending approved SKU to Planning; check logs.')
            return redirect('qc:sku_recipes')
        else:
            # Surface a clear top-level message so users notice validation errors
            messages.error(request, 'There are errors in the form. Please correct the highlighted fields and try again.')
    else:
        form = SkuRecipeForm(instance=recipe)

    return render(request, 'planning/sku_recipe_edit.html', {
        'form': form,
        'recipe': recipe,
        'page_title': page_title,
        'can_edit_approved': can_edit_approved,
        'can_admin_actions': can_admin_actions,
    })


@login_required
@permission_required('can_edit_jobcard')
def sku_recipe_bulk_upload(request):
    """Bulk upload SKU recipes from CSV/XLSX into Draft status for approval workflow."""
    if request.method == 'POST':
        upload_file = request.FILES.get('upload_file')
        if not upload_file:
            messages.error(request, 'Please choose a CSV or XLSX file to upload.')
            return redirect('qc:sku_recipe_bulk_upload')


        name = (upload_file.name or '').lower()
        rows = []
        # Map Google Sheet headers to model fields (robust, case-insensitive, and with all required fields)
        header_to_field = {
            'SKU': 'sku',
            'JOB NAME': 'job_name',
            'Material': 'material',
            'Color': 'color_spec',
            'Application': 'application',
            'Size W mm': 'size_w_mm',
            'Size H mm': 'size_h_mm',
            'Size W Inch': 'size_w_inch',
            'Size H Inch': 'size_h_inch',
            'Ups': 'ups',
            'Print Sheet Size': 'print_sheet_size',
            'Purchase Sheet Size': 'purchase_sheet_size',
            'Purchase Sheet ups': 'purchase_sheet_ups',
            'Cost': 'default_unit_cost',
            'Default Unit Cost': 'default_unit_cost',
            'Daily Demand': 'daily_demand',
            'AWC No.': 'awc_no',
            'AWC No': 'awc_no',
            'Die': 'die_cutting',
            'Notes': 'notes',
        }
        int_clean_fields = {'size_w_mm', 'size_h_mm', 'size_w_inch', 'size_h_inch', 'ups', 'purchase_sheet_ups', 'daily_demand'}
        def clean_intlike(val):
            try:
                if val is None or str(val).strip() == '':
                    return ''
                ival = int(float(val))
                return str(ival)
            except:
                return str(val) if val is not None else ''
        try:
            if name.endswith('.csv'):
                decoded = upload_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                for row in reader:
                    rows.append(row)
            elif name.endswith('.xlsx'):
                try:
                    import openpyxl
                except ImportError:
                    messages.error(request, 'openpyxl is required for XLSX upload.')
                    return redirect('qc:sku_recipe_bulk_upload')
                wb = openpyxl.load_workbook(upload_file, data_only=True)
                ws = wb.active
                # Find header row: look for row with 'SKU' and 'JOB NAME'
                header_row_idx = None
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    values = [str(c).strip() if c is not None else '' for c in row]
                    if 'SKU' in values and 'JOB NAME' in values:
                        header_row_idx = i
                        header = values
                        break
                if not header_row_idx:
                    messages.error(request, 'Could not find header row in Excel file. Make sure it matches the template.')
                    return redirect('qc:sku_recipe_bulk_upload')
                for values in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
                    row = {}
                    for idx, key in enumerate(header):
                        if key:
                            row[key] = values[idx] if idx < len(values) else None
                    rows.append(row)
            else:
                messages.error(request, 'Unsupported file type. Please upload CSV or XLSX.')
                return redirect('qc:sku_recipe_bulk_upload')
        except Exception as exc:
            messages.error(request, f'Could not read upload file: {exc}')
            return redirect('qc:sku_recipe_bulk_upload')

        if not rows:
            messages.error(request, 'No rows found in upload file.')
            return redirect('qc:sku_recipe_bulk_upload')

        created = 0
        updated = 0
        failed = 0
        sample_errors = []
        bulk_highlights = {}
        highlight_fields = {
            'sku', 'job_name', 'material', 'color_spec', 'application',
            'size_w_mm', 'size_h_mm', 'ups', 'print_sheet_size',
            'purchase_sheet_size', 'purchase_sheet_ups',
            'default_unit_cost', 'daily_demand',
            'awc_no', 'die_cutting', 'notes',
        }

        for idx, source in enumerate(rows, start=2):
            payload = {}
            lam_fnb = False
            for header, field in header_to_field.items():
                value = source.get(header, '')
                # Robust normalization for application
                if field == 'application':
                    if 'f+b' in (str(value or '').lower()):
                        lam_fnb = True
                    payload[field] = _normalize_application_input(value)
                elif field == 'color_spec':
                    payload[field] = _normalize_color_spec_input(value)
                # Remove decimals for mm columns
                elif field in {'size_w_mm', 'size_h_mm'}:
                    try:
                        if value is None or str(value).strip() == '':
                            payload[field] = ''
                        else:
                            payload[field] = str(int(float(value)))
                    except:
                        payload[field] = str(value) if value is not None else ''
                elif field in int_clean_fields:
                    payload[field] = clean_intlike(value)
                else:
                    payload[field] = '' if value is None else str(value).strip()
            payload['lamination_front_and_back'] = lam_fnb
            if not payload.get('sku'):
                continue
            existing = SkuRecipe.objects.filter(sku__iexact=payload['sku']).first()
            form = SkuRecipeForm(payload, instance=existing)
            if not form.is_valid():
                failed += 1
                if len(sample_errors) < 8:
                    error_text = '; '.join(
                        f"{name}: {', '.join([str(msg) for msg in msgs])}"
                        for name, msgs in form.errors.items()
                    )
                    sample_errors.append(f'Row {idx} ({payload["sku"]}): {error_text}')
                continue
            obj = form.save(commit=False)
            if not existing:
                obj.created_by = request.user
            obj.master_data_status = 'draft'
            obj.reviewed_by = None
            obj.reviewed_at = None
            obj.approved_by = None
            obj.approved_at = None
            obj.save()

            if existing:
                changed = [field for field in form.changed_data if field in highlight_fields]
                if not changed:
                    changed = [
                        field for field in highlight_fields
                        if str(getattr(existing, field, '') or '').strip() != str(form.cleaned_data.get(field, '') or '').strip()
                    ]
                if not changed:
                    changed = ['sku']
                bulk_highlights[str(obj.id)] = {
                    'type': 'updated',
                    'fields': changed,
                }
                updated += 1
            else:
                created_fields = [
                    field for field in form.cleaned_data
                    if field in highlight_fields and form.cleaned_data.get(field) not in (None, '')
                ]
                bulk_highlights[str(obj.id)] = {
                    'type': 'created',
                    'fields': created_fields,
                }
                created += 1

        if bulk_highlights:
            request.session['sku_recipe_bulk_highlights'] = bulk_highlights

        if created or updated:
            messages.success(
                request,
                f'Bulk upload complete. Draft recipes created {created}, updated {updated}, failed {failed}.',
            )
        if failed and sample_errors:
            messages.error(request, 'Sample row errors: ' + ' | '.join(sample_errors))

        return redirect('qc:sku_recipes')

    return render(request, 'planning/sku_recipe_bulk_upload.html')


@login_required
@permission_required('can_edit_jobcard')
def sku_recipe_template_download(request):
    """Return a CSV template for bulk SKU recipe upload."""
    headers = [
        'Sno.', 'SKU', 'JOB NAME', 'Order Status', 'Material', 'Color', 'Application',
        'Size W mm', 'Size H mm', 'Size W Inch', 'Size H Inch', 'Ups', 'Print Sheet Size',
        'Purchase Sheet Size', 'Purchase Sheet ups',
        'Default Unit Cost', 'Daily Demand', 'AWC No', 'Die', 'Notes'
    ]
    sample_row = [
        '1', 'SKU-001', 'Sample Job Name', 'Repeat', 'Art Card 300gsm', '4 color', 'UV',
        '100', '150', '3.94', '5.91', '4', '720x1020', '720x1020', '2',
        '5.00', '500', 'AWC-001', 'YES', 'Sample notes'
    ]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(sample_row)
    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="sku_recipe_upload_template.csv"'
    return response


@login_required
@permission_required('can_edit_jobcard')
@transaction.atomic
def pending_skus(request):
    """Central queue of SKUs that are still missing in SKU Recipe master data."""
    is_admin_user = _user_is_admin(request.user)
    if request.method == 'POST':
        action = (request.POST.get('action') or 'save').strip()
        sku = (request.POST.get('sku') or '').strip()
        po_doc_id = request.POST.get('po_doc_id')
        return_po = (request.POST.get('return_po') or '').strip()
        return_q = (request.POST.get('return_q') or '').strip()

        def _redirect_pending():
            params = {}
            if return_po:
                params['po'] = return_po
            if return_q:
                params['q'] = return_q
            url = reverse('qc:pending_skus')
            return redirect(f'{url}?{urlencode(params)}' if params else url)

        if action in {'delete', 'archive'}:
            recipe_id = request.POST.get('recipe_id')
            try:
                recipe_obj = SkuRecipe.objects.get(id=int(recipe_id))
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
                return _redirect_pending()

            if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                messages.error(request, 'Approved records can only be managed by admin users.')
                return _redirect_pending()

            if action == 'delete':
                recipe_obj.delete()
                messages.success(request, f'SKU recipe {recipe_obj.sku} deleted.')
            else:
                recipe_obj.is_active = False
                recipe_obj.archived_by = request.user
                recipe_obj.archived_at = timezone.now()
                recipe_obj.archive_reason = (request.POST.get('archive_reason') or '').strip()
                recipe_obj.save(update_fields=['is_active', 'archived_by', 'archived_at', 'archive_reason', 'updated_at'])
                messages.success(request, f'SKU recipe {recipe_obj.sku} archived.')

            return _redirect_pending()

        if action == 'ignore':
            try:
                po_doc = PoDocument.objects.get(id=int(po_doc_id)) if po_doc_id else None
            except (TypeError, ValueError, PoDocument.DoesNotExist):
                po_doc = None

            if not po_doc or not sku:
                messages.error(request, 'Invalid SKU or PO reference for ignore action.')
                return _redirect_pending()

            payload = po_doc.extracted_payload or {}
            ignored = { _sku_key(s) for s in (payload.get('new_skus_ignored') or []) if s }
            ignored.add(_sku_key(sku))
            payload['new_skus_ignored'] = sorted(ignored)
            po_doc.extracted_payload = payload
            po_doc.save(update_fields=['extracted_payload'])
            messages.success(request, f'SKU {sku} will be ignored and removed from pending processing.')
            return _redirect_pending()

        if not sku:
            messages.error(request, 'SKU is required.')
            return _redirect_pending()

        if action in {'submit_review', 'approve', 'back_to_draft'}:
            recipe = SkuRecipe.objects.filter(sku__iexact=sku).first()
            if not recipe:
                messages.error(request, f'No SKU recipe found for {sku}. Save recipe data first.')
                return _redirect_pending()

            current_status = (recipe.master_data_status or 'draft').lower()

            if action == 'submit_review':
                if current_status != 'draft':
                    messages.error(request, f'SKU {sku} can only move to Pending Review from Draft.')
                    return _redirect_pending()
                recipe.master_data_status = 'pending_review'
                recipe.reviewed_by = None
                recipe.reviewed_at = None
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.save(update_fields=['master_data_status', 'reviewed_by', 'reviewed_at', 'approved_by', 'approved_at', 'updated_at'])
                messages.success(request, f'SKU {sku} moved to Pending Review.')
                return _redirect_pending()

            if action == 'approve':
                if current_status != 'reviewed':
                    messages.error(request, f'SKU {sku} can only be Approved from Reviewed status.')
                    return _redirect_pending()
                missing_required = _missing_required_master_fields(recipe)
                if missing_required:
                    messages.error(
                        request,
                        f'SKU {sku} cannot be approved. Missing required master data: {", ".join(missing_required)}.',
                    )
                    return _redirect_pending()
                recipe.master_data_status = 'approved'
                recipe.approved_by = request.user
                recipe.approved_at = timezone.now()
                recipe.save(update_fields=['master_data_status', 'approved_by', 'approved_at', 'updated_at'])
                sync_result = _sync_new_jobs_for_approved_sku(sku, actor=request.user)
                messages.success(
                    request,
                    f'SKU {sku} approved for master data usage. Planning jobs refreshed: updated {sync_result["updated"]}, locked {sync_result["locked"]}, missing draft jobs {sync_result.get("missing_jobs", 0)}.',
                )
                return _redirect_pending()

            if action == 'back_to_draft':
                if current_status == 'draft':
                    messages.info(request, f'SKU {sku} is already in Draft.')
                    return _redirect_pending()
                recipe.master_data_status = 'draft'
                recipe.reviewed_by = None
                recipe.reviewed_at = None
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.save(update_fields=['master_data_status', 'reviewed_by', 'reviewed_at', 'approved_by', 'approved_at', 'updated_at'])
                messages.success(request, f'SKU {sku} moved back to Draft.')
                return _redirect_pending()

        job_name = (request.POST.get('job_name') or '').strip()
        material = (request.POST.get('material') or '').strip()
        color_spec = (request.POST.get('color_spec') or '').strip()
        application = (request.POST.get('application') or '').strip()
        department = (request.POST.get('department') or '').strip()
        print_sheet_size = (request.POST.get('print_sheet_size') or '').strip()
        purchase_sheet_size = (request.POST.get('purchase_sheet_size') or '').strip()
        purchase_sheet_ups = _to_optional_positive_int(request.POST.get('purchase_sheet_ups'))
        ups = _to_optional_positive_int(request.POST.get('ups'))
        daily_demand = _to_optional_decimal(request.POST.get('daily_demand'))
        awc_no = (request.POST.get('awc_no') or '').strip()
        die_cutting = (request.POST.get('die_cutting') or '').strip()

        unit_cost_raw = (request.POST.get('default_unit_cost') or '').strip()
        unit_cost = None
        if unit_cost_raw:
            try:
                unit_cost = Decimal(unit_cost_raw)
            except InvalidOperation:
                unit_cost = None

        if not job_name and not material:
            messages.error(request, 'Please enter at least Job Name or Material before saving.')
            return _redirect_pending()

        SkuRecipe.objects.update_or_create(
            sku=sku,
            defaults={
                'job_name': job_name,
                'material': material,
                'color_spec': color_spec,
                'application': application,
                'print_sheet_size': print_sheet_size,
                'purchase_sheet_size': purchase_sheet_size,
                'purchase_sheet_ups': purchase_sheet_ups,
                'ups': ups,
                'default_unit_cost': unit_cost,
                'daily_demand': daily_demand,
                'awc_no': awc_no,
                'die_cutting': die_cutting,
                'created_by': request.user,
                'master_data_status': 'draft',
                'reviewed_by': None,
                'reviewed_at': None,
                'approved_by': None,
                'approved_at': None,
            },
        )

        if po_doc_id:
            try:
                po_doc = PoDocument.objects.filter(id=int(po_doc_id)).first()
            except (TypeError, ValueError):
                po_doc = None
            if po_doc:
                payload = po_doc.extracted_payload or {}
                configured = set(payload.get('new_skus_configured') or [])
                configured.add(sku)
                payload['new_skus_configured'] = sorted(configured)
                po_doc.extracted_payload = payload
                po_doc.save(update_fields=['extracted_payload'])

        messages.success(request, f'SKU recipe saved for {sku}.')
        return _redirect_pending()

    po_filter = (request.GET.get('po') or '').strip()
    q = (request.GET.get('q') or '').strip()

    po_docs = PoDocument.objects.exclude(extracted_payload__isnull=True).order_by('-created_at')[:400]
    deduped_docs = []
    seen_po_numbers = set()
    for doc in po_docs:
        payload = doc.extracted_payload or {}
        po_number = (payload.get('po_number') or '').strip().upper()
        if po_number:
            if po_number in seen_po_numbers:
                continue
            seen_po_numbers.add(po_number)
        deduped_docs.append(doc)
    po_docs = deduped_docs[:200]
    all_pending_rows = _collect_pending_sku_rows(po_docs)

    po_summary_map = {}
    for row in all_pending_rows:
        po_key = row.get('po_number') or '-'
        current = po_summary_map.get(po_key)
        if not current:
            po_summary_map[po_key] = {
                'po_number': po_key,
                'count': 1,
                'po_doc_id': row.get('po_doc_id'),
            }
        else:
            current['count'] += 1

    pending_rows = all_pending_rows
    if po_filter:
        pending_rows = [row for row in pending_rows if (row.get('po_number') or '') == po_filter]
    if q:
        q_upper = q.upper()
        pending_rows = [
            row
            for row in pending_rows
            if q_upper in (row.get('sku') or '').upper()
            or q_upper in (row.get('po_number') or '').upper()
            or q_upper in (row.get('job_name') or '').upper()
        ]

    sku_values = sorted({row['sku'] for row in pending_rows if row.get('sku')})
    recipes_by_sku = {}
    if sku_values:
        recipe_query = Q()
        for sku in sku_values:
            recipe_query |= Q(sku__iexact=sku)
        recipes = SkuRecipe.objects.filter(recipe_query)
        recipes_by_sku = {recipe.sku.upper(): recipe for recipe in recipes}

    for row in pending_rows:
        recipe = recipes_by_sku.get(_sku_key(row.get('sku')))
        row['recipe'] = recipe
        row['recipe_status'] = recipe.master_data_status if recipe else 'missing'
        row['missing_required_fields'] = _missing_required_master_fields(recipe, row.get('job_name') or '')

    pending_rows.sort(key=lambda row: (row['po_number'], row['sku']))

    context = {
        'pending_rows': pending_rows,
        'pending_count': len(pending_rows),
        'po_summary': sorted(po_summary_map.values(), key=lambda x: x['po_number']),
        'po_filter': po_filter,
        'q': q,
        'can_admin_actions': is_admin_user,
    }
    return render(request, 'planning/pending_skus.html', context)


@login_required
@permission_required('can_edit_jobcard')
@transaction.atomic
def pending_skus_ignored(request):
    """Display pending SKUs that were marked ignored and no longer appear in the active pending queue."""
    is_admin_user = _user_is_admin(request.user)
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        sku = (request.POST.get('sku') or '').strip()
        po_doc_id = request.POST.get('po_doc_id')
        if action == 'unignore':
            try:
                po_doc = PoDocument.objects.get(id=int(po_doc_id)) if po_doc_id else None
            except (TypeError, ValueError, PoDocument.DoesNotExist):
                po_doc = None

            if not po_doc or not sku:
                messages.error(request, 'Invalid PO or SKU for unignore action.')
                return redirect('qc:pending_skus_ignored')

            payload = po_doc.extracted_payload or {}
            ignored = [s for s in (payload.get('new_skus_ignored') or []) if s]
            normalized = _sku_key(sku)
            kept = [s for s in ignored if _sku_key(s) != normalized]
            payload['new_skus_ignored'] = sorted(kept)
            po_doc.extracted_payload = payload
            po_doc.save(update_fields=['extracted_payload'])
            messages.success(request, f'SKU {sku} restored to the pending queue.')
            return redirect('qc:pending_skus_ignored')

    po_filter = (request.GET.get('po') or '').strip()
    q = (request.GET.get('q') or '').strip()

    docs = PoDocument.objects.exclude(extracted_payload__isnull=True).order_by('-created_at')[:400]
    rows = []
    for doc in docs:
        payload = doc.extracted_payload or {}
        ignored_skus = [s for s in (payload.get('new_skus_ignored') or []) if s]
        if not ignored_skus:
            continue

        po_number = payload.get('po_number') or '-'
        for sku in ignored_skus:
            if not sku:
                continue
            recipe = SkuRecipe.objects.filter(sku__iexact=sku).first()
            rows.append({
                'po_doc_id': doc.id,
                'po_number': po_number,
                'supplier': payload.get('supplier_name') or '-',
                'sku': sku,
                'job_name': recipe.job_name if recipe else '-',
                'recipe_status': recipe.master_data_status if recipe else 'missing',
                'recipe': recipe,
                'uploaded_at': doc.created_at,
            })

    if po_filter:
        rows = [row for row in rows if row['po_number'] == po_filter]
    if q:
        q_upper = q.upper()
        rows = [
            row
            for row in rows
            if q_upper in (row['sku'] or '').upper()
            or q_upper in (row['po_number'] or '').upper()
            or q_upper in (row['job_name'] or '').upper()
        ]

    po_summary = {}
    for row in rows:
        po_summary.setdefault(row['po_number'], {'po_number': row['po_number'], 'count': 0})
        po_summary[row['po_number']]['count'] += 1

    rows.sort(key=lambda row: (row['po_number'], row['sku']))

    return render(request, 'planning/pending_skus_ignored.html', {
        'rows': rows,
        'po_summary': sorted(po_summary.values(), key=lambda x: x['po_number']),
        'po_filter': po_filter,
        'q': q,
        'can_admin_actions': is_admin_user,
    })


@login_required
@permission_required('can_edit_jobcard')
@transaction.atomic
def pending_sku_master_entry(request):
    """Open a focused form for one pending SKU and send it through master-data approval flow."""
    sku = (request.GET.get('sku') or request.POST.get('sku') or '').strip()
    po_doc_id_raw = request.GET.get('po_doc_id') or request.POST.get('po_doc_id')
    return_po = (request.GET.get('return_po') or request.POST.get('return_po') or '').strip()
    return_q = (request.GET.get('return_q') or request.POST.get('return_q') or '').strip()
    is_readonly = (request.GET.get('readonly') or request.POST.get('readonly') or '').strip() in {'1', 'true', 'yes'}

    def _redirect_pending():
        params = {}
        if return_po:
            params['po'] = return_po
        if return_q:
            params['q'] = return_q
        url = reverse('qc:pending_skus')
        return redirect(f'{url}?{urlencode(params)}' if params else url)

    try:
        po_doc_id = int(po_doc_id_raw)
    except (TypeError, ValueError):
        po_doc_id = None

    if not sku or not po_doc_id:
        messages.error(request, 'Missing SKU or PO reference for master-data entry.')
        return _redirect_pending()

    po_doc = PoDocument.objects.filter(id=po_doc_id).first()
    if not po_doc:
        messages.error(request, 'PO document was not found.')
        return _redirect_pending()

    payload = po_doc.extracted_payload or {}
    po_number = payload.get('po_number') or '-'
    items = _sanitize_po_payload_items(payload)
    is_admin_user = _user_is_admin(request.user)

    suggested_item = None
    sku_key = _sku_key(sku)
    for item in items:
        if _sku_key(item.get('sku')) == sku_key:
            suggested_item = item
            break
    suggested_item = suggested_item or {}
    po_job_name = (suggested_item.get('job_name') or '').strip() or sku
    po_department = (payload.get('department') or '').strip()
    po_unit_cost = _to_decimal(suggested_item.get('unit_cost'))
    po_color_spec = _normalize_color_spec_input(
        suggested_item.get('color_spec') or suggested_item.get('colour') or suggested_item.get('color') or ''
    )
    po_application = _normalize_application_input(suggested_item.get('application') or payload.get('application') or '')

    recipe = SkuRecipe.objects.filter(sku__iexact=sku).first()

    if request.method == 'POST':
        if is_readonly:
            messages.error(request, 'Read-only mode does not allow edits.')
            return _redirect_pending()

        action = (request.POST.get('action') or 'save_draft').strip()
        if recipe and action == 'delete':
            if recipe.master_data_status == 'approved' and not is_admin_user:
                messages.error(request, 'Approved records can only be deleted by admin users.')
            else:
                recipe.delete()
                messages.success(request, f'SKU Recipe "{recipe.sku}" deleted.')
            return _redirect_pending()

        if recipe and action == 'archive':
            if recipe.master_data_status == 'approved' and not is_admin_user:
                messages.error(request, 'Approved records can only be archived by admin users.')
                return _redirect_pending()
            recipe.is_active = False
            recipe.archived_by = request.user
            recipe.archived_at = timezone.now()
            recipe.archive_reason = (request.POST.get('archive_reason') or '').strip()
            recipe.save(update_fields=['is_active', 'archived_by', 'archived_at', 'archive_reason', 'updated_at'])
            messages.success(request, f'SKU Recipe "{recipe.sku}" archived.')
            return _redirect_pending()

        posted = request.POST.copy()
        # Job name is sourced from PO parsing; keep it authoritative and non-editable.
        posted['job_name'] = po_job_name
        posted['sku'] = sku
        if not (posted.get('default_unit_cost') or '').strip() and po_unit_cost is not None:
            posted['default_unit_cost'] = str(po_unit_cost)
        if not (posted.get('color_spec') or '').strip() and po_color_spec:
            posted['color_spec'] = po_color_spec
        if not (posted.get('application') or '').strip() and po_application:
            posted['application'] = po_application
        form = SkuRecipeForm(posted, instance=recipe)
        if form.is_valid():
            action = (request.POST.get('action') or 'save_draft').strip()
            obj = form.save(commit=False)
            obj.sku = sku
            obj.job_name = po_job_name
            if not recipe:
                obj.created_by = request.user

            if action == 'submit_review':
                missing_required = _missing_required_master_fields(obj)
                if missing_required:
                    messages.error(
                        request,
                        f'SKU {sku} cannot be sent for QC review. Missing required data: {", ".join(missing_required)}.',
                    )
                else:
                    obj.master_data_status = 'pending_review'
                    obj.reviewed_by = None
                    obj.reviewed_at = None
                    obj.approved_by = None
                    obj.approved_at = None
                    obj.save()

                    configured = set(payload.get('new_skus_configured') or [])
                    configured.add(sku)
                    payload['new_skus_configured'] = sorted(configured)
                    po_doc.extracted_payload = payload
                    po_doc.save(update_fields=['extracted_payload'])

                    messages.success(request, f'SKU {sku} submitted for QC review.')
                    return _redirect_pending()
            else:
                obj.master_data_status = 'draft'
                obj.reviewed_by = None
                obj.reviewed_at = None
                obj.approved_by = None
                obj.approved_at = None
                obj.save()

                configured = set(payload.get('new_skus_configured') or [])
                configured.add(sku)
                payload['new_skus_configured'] = sorted(configured)
                po_doc.extracted_payload = payload
                po_doc.save(update_fields=['extracted_payload'])

                messages.success(request, f'SKU {sku} saved as Draft.')
                return _redirect_pending()
    else:
        initial = {
            'sku': sku,
            'job_name': po_job_name,
            'default_unit_cost': (recipe.default_unit_cost if recipe else None) or po_unit_cost,
            'color_spec': (recipe.color_spec if recipe else '') or po_color_spec,
            'application': (recipe.application if recipe else '') or po_application,
        }
        form = SkuRecipeForm(instance=recipe, initial=initial)

    form.fields['sku'].widget.attrs['readonly'] = True
    if is_readonly:
        for field in form.fields.values():
            field.disabled = True

    current_recipe = recipe
    if request.method == 'POST' and form.is_valid() and 'obj' in locals():
        current_recipe = obj

    mismatch_alerts = []
    if current_recipe:
        cost_alert = _build_cost_mismatch_note(current_recipe.default_unit_cost, po_unit_cost)
        if cost_alert:
            mismatch_alerts.append(cost_alert)

    context = {
        'form': form,
        'sku': sku,
        'po_doc_id': po_doc_id,
        'po_number': po_number,
        'return_po': return_po,
        'return_q': return_q,
        'suggested_job_name': po_job_name,
        'suggested_qty': _format_display_qty(suggested_item.get('quantity')),
        'suggested_delivery_date': suggested_item.get('delivery_date') or '-',
        'recipe_status': (current_recipe.master_data_status if current_recipe else 'missing'),
        'missing_required_fields': _missing_required_master_fields(current_recipe, po_job_name),
        'mismatch_alerts': mismatch_alerts,
        'is_readonly': is_readonly,
    }
    context['can_admin_actions'] = is_admin_user
    return render(request, 'planning/pending_sku_master_entry.html', context)


@login_required
@permission_required('can_edit_jobcard')
@transaction.atomic
def po_inbox(request):
    """PO intake queue after upload: split-ready documents with repeat/new counts."""
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        if action == 'delete_po_intake':
            if not _user_is_admin(request.user):
                messages.error(request, 'Only administrators can delete PO intake records.')
                return redirect('planning:po_inbox')

            po_doc_id = (request.POST.get('po_doc_id') or '').strip()
            po_number = (request.POST.get('po_number') or '').strip()

            if po_doc_id:
                try:
                    doc = PoDocument.objects.get(id=int(po_doc_id))
                except (TypeError, ValueError, PoDocument.DoesNotExist):
                    messages.error(request, 'Invalid PO document selected for delete action.')
                    return redirect('planning:po_inbox')

                try:
                    if doc.po_file:
                        doc.po_file.delete(save=False)
                except Exception:
                    pass
                doc.delete()

                messages.success(request, f'Deleted PO intake document {doc.id}.')
                return redirect('planning:po_inbox')

            if not po_number:
                messages.error(request, 'Invalid PO number for delete action.')
                return redirect('planning:po_inbox')

            docs_to_delete = PoDocument.objects.filter(extracted_payload__po_number__iexact=po_number)
            count = docs_to_delete.count()
            if count == 0:
                messages.error(request, f'No PO intake found for {po_number}.')
                return redirect('planning:po_inbox')

            for doc in docs_to_delete:
                try:
                    if doc.po_file:
                        doc.po_file.delete(save=False)
                except Exception:
                    pass
                doc.delete()

            messages.success(request, f'Deleted PO intake {po_number} and {count} document(s).')
            return redirect('planning:po_inbox')

    docs = PoDocument.objects.exclude(extracted_payload__isnull=True).order_by('-created_at')[:400]
    deduped_docs = []
    seen_po_numbers = set()
    for doc in docs:
        payload = doc.extracted_payload or {}
        po_number = (payload.get('po_number') or '').strip().upper()
        if po_number:
            if po_number in seen_po_numbers:
                continue
            seen_po_numbers.add(po_number)
        deduped_docs.append(doc)

    docs = deduped_docs[:200]
    rows = []
    for doc in docs:
        payload = doc.extracted_payload or {}
        items = _po_payload_items(payload)
        recipe_map = _build_recipe_map(items)
        _, _, _, missing_skus = _annotate_items_with_recipe(items, recipe_map)
        repeat_count, new_count = _history_repeat_new_counts(items)
        rows.append(
            {
                'doc': doc,
                'po_doc_id': doc.id,
                'po_number': payload.get('po_number') or '-',
                'supplier': payload.get('supplier_name') or '-',
                'item_count': len(items),
                'repeat_count': repeat_count,
                'new_count': new_count,
                'missing_count': len(missing_skus),
                'ignored_count': len([s for s in (payload.get('new_skus_ignored') or []) if s]),
                'repeat_jobs_created_count': payload.get('repeat_jobs_created_count') or 0,
                'repeat_jobs_updated_count': payload.get('repeat_jobs_updated_count') or 0,
                'repeat_jobs_locked_count': payload.get('repeat_jobs_locked_count') or 0,
                'repeat_jobs_missing_recipe_count': payload.get('repeat_jobs_missing_recipe_count') or 0,
                'new_skus_sent_to_planning_count': len(payload.get('new_skus_sent_to_planning') or []),
                'merged_duplicates': bool(payload.get('merged_duplicate_skus')),
                'merged_duplicate_skus': payload.get('merged_duplicate_skus') or [],
                'source_item_count': payload.get('source_item_count') or len(items),
            }
        )

    return render(request, 'planning/po_inbox.html', {'rows': rows, 'can_admin_actions': _user_is_admin(request.user)})




@login_required
@permission_required('can_edit_jobcard')
def upload_po(request):
    """Upload a PO PDF, extract its content, store it, and redirect to review."""
    if request.method == 'POST':
        pdf_file = request.FILES.get('po_pdf')
        if not pdf_file:
            messages.error(request, 'Please select a PDF file.')
            return redirect('planning:upload_po')

        if not pdf_file.name.lower().endswith('.pdf'):
            messages.error(request, 'Only PDF files are supported.')
            return redirect('planning:upload_po')

        try:
            extracted = extract_po_from_pdf(pdf_file)
        except ValueError as exc:
            messages.error(request, f'Extraction failed: {exc}')
            return redirect('planning:upload_po')
        except Exception as exc:
            messages.error(request, f'Unexpected error reading PDF: {exc}')
            return redirect('planning:upload_po')

        # Surface partial-extraction warning before saving so user sees it on review page.
        extraction_warning = extracted.pop('extraction_warning', None)
        expected_count = extracted.pop('expected_line_count', None)

        items = extracted.get('items', [])
        source_item_count = len(items)
        deduped_items, duplicate_skus = _deduplicate_po_items_by_sku(items)
        extracted['items'] = deduped_items
        extracted['source_item_count'] = source_item_count
        extracted['merged_duplicate_skus'] = sorted(set(duplicate_skus))

        # Reset file pointer so Django can save it
        pdf_file.seek(0)

        po_number = (extracted.get('po_number') or '').strip()
        existing_doc = None
        if po_number:
            existing_doc = PoDocument.objects.filter(extracted_payload__po_number=po_number).order_by('-id').first()

        if existing_doc:
            existing_payload = existing_doc.extracted_payload or {}
            existing_items, _ = _deduplicate_po_items_by_sku(existing_payload.get('items', []))
            merged_items, added_skus, updated_skus, ignored_lines = _merge_po_items_for_existing_po(
                existing_items,
                deduped_items,
            )

            merged_payload = dict(existing_payload)
            merged_payload.update(extracted)
            merged_payload['items'] = merged_items
            merged_payload['source_item_count'] = len(merged_items)
            merged_payload['merged_duplicate_skus'] = sorted(
                set(existing_payload.get('merged_duplicate_skus') or []) | set(duplicate_skus)
            )
            configured_skus = sorted(set(existing_payload.get('new_skus_configured') or []))
            if configured_skus:
                merged_payload['new_skus_configured'] = configured_skus

            existing_doc.po_file = pdf_file
            existing_doc.extracted_payload = merged_payload
            existing_doc.extraction_status = 'processed'
            existing_doc.uploaded_by = request.user
            existing_doc.save(update_fields=['po_file', 'extracted_payload', 'extraction_status', 'uploaded_by'])
            po_doc = existing_doc
        else:
            po_doc = PoDocument.objects.create(
                po_file=pdf_file,
                extracted_payload=extracted,
                extraction_status='processed',
                uploaded_by=request.user,
            )

        sync_result = _sync_repeat_jobs_from_po(po_doc, actor=request.user)
        item_count = len(extracted.get('items', []))
        if existing_doc and ignored_lines:
            preview = ', '.join(
                f"{row['sku']} ({row['qty'] if row['qty'] is not None else '-'})"
                for row in ignored_lines[:8]
            )
            remainder = len(ignored_lines) - 8
            if remainder > 0:
                preview += f" +{remainder} more"
            messages.warning(
                request,
                f"Ignored duplicate line(s) for same PO (same SKU and Qty): {preview}.",
            )

        if extraction_warning:
            messages.warning(
                request,
                f"Partial extraction: {extraction_warning}",
            )
        else:
            if existing_doc:
                final_item_count = len((existing_doc.extracted_payload or {}).get('items', []))
                msg = (
                    f"PO {extracted.get('po_number', '?')} updated. "
                    f"Unique added SKU(s): {len(added_skus)}; corrected SKU(s): {len(updated_skus)}; "
                    f"current PO line count: {final_item_count}."
                )
                if duplicate_skus:
                    msg += f" Duplicate SKU lines merged in upload: {', '.join(sorted(set(duplicate_skus)))}."
                msg += " Same PO + same SKU + same Qty lines are ignored."
            else:
                msg = (
                    f"PO {extracted.get('po_number', '?')} extracted with "
                    f"{item_count} of {expected_count or item_count} line items. Sent to PO Intake queue."
                )
                if duplicate_skus:
                    msg += f" Duplicate SKU lines merged: {', '.join(sorted(set(duplicate_skus)))}."
            if sync_result['created'] or sync_result['updated']:
                msg += (
                    f" Draft planning jobs synced from PO lines: created {sync_result['created']}, "
                    f"updated {sync_result['updated']}."
                )
            if sync_result['missing_recipe']:
                msg += f" SKU(s) pending approved master data: {sync_result['missing_recipe']}."
            messages.success(request, msg)
        return redirect('qc:po_review', doc_id=po_doc.id)

    return render(request, 'planning/po_upload.html')


@login_required
@permission_required('can_edit_jobcard')
def manual_po_entry(request):
    """Create a PO intake record manually without uploading a PDF."""
    if request.method == 'POST':
        po_number = (request.POST.get('po_number') or '').strip()
        pr_number = (request.POST.get('pr_number') or '').strip()

        if not po_number and not pr_number:
            messages.error(request, 'Either a PO Number or a PR Number is required.')
            return redirect('planning:manual_po_entry')

        items = []
        line_indexes = request.POST.getlist('item_index')
        for index in line_indexes:
            sku = (request.POST.get(f'manual_sku_{index}') or '').strip()
            if not sku:
                continue
            quantity = _to_int(request.POST.get(f'manual_quantity_{index}'))
            if quantity is None:
                messages.error(request, f'Quantity must be a valid number for line {index}.')
                return redirect('planning:manual_po_entry')

            unit_cost = _to_decimal(request.POST.get(f'manual_unit_cost_{index}'))
            if unit_cost is None:
                messages.error(request, f'Unit cost must be a valid number for line {index}.')
                return redirect('planning:manual_po_entry')

            net_total = None
            if quantity is not None and unit_cost is not None:
                net_total = unit_cost * Decimal(quantity)

            item = {
                'sku': sku,
                'job_name': (request.POST.get(f'manual_job_name_{index}') or '').strip() or sku,
                'quantity': quantity,
                'unit': (request.POST.get(f'manual_unit_{index}') or '').strip() or 'Pcs',
                'delivery_date': (request.POST.get(f'manual_delivery_date_{index}') or '').strip() or '',
                'unit_cost': _format_decimal_string(unit_cost),
                'net_total': _format_decimal_string(net_total),
            }
            items.append(item)

        if not items:
            messages.error(request, 'At least one PO line is required.')
            return redirect('planning:manual_po_entry')

        sku_keys = [_sku_key(item['sku']) for item in items if item.get('sku')]
        duplicate_sku_keys = [sku for sku in sku_keys if sku_keys.count(sku) > 1]
        if duplicate_sku_keys:
            messages.error(request, 'Duplicate SKUs are not allowed within the same PO. Please remove duplicate lines before saving.')
            return redirect('planning:manual_po_entry')

        supplier_name = (request.POST.get('supplier_name') or '').strip() or 'UTOPIA PRINTING & PACKAGING'
        buyer_name = (request.POST.get('buyer_name') or '').strip() or 'UTOPIA INDUSTRIES (PVT.) LTD.'
        grand_total = sum((Decimal(item['net_total']) if item.get('net_total') is not None else Decimal('0')) for item in items)

        payload = {
            'po_number': po_number,
            'pr_number': pr_number,
            'po_date': (request.POST.get('po_date') or '').strip(),
            'approval_date': (request.POST.get('approval_date') or '').strip(),
            'department': (request.POST.get('department') or '').strip(),
            'delivery_location': (request.POST.get('delivery_location') or '').strip(),
            'supplier_name': supplier_name,
            'buyer_name': buyer_name,
            'grand_total': _format_decimal_string(grand_total),
            'items': items,
            'source_item_count': len(items),
        }

        manual_file = ContentFile(b'', name=f'manual_po_{po_number}_{timezone.now().strftime("%Y%m%d%H%M%S")}.txt')
        po_doc = PoDocument.objects.create(
            po_file=manual_file,
            extracted_payload=payload,
            extraction_status='processed',
            uploaded_by=request.user,
        )

        sync_result = _sync_repeat_jobs_from_po(po_doc, actor=request.user)
        ref_label = f'PO {po_number}' if po_number else f'PR {pr_number}'
        messages.success(request, f'Manual {ref_label} created with {len(items)} line(s).')
        if sync_result['created'] or sync_result['updated']:
            messages.success(
                request,
                f'Draft planning jobs synced from PO lines: created {sync_result["created"]}, updated {sync_result["updated"]}.',
            )
        if sync_result['missing_recipe']:
            messages.warning(request, f'SKU(s) pending approved master data: {sync_result["missing_recipe"]}.')

        return redirect('qc:po_review', doc_id=po_doc.id)

    return render(request, 'planning/manual_po_entry.html')


@login_required
@permission_required('can_edit_jobcard')
def po_review(request, doc_id):
    """Review extracted PO data and create PlanningJob records."""
    po_doc = get_object_or_404(PoDocument, id=doc_id)
    payload = po_doc.extracted_payload or {}
    items = _po_payload_items(payload)
    sku_counts = {}
    for item in payload.get('items', []) or []:
        sku_key = _sku_key(item.get('sku'))
        if sku_key:
            sku_counts[sku_key] = sku_counts.get(sku_key, 0) + 1
    duplicate_skus = [sku for sku, count in sku_counts.items() if count > 1]
    if duplicate_skus:
        messages.error(
            request,
            f'Duplicate SKUs are not allowed in the same PO. Please remove duplicate lines for: {", ".join(sorted(duplicate_skus))}.',
        )
    ignored_skus = sorted({s for s in (payload.get('new_skus_ignored') or []) if s})
    po_number = payload.get('po_number', '')
    configured_new_skus = {_sku_key(sku) for sku in (payload.get('new_skus_configured') or []) if sku}
    recipe_map = _build_recipe_map(items)
    annotated_items, repeat_count, new_count, missing_skus = _annotate_items_with_recipe(items, recipe_map)

    item_sku_keys = {_sku_key(item.get('sku')) for item in annotated_items if item.get('sku')}
    existing_jobs_by_sku = {}
    if po_number and item_sku_keys:
        existing_jobs = PlanningJob.objects.filter(po_number=po_number).order_by('-updated_at', '-id')
        for job in existing_jobs:
            key = _sku_key(job.sku)
            if key in item_sku_keys and key not in existing_jobs_by_sku:
                existing_jobs_by_sku[key] = job

    existing_any_jobs_skus = set()
    if item_sku_keys:
        sku_any_query = Q()
        for sku_key in item_sku_keys:
            sku_any_query |= Q(sku__iexact=sku_key)
        existing_any_jobs_skus = {
            _sku_key(sku)
            for sku in PlanningJob.objects.filter(sku_any_query).values_list('sku', flat=True)
            if sku
        }

    seen_skus_in_payload = set()
    for item in annotated_items:
        sku_key = _sku_key(item.get('sku'))
        is_first_production = bool(
            sku_key
            and sku_key not in existing_any_jobs_skus
            and sku_key not in seen_skus_in_payload
        )
        # Keep existing planning-job forward flags for internal create/update logic.
        item['forward_flag'] = 'New' if is_first_production else 'Repeat'
        item['is_first_production'] = is_first_production
        existing_job = existing_jobs_by_sku.get(sku_key)
        item['existing_job_id'] = existing_job.id if existing_job else None
        item['existing_jc_number'] = existing_job.jc_number if existing_job else ''
        if sku_key:
            seen_skus_in_payload.add(sku_key)

    # Display counts based on approved SKU master data availability, not existing PlanningJob history.
    repeat_count = sum(1 for item in annotated_items if item.get('is_repeat'))
    new_count = sum(1 for item in annotated_items if not item.get('is_repeat'))

    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'ignore':
            sku = (request.POST.get('sku') or '').strip()
            if not sku:
                messages.error(request, 'SKU is required for ignore action.')
                return redirect('qc:po_review', doc_id=po_doc.id)

            ignored = {
                _sku_key(s)
                for s in (payload.get('new_skus_ignored') or [])
                if s
            }
            ignored.add(_sku_key(sku))
            payload['new_skus_ignored'] = sorted(ignored)
            po_doc.extracted_payload = payload
            po_doc.save(update_fields=['extracted_payload'])
            messages.success(request, f'SKU {sku} ignored and removed from PO intake review.')
            return redirect('qc:po_review', doc_id=po_doc.id)

        if action == 'update_po_number':
            manual_po_number = (request.POST.get('manual_po_number') or '').strip()
            if not manual_po_number:
                messages.error(request, 'PO number is required to update the PO intake record.')
                return redirect('qc:po_review', doc_id=po_doc.id)

            payload['po_number'] = manual_po_number
            po_doc.extracted_payload = payload
            po_doc.save(update_fields=['extracted_payload'])
            messages.success(request, f'PO number updated to {manual_po_number}.')
            return redirect('qc:po_review', doc_id=po_doc.id)

        if action == 'add_manual_item':
            sku = (request.POST.get('manual_sku') or '').strip()
            if not sku:
                messages.error(request, 'SKU is required to add a manual PO line.')
                return redirect('qc:po_review', doc_id=po_doc.id)

            sku_key = _sku_key(sku)
            existing_skus = {_sku_key(item.get('sku')) for item in payload.get('items', []) if item.get('sku')}
            if sku_key in existing_skus:
                messages.error(request, f'SKU {sku} is already present on this PO. Duplicate SKUs are not allowed.')
                return redirect('qc:po_review', doc_id=po_doc.id)

            quantity = _to_int(request.POST.get('manual_quantity'))
            if quantity is None:
                messages.error(request, 'Quantity must be a valid number to add a manual PO line.')
                return redirect('qc:po_review', doc_id=po_doc.id)

            unit_cost_value = _to_decimal(request.POST.get('manual_unit_cost'))
            net_total_value = _to_decimal(request.POST.get('manual_net_total'))
            manual_item = {
                'sku': sku,
                'job_name': (request.POST.get('manual_job_name') or '').strip() or sku,
                'quantity': quantity,
                'unit': (request.POST.get('manual_unit') or '').strip() or '',
                'delivery_date': (request.POST.get('manual_delivery_date') or '').strip() or '',
                'unit_cost': _format_decimal_string(unit_cost_value),
                'net_total': _format_decimal_string(net_total_value),
                'print_sheet_size': (request.POST.get('manual_print_sheet_size') or '').strip() or '',
                'purchase_sheet_size': (request.POST.get('manual_purchase_sheet_size') or '').strip() or '',
                'ups': _to_optional_positive_int(request.POST.get('manual_ups')),
                'machine_name': (request.POST.get('manual_machine_name') or '').strip() or '',
            }
            payload['items'] = list(payload.get('items', []))
            payload['items'].append(manual_item)
            po_doc.extracted_payload = payload
            po_doc.save(update_fields=['extracted_payload'])
            messages.success(request, f'Manual PO line for SKU {sku} added.')
            return redirect('qc:po_review', doc_id=po_doc.id)

        if action == 'create_jobs':
            sku_counts = {}
            for item in annotated_items:
                sku_key = _sku_key(item.get('sku'))
                if not sku_key:
                    continue
                sku_counts[sku_key] = sku_counts.get(sku_key, 0) + 1

            duplicate_skus = [sku for sku, count in sku_counts.items() if count > 1]
            if duplicate_skus:
                messages.error(
                    request,
                    'Duplicate SKUs are not allowed in the same PO. Remove duplicate lines before creating jobs.',
                )
                return redirect('qc:po_review', doc_id=po_doc.id)

            with transaction.atomic():
                created_count = 0
                updated_count = 0
                skipped_count = 0
                locked_count = 0
                missing_recipe_count = 0
                po_date_raw = payload.get('po_date')
                po_date = _parse_iso_date(po_date_raw)
                delivery_location = payload.get('delivery_location', '')
                department = payload.get('department', '')

                for item in annotated_items:
                    sku = (item.get('sku') or '').strip()
                    job_name = (item.get('job_name') or '').strip() or sku
                    if not sku:
                        skipped_count += 1
                        continue

                    field_prefix = f"item_{item['line_no']}_"
                    skip_flag = request.POST.get(f"{field_prefix}skip") == '1'

                    if skip_flag:
                        skipped_count += 1
                        continue

                    recipe = recipe_map.get(_sku_key(sku))
                    if not recipe:
                        missing_recipe_count += 1
                        continue

                    sku_key = _sku_key(sku)
                    is_first_production = bool(
                        sku_key
                        and sku_key not in existing_any_jobs_skus
                    )

                    delivery_date = _parse_iso_date(item.get('delivery_date'))
                    plan_date = delivery_date or po_date

                    existing_job = existing_jobs_by_sku.get(sku_key)
                    if existing_job:
                        if _normalize_status(existing_job.status) != 'draft':
                            locked_count += 1
                            continue
                        existing_repeat_flag = (existing_job.repeat_flag or '').strip().lower()
                        if existing_repeat_flag in {'new', 'repeat'}:
                            forward_as_new = existing_repeat_flag == 'new'
                        else:
                            prior_jobs_exist = PlanningJob.objects.filter(sku__iexact=sku).exclude(id=existing_job.id).exists()
                            forward_as_new = not prior_jobs_exist
                        jc_number = existing_job.jc_number
                    else:
                        forward_as_new = is_first_production
                        jc_number = allocate_next_jc_number(plan_date)

                    current_requirement = existing_job.requirement if existing_job else ''

                    qty = item.get('quantity')
                    order_qty = int(qty) if qty is not None else None

                    unit_cost_val = item.get('unit_cost')
                    unit_cost_dec = Decimal(str(unit_cost_val)) if unit_cost_val is not None else None

                    defaults = {
                        'po_number': po_number,
                        'sku': sku,
                        'job_name': recipe.job_name or job_name,
                        'order_qty': order_qty,
                        'department': department,
                        'destination': delivery_location,
                        'delivery_date': delivery_date,
                        'unit_cost': unit_cost_dec if unit_cost_dec is not None else recipe.default_unit_cost,
                        'status': 'draft',
                        'repeat_flag': 'New' if forward_as_new else 'Repeat',
                        'requirement': _sync_new_sku_requirement(current_requirement, forward_as_new),
                        'material': recipe.material,
                        'color_spec': recipe.color_spec,
                        'application': recipe.application,
                        'size_w_mm': recipe.size_w_mm,
                        'size_h_mm': recipe.size_h_mm,
                        'ups': recipe.ups,
                        'print_sheet_size': recipe.print_sheet_size,
                        'purchase_sheet_size': recipe.purchase_sheet_size,
                    }
                    if plan_date:
                        defaults['plan_date'] = plan_date

                    job_obj, created = PlanningJob.objects.update_or_create(
                        jc_number=jc_number,
                        defaults=defaults,
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                    existing_jobs_by_sku[sku_key] = job_obj
                    if sku_key:
                        existing_any_jobs_skus.add(sku_key)

            po_doc.extraction_status = 'processed'
            po_doc.save(update_fields=['extraction_status'])

            if created_count == 0 and updated_count == 0:
                if missing_recipe_count > 0:
                    messages.warning(
                        request,
                        'This PO contains only new SKUs with missing master data. Please configure them in Pending SKUs before sending to planning.',
                    )
                    return redirect('qc:pending_skus')
                messages.warning(
                    request,
                    f'No jobs created. Skipped {skipped_count}, missing-recipe {missing_recipe_count}, locked-skip {locked_count}. Add missing SKU master data from Pending SKUs and run create again.',
                )
                return redirect('qc:pending_skus')

            messages.success(
                request,
                f'Done. Created {created_count}, updated {updated_count}, skipped {skipped_count}, missing-recipe {missing_recipe_count}, locked-skip {locked_count} planning job(s).',
            )
            if missing_recipe_count > 0:
                messages.warning(
                    request,
                    f'{missing_recipe_count} SKU(s) are still pending master data. Open Pending SKUs tab to configure them.',
                )
            return redirect('qc:approval_queue')

    context = {
        'po_doc': po_doc,
        'payload': payload,
        'items': annotated_items,
        'items_json': json.dumps(annotated_items),
        'repeat_count': repeat_count,
        'new_count': new_count,
        'configured_new_count': len(configured_new_skus),
        'missing_skus': missing_skus,
        'ignored_skus': ignored_skus,
        'ignored_count': len(ignored_skus),
    }
    return render(request, 'planning/po_review.html', context)


@login_required
@permission_required('can_edit_jobcard')
@transaction.atomic
def po_new_skus(request, doc_id):
    po_doc = get_object_or_404(PoDocument, id=doc_id)
    payload = po_doc.extracted_payload or {}
    items = _po_payload_items(payload)

    recipe_map = _build_recipe_map(items)
    _, _, _, missing_skus = _annotate_items_with_recipe(items, recipe_map)
    missing_recipe_defaults = {}
    if missing_skus:
        recipe_query = Q()
        for sku in missing_skus:
            recipe_query |= Q(sku__iexact=sku)
        for recipe in SkuRecipe.objects.filter(recipe_query):
            missing_recipe_defaults[recipe.sku.upper()] = recipe
    missing_sku_rows = [
        {
            'sku': sku,
            'recipe': missing_recipe_defaults.get(_sku_key(sku)),
        }
        for sku in missing_skus
    ]

    if request.method == 'POST':
        created_count = 0
        saved_skus = []
        for sku in missing_skus:
            prefix = f"sku_{sku}"
            job_name = (request.POST.get(f"{prefix}_job_name") or '').strip()
            material = (request.POST.get(f"{prefix}_material") or '').strip()
            color_spec = (request.POST.get(f"{prefix}_color_spec") or '').strip()
            application = (request.POST.get(f"{prefix}_application") or '').strip()
            print_sheet_size = (request.POST.get(f"{prefix}_print_sheet_size") or '').strip()
            purchase_sheet_size = (request.POST.get(f"{prefix}_purchase_sheet_size") or '').strip()
            ups = _to_optional_positive_int(request.POST.get(f"{prefix}_ups"))

            unit_cost_raw = (request.POST.get(f"{prefix}_default_unit_cost") or '').strip()
            unit_cost = None
            if unit_cost_raw:
                try:
                    unit_cost = Decimal(unit_cost_raw)
                except InvalidOperation:
                    unit_cost = None

            if not job_name and not material:
                # Keep save requirements simple, but avoid empty recipe rows.
                continue

            SkuRecipe.objects.update_or_create(
                sku=sku,
                defaults={
                    'job_name': job_name,
                    'material': material,
                    'color_spec': color_spec,
                    'application': application,
                    'print_sheet_size': print_sheet_size,
                    'purchase_sheet_size': purchase_sheet_size,
                    'ups': ups,
                    'default_unit_cost': unit_cost,
                    'created_by': request.user,
                    'master_data_status': 'draft',
                    'reviewed_by': None,
                    'reviewed_at': None,
                    'approved_by': None,
                    'approved_at': None,
                },
            )
            created_count += 1
            saved_skus.append(sku)

        if saved_skus:
            configured = set(payload.get('new_skus_configured') or [])
            configured.update(saved_skus)
            payload['new_skus_configured'] = sorted(configured)
            po_doc.extracted_payload = payload
            po_doc.save(update_fields=['extracted_payload'])

        messages.success(
            request,
            f'SKU recipes saved: {created_count}. Planning jobs remain in draft until SKU master approval unlocks refresh.',
        )
        return redirect('qc:po_review', doc_id=po_doc.id)

    return render(
        request,
        'planning/po_new_skus.html',
        {
            'po_doc': po_doc,
            'payload': payload,
            'missing_skus': missing_skus,
            'missing_sku_rows': missing_sku_rows,
            'missing_recipe_defaults': missing_recipe_defaults,
            'example_form': SkuRecipeForm(),
        },
    )
