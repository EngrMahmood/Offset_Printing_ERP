import openpyxl
import csv

# Path to your Google Sheet master data (already downloaded as .xlsx)
wb = openpyxl.load_workbook('planning/docs/planning_2026_master.xlsx', data_only=True)
ws = wb['Master_Data']


# Find header row (row 2)
header_row = 2
headers = [str(ws.cell(row=header_row, column=c).value).strip() if ws.cell(row=header_row, column=c).value is not None else '' for c in range(1, ws.max_column+1)]

# Map for renaming headers to match model fields
header_rename = {
    'Cost': 'Default Unit Cost',
    'Machine Name': 'Machine',
    'Purchase Material': 'Purchase Material',
    'AWC No.': 'AWC No',
}
headers = [header_rename.get(h, h) for h in headers]

# Indices for size columns
def clean_size(val):
    try:
        if val is None or str(val).strip() == '':
            return ''
        ival = int(float(val))
        return str(ival)
    except:
        return str(val)

size_cols = ['Size W mm', 'Size H mm', 'Size W Inch', 'Size H Inch', 'Ups', 'Purchase Sheet ups', 'Daily Demand']
size_idx = [i for i, h in enumerate(headers) if h in size_cols]

# Collect first 20 data rows after header, clean decimals
rows = []
for r in range(header_row+1, header_row+21):
    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
    if any(x is not None and str(x).strip() != '' for x in row):
        # Clean size columns
        for idx in size_idx:
            row[idx] = clean_size(row[idx])
        rows.append(row)

# Write to new .xlsx for upload
dst_wb = openpyxl.Workbook()
dst_ws = dst_wb.active
dst_ws.title = 'Master_Data'
dst_ws.append(headers)
for row in rows:
    dst_ws.append(row)
dst_wb.save('planning/docs/sku_recipe_upload_sample.xlsx')

# Also write to CSV for reference
with open('planning/docs/sku_recipe_upload_sample.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
print('Sample files created.')
