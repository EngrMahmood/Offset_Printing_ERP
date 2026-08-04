"""Review user-updated all_phases_missing_master_data_update.xlsx."""

from __future__ import annotations

import os
import sys
from collections import Counter, defaultdict
from pathlib import Path

import django

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'Offset_ERP.settings')
django.setup()

import openpyxl
from core.models import ProductType
from planning.models import SkuRecipe
from planning.sku_migration_phases import build_sku_phase_map, get_sku_migration_phase

PATH = ROOT / 'all_phases_missing_master_data_update.xlsx'

JP_MAP = {
    'print_and_pack': 'print_and_pack',
    'Print + Pack': 'print_and_pack',
    'cut_and_pack': 'cut_and_pack',
    'Cut & Pack (no printing)': 'cut_and_pack',
}

# Sheets to skip during review/import
IGNORED_SHEETS = {'phase 1 (2)'}

# User-confirmed corrections (override file values when checking)
SKU_OVERRIDES = {
    'stickercartonpillowb2bgtext': {
        'resolved_job_process': 'print_and_pack',
        'resolved_print_passes': '1',
    },
    'wrappaper-nytmfgussetbedpillownavyqueenauup': {
        'resolved_job_process': 'cut_and_pack',
        'resolved_print_passes': '',
    },
}

VALID_PRINT_PASSES = {1, 2, 3, 4}


def norm(value):
    if value is None:
        return ''
    return str(value).strip()


def parse_passes(value):
    if value in (None, ''):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 'INVALID'


def resolved_product_type(row):
    for key in ('corrected_product_type', 'product_type', 'suggested_product_type'):
        if norm(row.get(key)):
            return norm(row[key])
    return ''


def resolved_job_process(row):
    for key in ('corrected_job_process', 'job_process_type', 'job_process_label'):
        raw = norm(row.get(key))
        if raw:
            return JP_MAP.get(raw, raw)
    return ''


def resolved_print_passes(row):
    for key in ('corrected_print_passes', 'print_passes'):
        if key in row and norm(row.get(key)) != '':
            return row.get(key)
    return ''


def load_rows():
    wb = openpyxl.load_workbook(PATH, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().casefold() in IGNORED_SHEETS:
            continue
        if sheet_name.strip().casefold() in {'erp valid values', 'valid values'}:
            continue
        ws = wb[sheet_name]
        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [norm(cell.value) for cell in header_cells]
        if 'sku' not in [h.casefold() for h in headers]:
            continue
        hmap = {h.casefold(): idx for idx, h in enumerate(headers)}
        phase_guess = None
        if sheet_name.lower().startswith('phase'):
            digits = ''.join(ch for ch in sheet_name if ch.isdigit())
            if digits:
                phase_guess = int(digits)
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values:
                continue
            sku_idx = hmap.get('sku')
            sku = norm(values[sku_idx]) if sku_idx is not None and sku_idx < len(values) else ''
            if not sku:
                continue
            row = {'sheet': sheet_name, 'sku': sku}
            for key, idx in hmap.items():
                if idx < len(values):
                    row[key] = norm(values[idx])
            if not row.get('phase') and phase_guess:
                row['phase'] = str(phase_guess)
            row['resolved_product_type'] = resolved_product_type(row)
            row['resolved_job_process'] = resolved_job_process(row)
            row['resolved_print_passes'] = resolved_print_passes(row)
            override = SKU_OVERRIDES.get(row['sku'].casefold())
            if override:
                row.update(override)
            rows.append(row)
    return wb.sheetnames, rows


def main():
    if not PATH.exists():
        print(f'MISSING: {PATH}')
        return 1

    sheet_names, rows = load_rows()
    print('File:', PATH.name)
    print('Sheets:', sheet_names)
    print('Total SKUs:', len(rows))

    erp_product_types = set(ProductType.objects.values_list('name', flat=True))
    phase_map = build_sku_phase_map()

    by_sku = defaultdict(list)
    for row in rows:
        by_sku[row['sku'].casefold()].append(row)
    duplicates = {key: items for key, items in by_sku.items() if len(items) > 1}
    print('Duplicate SKUs:', len(duplicates))

    errors = []
    warnings = []
    ready = []
    for row in rows:
        sku = row['sku']
        issues = []
        product_type = row['resolved_product_type']
        job_process = row['resolved_job_process']
        print_passes = parse_passes(row['resolved_print_passes'])

        recipe = SkuRecipe.objects.filter(sku__iexact=sku).first()
        if not recipe:
            issues.append('SKU not in database')
        else:
            db_phase = get_sku_migration_phase(sku, phase_map)
            row['db_phase'] = db_phase
            file_phase = norm(row.get('phase'))
            if file_phase.isdigit() and int(file_phase) != db_phase:
                warnings.append((sku, f'file phase {file_phase} vs DB phase {db_phase}'))

        if not product_type:
            issues.append('missing product_type')
        elif product_type not in erp_product_types:
            issues.append(f'invalid product_type: {product_type!r}')

        if not job_process:
            issues.append('missing job_process')
        elif job_process not in {'print_and_pack', 'cut_and_pack'}:
            issues.append(f'invalid job_process: {row.get("job_process_type") or row.get("job_process_label")!r}')

        if job_process == 'cut_and_pack':
            if print_passes not in (None, ''):
                issues.append('cut_and_pack must not have print_passes')
        elif job_process == 'print_and_pack':
            if print_passes in (None, ''):
                issues.append('print_and_pack missing print_passes')
            elif print_passes == 'INVALID' or print_passes not in VALID_PRINT_PASSES:
                issues.append(f'invalid print_passes: {row.get("resolved_print_passes")!r}')

        if issues:
            errors.append((sku, issues, row))
        else:
            ready.append(row)

    print('\n=== REVIEW SUMMARY ===')
    print('Ready to apply:', len(ready))
    print('With errors:', len(errors))
    print('Warnings:', len(warnings))

    by_phase = Counter(int(row.get('phase') or row.get('db_phase') or 0) for row in rows)
    print('Rows by phase:', dict(sorted(by_phase.items())))

    print('\nProduct types (ready rows):')
    for name, count in Counter(row['resolved_product_type'] for row in ready).most_common():
        print(f'  {name}: {count}')

    print('\nJob process (ready rows):')
    for name, count in Counter(row['resolved_job_process'] for row in ready).most_common():
        label = 'Print + Pack' if name == 'print_and_pack' else 'Cut & Pack (no printing)'
        print(f'  {label}: {count}')

    print('\nPrint passes (ready rows):')
    for name, count in Counter(row['resolved_print_passes'] for row in ready).most_common():
        print(f'  {name or "(blank)"}: {count}')

    field_changes = Counter()
    for row in ready:
        recipe = SkuRecipe.objects.get(sku__iexact=row['sku'])
        product_type = row['resolved_product_type']
        job_process = row['resolved_job_process']
        print_passes = parse_passes(row['resolved_print_passes'])
        new_passes = print_passes if job_process == 'print_and_pack' else None
        if norm(recipe.product_type) != product_type:
            field_changes['product_type'] += 1
        if (recipe.job_process_type or 'print_and_pack') != job_process:
            field_changes['job_process_type'] += 1
        if recipe.print_passes != new_passes:
            field_changes['print_passes'] += 1
    print('\nFields that would change:', dict(field_changes))

    if errors:
        print('\nErrors:')
        for sku, issues, _row in errors[:30]:
            print(f'  {sku}')
            for issue in issues:
                print(f'    - {issue}')
        if len(errors) > 30:
            print(f'  ... and {len(errors) - 30} more')

    if warnings:
        print('\nWarnings:')
        for sku, message in warnings[:15]:
            print(f'  {sku}: {message}')

    return 0 if not errors else 1


if __name__ == '__main__':
    raise SystemExit(main())
