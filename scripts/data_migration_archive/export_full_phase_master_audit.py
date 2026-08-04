"""Full phase-wise master SKU audit with update-file overlay."""

from __future__ import annotations

import os
import sys
from collections import Counter, defaultdict
from pathlib import Path

import django

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'Offset_ERP.settings')
django.setup()

import openpyxl
from openpyxl.utils import get_column_letter

from core.models import ProductType
from planning.models import SkuRecipe
from planning.sku_migration_phases import build_sku_phase_map, get_sku_migration_phase

UPDATE_PATH = ROOT / 'all_phases_missing_master_data_update.xlsx'
OUT_PATH = ROOT / 'migration_reports' / 'full_phase_master_audit.xlsx'

JOB_PROCESS_LABELS = {
    'print_and_pack': 'Print + Pack',
    'cut_and_pack': 'Cut & Pack (no printing)',
}

VALID_PRINT_PASSES = {1, 2, 3, 4}
VALID_PRODUCT_TYPES = set(ProductType.objects.values_list('name', flat=True))

MASTER_COLUMNS = [
    'phase',
    'sku',
    'master_data_status',
    'legacy_produced',
    'is_active',
    'job_name',
    'material',
    'color_spec',
    'application',
    'product_type',
    'job_process_type',
    'job_process_label',
    'print_passes',
    'machine_name',
    'plate_set_no',
    'size_w_mm',
    'size_h_mm',
    'ups',
    'print_sheet_size',
    'purchase_sheet_size',
    'purchase_sheet_ups',
    'default_unit_cost',
    'daily_demand',
    'awc_no',
    'die_cutting',
    'notes',
    'remarks',
    'lamination_front_and_back',
    'missing_product_type',
    'missing_print_passes',
    'invalid_product_type',
    'in_update_file',
    'update_product_type',
    'update_job_process',
    'update_job_process_label',
    'update_print_passes',
    'fields_will_change',
]


def _norm(value):
    if value is None:
        return ''
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    return str(value).strip()


def _ser_decimal(value):
    if value is None:
        return ''
    return str(value)


def load_update_overlay():
    """Load approved update rows (ignore Phase 1 (2))."""
    import importlib.util

    spec = importlib.util.spec_from_file_location(
        'rev', ROOT / 'scripts' / 'data_migration_archive' / 'review_master_data_update.py',
    )
    rev = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(rev)

    overlay = {}
    _, rows = rev.load_rows()
    for row in rows:
        sku_key = row['sku'].casefold()
        overlay[sku_key] = {
            'sheet': row.get('sheet', ''),
            'product_type': row['resolved_product_type'],
            'job_process': row['resolved_job_process'],
            'job_process_label': JOB_PROCESS_LABELS.get(row['resolved_job_process'], ''),
            'print_passes': rev.parse_passes(row['resolved_print_passes']),
            'print_passes_raw': row['resolved_print_passes'],
        }
    return overlay


def recipe_to_row(recipe, phase, overlay):
    sku_key = recipe.sku.casefold()
    job_process = (recipe.job_process_type or 'print_and_pack').strip()
    product_type = _norm(recipe.product_type)
    update = overlay.get(sku_key)

    missing_pt = 'Y' if not product_type else 'N'
    invalid_pt = 'Y' if product_type and product_type not in VALID_PRODUCT_TYPES else 'N'
    missing_pp = 'N'
    if job_process == 'print_and_pack' and recipe.print_passes is None:
        missing_pp = 'Y'

    update_pt = update['product_type'] if update else ''
    update_jp = update['job_process'] if update else ''
    update_jpl = update['job_process_label'] if update else ''
    update_pp = ''
    if update:
        pp = update['print_passes']
        update_pp = '' if pp in (None, '') else str(pp)

    changes = []
    if update:
        if update_pt and update_pt != product_type:
            changes.append('product_type')
        if update_jp and update_jp != job_process:
            changes.append('job_process_type')
        new_pp = update['print_passes'] if update_jp == 'print_and_pack' else None
        if update_jp == 'cut_and_pack':
            new_pp = None
        if new_pp != recipe.print_passes:
            if update_jp or update_pt or update_pp != '':
                changes.append('print_passes')

    return {
        'phase': phase,
        'sku': recipe.sku,
        'master_data_status': recipe.master_data_status,
        'legacy_produced': 'Y' if recipe.legacy_produced else 'N',
        'is_active': 'Y' if recipe.is_active else 'N',
        'job_name': _norm(recipe.job_name),
        'material': _norm(recipe.material),
        'color_spec': _norm(recipe.color_spec),
        'application': _norm(recipe.application),
        'product_type': product_type,
        'job_process_type': job_process,
        'job_process_label': JOB_PROCESS_LABELS.get(job_process, job_process),
        'print_passes': '' if recipe.print_passes is None else recipe.print_passes,
        'machine_name': _norm(recipe.machine_name),
        'plate_set_no': _norm(recipe.plate_set_no),
        'size_w_mm': recipe.size_w_mm if recipe.size_w_mm is not None else '',
        'size_h_mm': recipe.size_h_mm if recipe.size_h_mm is not None else '',
        'ups': recipe.ups if recipe.ups is not None else '',
        'print_sheet_size': _norm(recipe.print_sheet_size),
        'purchase_sheet_size': _norm(recipe.purchase_sheet_size),
        'purchase_sheet_ups': recipe.purchase_sheet_ups if recipe.purchase_sheet_ups is not None else '',
        'default_unit_cost': _ser_decimal(recipe.default_unit_cost),
        'daily_demand': _ser_decimal(recipe.daily_demand),
        'awc_no': _norm(recipe.awc_no),
        'die_cutting': _norm(recipe.die_cutting),
        'notes': _norm(recipe.notes),
        'remarks': _norm(recipe.remarks),
        'lamination_front_and_back': 'Y' if recipe.lamination_front_and_back else 'N',
        'missing_product_type': missing_pt,
        'missing_print_passes': missing_pp,
        'invalid_product_type': invalid_pt,
        'in_update_file': 'Y' if update else 'N',
        'update_product_type': update_pt,
        'update_job_process': update_jp,
        'update_job_process_label': update_jpl,
        'update_print_passes': update_pp,
        'fields_will_change': ', '.join(changes),
    }


def build_phase_rows(phase_map, overlay, phase_num):
    rows = []
    for recipe in SkuRecipe.objects.all().order_by('sku').iterator(chunk_size=500):
        if get_sku_migration_phase(recipe.sku, phase_map=phase_map) != phase_num:
            continue
        rows.append(recipe_to_row(recipe, phase_num, overlay))
    return rows


def write_sheet(ws, columns, rows):
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, '') for col in columns])
    for idx, col in enumerate(columns, 1):
        width = 55 if col == 'sku' else min(42, max(len(col) + 2, 12))
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_summary(phase_rows, overlay):
    lines = [
        ['Section', 'Metric', 'Value'],
        ['', '', ''],
        ['Phase totals', 'Phase 1 SKUs (never in planning/production)', len(phase_rows[1])],
        ['Phase totals', 'Phase 2 SKUs (planning only)', len(phase_rows[2])],
        ['Phase totals', 'Phase 3 SKUs (released to production)', len(phase_rows[3])],
        ['Phase totals', 'Grand total SKUs', sum(len(v) for v in phase_rows.values())],
        ['', '', ''],
    ]
    for phase in (1, 2, 3):
        subset = phase_rows[phase]
        lines.append([f'Phase {phase}', 'In update file', sum(1 for r in subset if r['in_update_file'] == 'Y')])
        lines.append([f'Phase {phase}', 'Will change from update file', sum(1 for r in subset if r['fields_will_change'])])
        lines.append([f'Phase {phase}', 'Missing product_type', sum(1 for r in subset if r['missing_product_type'] == 'Y')])
        lines.append([f'Phase {phase}', 'Invalid product_type', sum(1 for r in subset if r['invalid_product_type'] == 'Y')])
        lines.append([f'Phase {phase}', 'Missing print_passes (print_and_pack)', sum(1 for r in subset if r['missing_print_passes'] == 'Y')])
        lines.append([f'Phase {phase}', '', ''])

    lines.append(['Update file', 'Total rows loaded (excl Phase 1 (2))', len(overlay)])
    lines.append(['Update file', 'Source', UPDATE_PATH.name])
    lines.append(['', '', ''])
    lines.append(['ERP valid product types', ', '.join(sorted(VALID_PRODUCT_TYPES)), ''])
    lines.append(['ERP job process', 'Print + Pack / Cut & Pack (no printing)', ''])
    lines.append(['ERP print passes', '1, 2, 3, 4 (blank for Cut & Pack)', ''])
    return lines


def main():
    OUT_PATH.parent.mkdir(exist_ok=True)
    phase_map = build_sku_phase_map()
    overlay = load_update_overlay()

    phase_rows = {
        1: build_phase_rows(phase_map, overlay, 1),
        2: build_phase_rows(phase_map, overlay, 2),
        3: build_phase_rows(phase_map, overlay, 3),
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet('Audit Summary')
    for line in build_summary(phase_rows, overlay):
        ws_summary.append(line)
    ws_summary.column_dimensions['A'].width = 28
    ws_summary.column_dimensions['B'].width = 42
    ws_summary.column_dimensions['C'].width = 20

    for phase in (1, 2, 3):
        ws = wb.create_sheet(f'Phase {phase} All SKUs')
        write_sheet(ws, MASTER_COLUMNS, phase_rows[phase])

    ws_updates = wb.create_sheet('Update File Changes')
    update_columns = [
        'phase', 'sku', 'update_product_type', 'update_job_process_label',
        'update_print_passes', 'current_product_type', 'current_job_process_label',
        'current_print_passes', 'fields_will_change',
    ]
    update_rows = []
    for phase in (1, 2, 3):
        for row in phase_rows[phase]:
            if row['in_update_file'] != 'Y':
                continue
            update_rows.append({
                'phase': phase,
                'sku': row['sku'],
                'update_product_type': row['update_product_type'],
                'update_job_process_label': row['update_job_process_label'],
                'update_print_passes': row['update_print_passes'],
                'current_product_type': row['product_type'],
                'current_job_process_label': row['job_process_label'],
                'current_print_passes': row['print_passes'],
                'fields_will_change': row['fields_will_change'],
            })
    write_sheet(ws_updates, update_columns, update_rows)

    ws_fields = wb.create_sheet('Master Field List')
    field_docs = [
        ['field_name', 'description', 'required_for_approval'],
        ['sku', 'Unique SKU code', 'Yes'],
        ['job_name', 'Job / product name', 'Yes'],
        ['material', 'Material type', 'Yes'],
        ['product_type', 'ERP dropdown: A4 Rim, Label, Insert Card, etc.', 'Yes'],
        ['job_process_type', 'print_and_pack or cut_and_pack', 'Yes'],
        ['print_passes', '1-4 for Print + Pack; blank for Cut & Pack', 'Yes (Print + Pack)'],
        ['color_spec', 'Print color', 'Yes (Print + Pack)'],
        ['application', 'Application / lamination', 'Yes'],
        ['machine_name', 'Machine', 'No'],
        ['plate_set_no', 'Plate set number', 'Yes (Print + Pack)'],
        ['size_w_mm', 'Width mm', 'Yes'],
        ['size_h_mm', 'Height mm', 'Yes'],
        ['ups', 'UPS', 'Yes'],
        ['print_sheet_size', 'Print sheet size', 'Yes'],
        ['purchase_sheet_size', 'Purchase sheet size', 'Yes'],
        ['purchase_sheet_ups', 'Purchase sheet UPS', 'Yes'],
        ['awc_no', 'AWC #', 'Yes'],
        ['die_cutting', 'Die cutting', 'Yes'],
        ['default_unit_cost', 'Default unit cost', 'No'],
        ['daily_demand', 'Daily demand', 'No'],
        ['notes', 'Notes / remarks from legacy sheet', 'No'],
        ['remarks', 'Internal remarks', 'No'],
        ['legacy_produced', 'From legacy Google Sheet', 'No'],
        ['master_data_status', 'draft / pending_review / reviewed / approved', 'No'],
    ]
    for line in field_docs:
        ws_fields.append(line)

    wb.save(OUT_PATH)

    print(f'Wrote {OUT_PATH}')
    print('Phase 1 all SKUs:', len(phase_rows[1]))
    print('Phase 2 all SKUs:', len(phase_rows[2]))
    print('Phase 3 all SKUs:', len(phase_rows[3]))
    print('Update file overlays:', len(overlay))
    for phase in (1, 2, 3):
        subset = phase_rows[phase]
        print(
            f'Phase {phase}: in_update_file={sum(1 for r in subset if r["in_update_file"]=="Y")} '
            f'will_change={sum(1 for r in subset if r["fields_will_change"])}'
        )


if __name__ == '__main__':
    main()
