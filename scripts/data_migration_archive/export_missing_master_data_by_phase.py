"""Export phase-wise CSV/Excel of SKUs with missing or invalid master data."""

from __future__ import annotations

import csv
import os
import sys
from pathlib import Path

import django

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'Offset_ERP.settings')
django.setup()

from core.models import ProductType
from planning.models import PlanningJob, SkuRecipe
from planning.sku_migration_phases import build_sku_phase_map, get_sku_migration_phase
from planning.sku_sheet_import import (
    _normalize_job_process_type,
    _sheet_row_get,
    dedupe_sheet_rows,
    parse_sheet_rows,
)

SHEET_PATH = ROOT / 'Offset_Master Data_Planning_ERP 1.xlsb'
OUT_DIR = ROOT / 'migration_reports'

# ERP UI dropdowns (from ProductType master + SkuRecipe form)
ERP_JOB_PROCESS_LABELS = {
    'print_and_pack': 'Print + Pack',
    'cut_and_pack': 'Cut & Pack (no printing)',
}
ERP_PRINT_PASSES = {1, 2, 3}

# Prefix inference used invalid names — map to valid ERP product types
INVALID_TO_ERP_PRODUCT_TYPE = {
    'Care Label': 'Label',
    'Size Label': 'Label',
    'Law Label': 'Label',
    'Warning Label': 'Label',
    'Importer Label': 'Label',
    'Importer Sleeve': 'Label',
    'Wrap Paper': 'Stiffeners',
    'Belly Band': 'Insert Card',
    'Catalog': 'Report Books',
    'Inner Paper': 'Insert Card',
    'Color Box': 'Stiffeners',
    'Poly Bag': 'Label',
    'Identification Card': 'Visiting/Employee Card',
    'Header Card': 'Insert Card',
}

COLUMNS = [
    'phase',
    'sku',
    'master_data_status',
    'legacy_produced',
    'in_legacy_sheet',
    'active_jobs',
    'job_process_type',
    'job_process_label',
    'product_type',
    'print_passes',
    'missing_product_type',
    'invalid_product_type',
    'suggested_product_type',
    'missing_print_passes',
    'invalid_print_passes',
    'passes_required',
    'missing_job_process',
    'job_process_sheet_mismatch',
    'sheet_product_type',
    'sheet_job_process',
    'sheet_no_of_passes',
    'material',
    'job_name',
]


def _norm(value):
    return str(value or '').strip()


def _erp_product_types():
    return set(ProductType.objects.values_list('name', flat=True))


def _sheet_passes_display(raw):
    if raw is None or _norm(raw) in ('', '0', '0.0'):
        return ''
    try:
        return str(int(float(raw)))
    except (TypeError, ValueError):
        return _norm(raw)


def _suggest_product_type(current, valid_types):
    if not current:
        return ''
    if current in valid_types:
        return current
    if current in INVALID_TO_ERP_PRODUCT_TYPE:
        return INVALID_TO_ERP_PRODUCT_TYPE[current]
    upper = current.upper()
    if 'LABEL' in upper:
        return 'Label'
    if 'INSERT' in upper or 'CARD' in upper:
        return 'Insert Card'
    if 'STICKER' in upper:
        return 'Sticker'
    if 'FLUFF' in upper:
        return 'Fluffing Instruction'
    if 'RIM' in upper or 'A4' in upper:
        return 'A4 Rim'
    if 'LETTER' in upper:
        return 'Letter Head'
    if 'REPORT' in upper or 'BOOK' in upper:
        return 'Report Books'
    if 'STIFF' in upper:
        return 'Stiffeners'
    if 'VISIT' in upper or 'EMPLOYEE' in upper:
        return 'Visiting/Employee Card'
    return ''


def load_sheet_index():
    if not SHEET_PATH.exists():
        return {}
    rows = dedupe_sheet_rows(parse_sheet_rows(SHEET_PATH))
    return {_norm(_sheet_row_get(row, 'SKU')).casefold(): row for row in rows if _norm(_sheet_row_get(row, 'SKU'))}


def build_gap_rows(phase_map, sheet_by_sku, valid_product_types):
    rows = []
    for recipe in SkuRecipe.objects.all().order_by('sku').iterator(chunk_size=500):
        phase = get_sku_migration_phase(recipe.sku, phase_map=phase_map)
        sheet_row = sheet_by_sku.get(recipe.sku.casefold())
        in_sheet = 'Y' if sheet_row else 'N'

        sheet_pt = _norm(_sheet_row_get(sheet_row, 'Product Type')) if sheet_row else ''
        sheet_jp = _norm(_sheet_row_get(sheet_row, 'Job Process')) if sheet_row else ''
        sheet_jp_norm = _normalize_job_process_type(sheet_jp) if sheet_jp else ''
        sheet_passes = _sheet_passes_display(_sheet_row_get(sheet_row, 'No. of Passes')) if sheet_row else ''

        job_process = (recipe.job_process_type or '').strip()
        job_process_display = job_process or 'print_and_pack'
        if not job_process:
            job_process = 'print_and_pack'
        job_process_label = ERP_JOB_PROCESS_LABELS.get(job_process, job_process)

        product_type = _norm(recipe.product_type)
        missing_pt = 'Y' if not product_type else 'N'
        invalid_pt = 'Y' if product_type and product_type not in valid_product_types else 'N'
        suggested_pt = _suggest_product_type(product_type, valid_product_types) if missing_pt == 'Y' or invalid_pt == 'Y' else product_type

        missing_jp = 'Y' if not _norm(recipe.job_process_type) else 'N'
        jp_mismatch = 'N'
        if sheet_jp_norm and sheet_jp_norm != job_process:
            jp_mismatch = 'Y'

        passes_required = 'Y' if job_process == 'print_and_pack' else 'N'
        missing_pp = 'N'
        invalid_pp = 'N'
        if job_process == 'print_and_pack' and recipe.print_passes is None:
            missing_pp = 'Y'
        elif job_process == 'cut_and_pack' and recipe.print_passes is not None:
            invalid_pp = 'Y'
        elif recipe.print_passes is not None and recipe.print_passes not in ERP_PRINT_PASSES:
            invalid_pp = 'Y'

        has_gap = (
            missing_pt == 'Y' or invalid_pt == 'Y' or missing_pp == 'Y' or invalid_pp == 'Y'
            or missing_jp == 'Y' or jp_mismatch == 'Y'
        )
        if not has_gap:
            continue

        active_jobs = PlanningJob.objects.filter(sku__iexact=recipe.sku, is_active=True).count()
        rows.append({
            'phase': phase,
            'sku': recipe.sku,
            'master_data_status': recipe.master_data_status,
            'legacy_produced': 'Y' if recipe.legacy_produced else 'N',
            'in_legacy_sheet': in_sheet,
            'active_jobs': active_jobs,
            'job_process_type': job_process,
            'job_process_label': job_process_label,
            'product_type': product_type,
            'print_passes': '' if recipe.print_passes is None else recipe.print_passes,
            'missing_product_type': missing_pt,
            'invalid_product_type': invalid_pt,
            'suggested_product_type': suggested_pt,
            'missing_print_passes': missing_pp,
            'invalid_print_passes': invalid_pp,
            'passes_required': passes_required,
            'missing_job_process': missing_jp,
            'job_process_sheet_mismatch': jp_mismatch,
            'sheet_product_type': sheet_pt,
            'sheet_job_process': sheet_jp,
            'sheet_no_of_passes': sheet_passes,
            'material': _norm(recipe.material),
            'job_name': _norm(recipe.job_name),
        })
    return rows


def write_csv(path, rows):
    with open(path, 'w', newline='', encoding='utf-8-sig') as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path, rows):
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Gaps'
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row[col] for col in COLUMNS])
    for idx, col in enumerate(COLUMNS, 1):
        width = 55 if col == 'sku' else min(40, max(len(col), 12))
        ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def main():
    OUT_DIR.mkdir(exist_ok=True)
    valid_product_types = _erp_product_types()
    phase_map = build_sku_phase_map()
    sheet_by_sku = load_sheet_index()
    all_gaps = build_gap_rows(phase_map, sheet_by_sku, valid_product_types)

    # Reference sheet for valid ERP values
    ref_path = OUT_DIR / 'erp_valid_master_values.csv'
    with open(ref_path, 'w', newline='', encoding='utf-8-sig') as handle:
        writer = csv.writer(handle)
        writer.writerow(['field', 'valid_value'])
        for name in sorted(valid_product_types):
            writer.writerow(['product_type', name])
        for key, label in ERP_JOB_PROCESS_LABELS.items():
            writer.writerow(['job_process_type', key])
            writer.writerow(['job_process_label', label])
        for p in sorted(ERP_PRINT_PASSES):
            writer.writerow(['print_passes', p])
        writer.writerow(['print_passes', '(blank/null for Cut & Pack only)'])

    for phase in (1, 2, 3):
        phase_rows = [row for row in all_gaps if row['phase'] == phase]
        write_csv(OUT_DIR / f'phase{phase}_missing_master_data.csv', phase_rows)
        write_xlsx(OUT_DIR / f'phase{phase}_missing_master_data.xlsx', phase_rows)

    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet('ERP Valid Values').append(['field', 'valid_value'])
    ws_ref = wb['ERP Valid Values']
    with open(ref_path, encoding='utf-8-sig') as handle:
        for row in csv.reader(handle):
            ws_ref.append(row)
    for phase in (1, 2, 3):
        phase_rows = [row for row in all_gaps if row['phase'] == phase]
        ws = wb.create_sheet(title=f'Phase {phase}')
        ws.append(COLUMNS)
        for row in phase_rows:
            ws.append([row[col] for col in COLUMNS])
    wb.save(OUT_DIR / 'all_phases_missing_master_data.xlsx')

    print('ERP valid product types:', sorted(valid_product_types))
    print(f'Exported to {OUT_DIR}')
    for phase in (1, 2, 3):
        subset = [r for r in all_gaps if r['phase'] == phase]
        pt = sum(1 for r in subset if r['missing_product_type'] == 'Y')
        inv_pt = sum(1 for r in subset if r['invalid_product_type'] == 'Y')
        pp = sum(1 for r in subset if r['missing_print_passes'] == 'Y')
        inv_pp = sum(1 for r in subset if r['invalid_print_passes'] == 'Y')
        jp = sum(1 for r in subset if r['missing_job_process'] == 'Y' or r['job_process_sheet_mismatch'] == 'Y')
        print(
            f'Phase {phase}: {len(subset)} rows | blank product_type={pt} | '
            f'invalid product_type={inv_pt} | missing passes={pp} | invalid passes={inv_pp} | job_process={jp}'
        )


if __name__ == '__main__':
    main()
