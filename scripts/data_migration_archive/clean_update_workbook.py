"""Remove obsolete Phase 1 (2) sheet from the user update workbook."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PATH = ROOT / 'all_phases_missing_master_data_update.xlsx'
SHEET_TO_REMOVE = 'Phase 1 (2)'


def main() -> int:
    if not PATH.exists():
        print(f'MISSING: {PATH}')
        return 1

    wb = openpyxl.load_workbook(PATH)
    if SHEET_TO_REMOVE not in wb.sheetnames:
        print(f'Already clean — no {SHEET_TO_REMOVE!r} sheet.')
        print('Sheets:', wb.sheetnames)
        return 0

    del wb[SHEET_TO_REMOVE]
    try:
        wb.save(PATH)
        print(f'Removed {SHEET_TO_REMOVE!r} from {PATH.name}')
    except PermissionError:
        backup = ROOT / 'all_phases_missing_master_data_update_clean.xlsx'
        wb.save(backup)
        print(f'Original file is open — close Excel and re-run, or use: {backup.name}')
        return 1

    print('Sheets:', wb.sheetnames)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
