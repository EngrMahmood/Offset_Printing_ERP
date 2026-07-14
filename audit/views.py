from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.utils import timezone
from core.models import JobCard, Production, Dispatch, JobCardWipStatus
from printing_plates.models import PlateRequest
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import io

def get_audit_data():
    """
    Scans the database in real-time to find gaps/missing entries.
    """
    # 1. Pre-Press Gaps (missing plate set number or AWC number on approved/released print JobCards)
    pre_press_gaps = []
    pre_press_qs = JobCard.objects.filter(
        is_active=True,
        is_print_job=True,
        status__in=['production_approved', 'released', 'in_production']
    ).select_related('planning_job', 'machine_name')
    for jc in pre_press_qs:
        gaps = []
        if not jc.plate_set_no:
            gaps.append("Missing Plate Set No")
        awc = jc.planning_job.awc_no_display if jc.planning_job else ''
        if not awc:
            gaps.append("Missing AWC No")
            
        if gaps:
            pre_press_gaps.append({
                'job_card_id': jc.id,
                'job_card_no': jc.job_card_no,
                'sku': jc.SKU,
                'gaps': ", ".join(gaps),
                'machine': jc.machine_name_display or '-',
                'status': jc.workflow_status_label,
            })

    # 2. Press Gaps (Printing)
    press_gaps = []
    # Print JobCards in released/in_production with no printing entries
    press_qs = JobCard.objects.filter(
        is_active=True,
        is_print_job=True,
        status__in=['released', 'in_production']
    ).select_related('planning_job', 'machine_name')
    for jc in press_qs:
        printing_exists = jc.productions.filter(is_active=True, entry_type='printing').exists()
        packing_exists = jc.productions.filter(is_active=True, entry_type='packing').exists()
        
        if not printing_exists:
            if packing_exists:
                # CRITICAL: Packing entries logged but Printing entries are missing!
                press_gaps.append({
                    'job_card_id': jc.id,
                    'job_card_no': jc.job_card_no,
                    'sku': jc.SKU,
                    'gap_type': 'critical',
                    'gaps': "CRITICAL: Packing entries logged but Printing (Press) entries are missing!",
                    'machine': jc.machine_name_display or '-',
                    'status': jc.workflow_status_label,
                })
            else:
                press_gaps.append({
                    'job_card_id': jc.id,
                    'job_card_no': jc.job_card_no,
                    'sku': jc.SKU,
                    'gap_type': 'warning',
                    'gaps': "Released/In Production but has no Printing entries logged.",
                    'machine': jc.machine_name_display or '-',
                    'status': jc.workflow_status_label,
                })

    # Printing entries missing operator or supervisor
    prod_printing_missing = Production.objects.filter(
        is_active=True,
        entry_type='printing',
        job_card__is_active=True
    ).filter(Q(operator__isnull=True) | Q(supervisor__isnull=True)).select_related('job_card', 'operator', 'supervisor')
    for p in prod_printing_missing:
        missing = []
        if not p.operator:
            missing.append("Operator")
        if not p.supervisor:
            missing.append("Supervisor")
        press_gaps.append({
            'job_card_id': p.job_card.id,
            'job_card_no': p.job_card.job_card_no,
            'sku': p.job_card.SKU,
            'gap_type': 'info',
            'gaps': f"Printing Entry (ID: {p.id}, Date: {p.date}) missing {', '.join(missing)}.",
            'machine': p.machine or p.job_card.machine_name_display or '-',
            'status': p.job_card.workflow_status_label,
        })

    # 3. Post-Press Gaps (Packing/Sorting)
    post_press_gaps = []
    # Print JobCards in production/completed with no packing production entries
    post_press_qs = JobCard.objects.filter(
        is_active=True,
        is_print_job=True,
        status__in=['in_production', 'completed']
    ).select_related('planning_job', 'machine_name')
    for jc in post_press_qs:
        packing_exists = jc.productions.filter(is_active=True, entry_type='packing').exists()
        if not packing_exists:
            post_press_gaps.append({
                'job_card_id': jc.id,
                'job_card_no': jc.job_card_no,
                'sku': jc.SKU,
                'gaps': "In Production/Completed but has no Packing entries logged.",
                'machine': jc.machine_name_display or '-',
                'status': jc.workflow_status_label,
            })
            
    # Packing entries missing sorter
    prod_packing_missing = Production.objects.filter(
        is_active=True,
        entry_type='packing',
        job_card__is_active=True,
        sorter__isnull=True
    ).select_related('job_card')
    for p in prod_packing_missing:
        post_press_gaps.append({
            'job_card_id': p.job_card.id,
            'job_card_no': p.job_card.job_card_no,
            'sku': p.job_card.SKU,
            'gaps': f"Packing Entry (ID: {p.id}, Date: {p.date}) missing Sorter assignment.",
            'machine': p.machine or p.job_card.machine_name_display or '-',
            'status': p.job_card.workflow_status_label,
        })

    # 4. QC Gaps (JobCards waiting in QC)
    qc_gaps = []
    qc_qs = JobCard.objects.filter(
        is_active=True,
        status='pending_qc'
    ).select_related('planning_job', 'machine_name')
    for jc in qc_qs:
        qc_gaps.append({
            'job_card_id': jc.id,
            'job_card_no': jc.job_card_no,
            'sku': jc.SKU,
            'gaps': "Job Card is pending QC Review / approval.",
            'machine': jc.machine_name_display or '-',
            'status': jc.workflow_status_label,
        })

    # 5. Dispatch Gaps (Completed but no/incomplete dispatch)
    dispatch_gaps = []
    completed_jcs = JobCard.objects.filter(
        is_active=True,
        status='completed'
    ).select_related('planning_job', 'machine_name')
    for jc in completed_jcs:
        total_dispatched = jc.dispatch_set.filter(is_active=True).aggregate(total=Sum('dispatch_qty'))['total'] or 0
        if total_dispatched == 0:
            dispatch_gaps.append({
                'job_card_id': jc.id,
                'job_card_no': jc.job_card_no,
                'sku': jc.SKU,
                'gaps': f"Completed but has no Dispatch logs. Order Qty: {jc.order_qty}.",
                'machine': jc.machine_name_display or '-',
                'status': jc.workflow_status_label,
            })
        elif total_dispatched < jc.order_qty:
            dispatch_gaps.append({
                'job_card_id': jc.id,
                'job_card_no': jc.job_card_no,
                'sku': jc.SKU,
                'gaps': f"Incomplete Dispatch: dispatched {total_dispatched} of {jc.order_qty} (Short: {jc.order_qty - total_dispatched}).",
                'machine': jc.machine_name_display or '-',
                'status': jc.workflow_status_label,
            })
            
    return {
        'pre_press': pre_press_gaps,
        'press': press_gaps,
        'post_press': post_press_gaps,
        'qc': qc_gaps,
        'dispatch': dispatch_gaps,
        'total_gaps_count': len(pre_press_gaps) + len(press_gaps) + len(post_press_gaps) + len(qc_gaps) + len(dispatch_gaps)
    }

def get_pending_work_data(query_q=''):
    """
    Computes JobCard-wise pending work process-wise.
    """
    # Base JobCard Query
    job_cards = JobCard.objects.filter(is_active=True).select_related(
        'planning_job', 'machine_name', 'production_wip_status__status'
    )
    if query_q:
        job_cards = job_cards.filter(
            Q(job_card_no__icontains=query_q)
            | Q(SKU__icontains=query_q)
            | Q(destination__icontains=query_q)
            | Q(planning_job__job_name__icontains=query_q)
        )

    # 1. Pre-Press Pending (approved/released job cards where plates are not ready)
    pre_press_pending = []
    pre_press_candidates = job_cards.filter(
        is_print_job=True,
        status__in=['planning_approved', 'production_approved', 'released']
    )
    for jc in pre_press_candidates:
        # Check if plate request is available/ready for production
        plate_requests = jc.plate_requests.all()
        is_ready = any(pr.status == 'available_for_production' for pr in plate_requests)
        
        if not is_ready:
            # Determine plate status label
            status_text = "No Request"
            if plate_requests.exists():
                latest_pr = plate_requests.order_by('-updated_at').first()
                status_text = latest_pr.get_status_display()
                
            wip_status = jc.production_wip_status.status.name if hasattr(jc, 'production_wip_status') else 'No Status'
            pre_press_pending.append({
                'job_card': jc,
                'plate_status': status_text,
                'wip_status': wip_status,
            })

    # 2. Press Pending (released/in_production where printing sheets are not complete)
    press_pending = []
    press_candidates = job_cards.filter(
        is_print_job=True,
        status__in=['released', 'in_production']
    )
    for jc in press_candidates:
        sheets_required = jc.total_sheet_quantity_display or 0
        if sheets_required == 0:
            continue
            
        printed_sheets = jc.productions.filter(is_active=True, entry_type='printing').aggregate(total=Sum('output_sheets'))['total'] or 0
        if printed_sheets < sheets_required:
            balance = sheets_required - printed_sheets
            progress = round((printed_sheets / sheets_required) * 100, 1) if sheets_required > 0 else 0
            wip_status = jc.production_wip_status.status.name if hasattr(jc, 'production_wip_status') else 'No Status'
            
            press_pending.append({
                'job_card': jc,
                'sheets_required': sheets_required,
                'printed_sheets': printed_sheets,
                'balance': balance,
                'progress': progress,
                'wip_status': wip_status,
            })

    # 3. Post-Press Pending (in_production where packed quantity is less than order quantity)
    post_press_pending = []
    post_press_candidates = job_cards.filter(
        status__in=['released', 'in_production']
    )
    for jc in post_press_candidates:
        order_qty = jc.order_qty or 0
        if order_qty == 0:
            continue
            
        packed_qty = jc.productions.filter(is_active=True, entry_type='packing').aggregate(total=Sum('packing_qty'))['total'] or 0
        if packed_qty < order_qty:
            balance = order_qty - packed_qty
            progress = round((packed_qty / order_qty) * 100, 1) if order_qty > 0 else 0
            wip_status = jc.production_wip_status.status.name if hasattr(jc, 'production_wip_status') else 'No Status'
            
            post_press_pending.append({
                'job_card': jc,
                'order_qty': order_qty,
                'packed_qty': packed_qty,
                'balance': balance,
                'progress': progress,
                'wip_status': wip_status,
            })

    return {
        'pre_press': pre_press_pending,
        'press': press_pending,
        'post_press': post_press_pending,
    }


@login_required
def dashboard(request):
    q = (request.GET.get('q') or '').strip()
    active_tab = request.GET.get('tab', 'audit')
    
    audit_data = get_audit_data()
    pending_work = get_pending_work_data(query_q=q)
    
    context = {
        'audit': audit_data,
        'pending': pending_work,
        'q': q,
        'active_tab': active_tab,
        'today': timezone.now().date(),
    }
    return render(request, 'audit/dashboard.html', context)


@login_required
def export_excel(request):
    """
    Generates a high-fidelity Excel report for audit and pending work.
    """
    wb = openpyxl.Workbook()
    
    # 1. Audit Gaps Sheet
    ws1 = wb.active
    ws1.title = "Audit Gaps"
    
    audit = get_audit_data()
    
    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F6EA7", end_color="2F6EA7", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border_side = Side(style='thin', color='DDDDDD')
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Title Block
    ws1.merge_cells("A1:E1")
    ws1["A1"] = "ERP Operational Entry Audit Report"
    ws1["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4F7A")
    ws1["A1"].alignment = left_align
    ws1.row_dimensions[1].height = 25
    
    headers1 = ["Team", "Job Card No", "SKU", "Audit Alert Description", "Target Machine"]
    ws1.append([]) # empty row
    ws1.append(headers1)
    ws1.row_dimensions[3].height = 20
    
    # Style header row
    for col_idx in range(1, 6):
        cell = ws1.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = left_align
        cell.border = thin_border
        
    all_gaps = []
    for row in audit['pre_press']:
        all_gaps.append(["Pre-Press", row['job_card_no'], row['sku'], row['gaps'], row['machine']])
    for row in audit['press']:
        all_gaps.append(["Press (Printing)", row['job_card_no'], row['sku'], row['gaps'], row['machine']])
    for row in audit['post_press']:
        all_gaps.append(["Post-Press", row['job_card_no'], row['sku'], row['gaps'], row['machine']])
    for row in audit['qc']:
        all_gaps.append(["QC", row['job_card_no'], row['sku'], row['gaps'], row['machine']])
    for row in audit['dispatch']:
        all_gaps.append(["Dispatch", row['job_card_no'], row['sku'], row['gaps'], row['machine']])
        
    for row in all_gaps:
        ws1.append(row)
        curr_row = ws1.max_row
        for col_idx in range(1, 6):
            cell = ws1.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if col_idx in [1, 2, 5]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Auto-fit columns
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 2. Pending Work Sheet
    ws2 = wb.create_sheet(title="Pending Process Work")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "Process-wise Outstanding Production Work"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="2F8D4E")
    ws2["A1"].alignment = left_align
    ws2.row_dimensions[1].height = 25
    
    headers2 = ["Process Tab", "Job Card No", "SKU", "Job Name", "Order/Sheets Qty", "Completed Qty", "Outstanding Balance", "Current WIP Status"]
    ws2.append([])
    ws2.append(headers2)
    ws2.row_dimensions[3].height = 20
    
    # Style headers
    for col_idx in range(1, 9):
        cell = ws2.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = PatternFill(start_color="2F8D4E", end_color="2F8D4E", fill_type="solid")
        cell.alignment = left_align
        cell.border = thin_border
        
    pending = get_pending_work_data()
    pending_rows = []
    for item in pending['pre_press']:
        pending_rows.append([
            "Pre-Press (Plates)", 
            item['job_card'].job_card_no, 
            item['job_card'].SKU, 
            item['job_card'].planning_job.job_name if item['job_card'].planning_job else '-',
            "-", "-", item['plate_status'], item['wip_status']
        ])
    for item in pending['press']:
        pending_rows.append([
            "Press (Printing)", 
            item['job_card'].job_card_no, 
            item['job_card'].SKU, 
            item['job_card'].planning_job.job_name if item['job_card'].planning_job else '-',
            item['sheets_required'], item['printed_sheets'], item['balance'], item['wip_status']
        ])
    for item in pending['post_press']:
        pending_rows.append([
            "Post-Press (Packing)", 
            item['job_card'].job_card_no, 
            item['job_card'].SKU, 
            item['job_card'].planning_job.job_name if item['job_card'].planning_job else '-',
            item['order_qty'], item['packed_qty'], item['balance'], item['wip_status']
        ])
        
    for row in pending_rows:
        ws2.append(row)
        curr_row = ws2.max_row
        for col_idx in range(1, 9):
            cell = ws2.cell(row=curr_row, column=col_idx)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if col_idx in [1, 2, 8]:
                cell.alignment = center_align
            elif col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = left_align
                
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Return Workbook
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=Operational_Audit_Pending_{timezone.now().date()}.xlsx"
    wb.save(response)
    return response


@login_required
def export_pdf(request):
    """
    Generates a structured PDF report of the operational pending work.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1D2735'),
        spaceAfter=12
    )
    section_style = ParagraphStyle(
        'SecTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor('#2F6EA7'),
        spaceBefore=10,
        spaceAfter=6
    )
    body_style = ParagraphStyle(
        'BodyText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10
    )
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white
    )
    
    story = []
    
    # 1. Document Title
    story.append(Paragraph("Offset ERP — Production Pending Process Work Report", title_style))
    story.append(Spacer(1, 10))
    
    pending = get_pending_work_data()
    
    # 2. Pre-Press Pending Section
    story.append(Paragraph("1. Outstanding Pre-Press Work (Plates)", section_style))
    pre_press_data = [[
        Paragraph("Job Card No", header_style),
        Paragraph("SKU", header_style),
        Paragraph("Job Name", header_style),
        Paragraph("Machine", header_style),
        Paragraph("WIP Status", header_style),
        Paragraph("Plates Status", header_style),
    ]]
    for item in pending['pre_press'][:50]: # limit rows to avoid bloating
        pre_press_data.append([
            Paragraph(item['job_card'].job_card_no, body_style),
            Paragraph(item['job_card'].SKU, body_style),
            Paragraph(item['job_card'].planning_job.job_name if item['job_card'].planning_job else '-', body_style),
            Paragraph(item['job_card'].machine_name_display or '-', body_style),
            Paragraph(item['wip_status'], body_style),
            Paragraph(item['plate_status'], body_style),
        ])
    t1 = Table(pre_press_data, colWidths=[80, 100, 220, 100, 100, 100])
    t1.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2F6EA7')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    story.append(t1)
    story.append(Spacer(1, 15))
    
    # 3. Press Pending Section
    story.append(Paragraph("2. Outstanding Press Work (Printing)", section_style))
    press_data = [[
        Paragraph("Job Card No", header_style),
        Paragraph("SKU", header_style),
        Paragraph("Job Name", header_style),
        Paragraph("Sheets Req.", header_style),
        Paragraph("Printed", header_style),
        Paragraph("Balance", header_style),
        Paragraph("WIP Status", header_style),
    ]]
    for item in pending['press'][:50]:
        press_data.append([
            Paragraph(item['job_card'].job_card_no, body_style),
            Paragraph(item['job_card'].SKU, body_style),
            Paragraph(item['job_card'].planning_job.job_name if item['job_card'].planning_job else '-', body_style),
            Paragraph(str(item['sheets_required']), body_style),
            Paragraph(str(item['printed_sheets']), body_style),
            Paragraph(str(item['balance']), body_style),
            Paragraph(item['wip_status'], body_style),
        ])
    t2 = Table(press_data, colWidths=[80, 100, 220, 80, 80, 80, 80])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2F8D4E')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FBFD')])
    ]))
    story.append(t2)
    story.append(Spacer(1, 15))

    # 4. Post-Press Pending Section
    story.append(Paragraph("3. Outstanding Post-Press Work (Packing)", section_style))
    post_press_data = [[
        Paragraph("Job Card No", header_style),
        Paragraph("SKU", header_style),
        Paragraph("Job Name", header_style),
        Paragraph("Order Qty", header_style),
        Paragraph("Packed", header_style),
        Paragraph("Balance", header_style),
        Paragraph("WIP Status", header_style),
    ]]
    for item in pending['post_press'][:50]:
        post_press_data.append([
            Paragraph(item['job_card'].job_card_no, body_style),
            Paragraph(item['job_card'].SKU, body_style),
            Paragraph(item['job_card'].planning_job.job_name if item['job_card'].planning_job else '-', body_style),
            Paragraph(str(item['order_qty']), body_style),
            Paragraph(str(item['packed_qty']), body_style),
            Paragraph(str(item['balance']), body_style),
            Paragraph(item['wip_status'], body_style),
        ])
    t3 = Table(post_press_data, colWidths=[80, 100, 220, 80, 80, 80, 80])
    t3.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#C08128')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#FFFDFB')])
    ]))
    story.append(t3)

    doc.build(story)
    
    # Return PDF response
    pdf_val = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f"attachment; filename=Outstanding_Operational_Work_{timezone.now().date()}.pdf"
    response.write(pdf_val)
    return response
