import csv
import io
import logging
import re
from decimal import Decimal, InvalidOperation
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.db.models.functions import Upper
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from core.views import permission_required
from core.models import JobCard
from core.jobcard_service import job_card_queue_queryset
from planning import views as planning_views
from planning.forms import SkuRecipeForm
from planning.models import PlanningJob, PoDocument, SkuRecipe
from workflow.services import (
    _annotate_items_with_recipe,
    _build_cost_mismatch_note,
    _build_recipe_map,
    _collect_pending_sku_rows,
    _format_display_qty,
    _missing_required_master_fields,
    _warning_master_fields,
    _normalize_application_input,
    _normalize_color_spec_input,
    _normalize_status,
    _po_payload_items,
    _sku_key,
    _sync_new_jobs_for_approved_sku,
    _to_decimal,
    _to_optional_decimal,
    _to_optional_positive_int,
    _sanitize_po_payload_items,
    _user_is_admin,
    execute_job_card_action,
)

logger = logging.getLogger(__name__)

# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_view_approval_queue')
def approval_queue(request):
    """JobCard approval queue for QC inspection only."""
    qc_jobs = job_card_queue_queryset('qc')
    queue_q = (request.GET.get('q') or '').strip()

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        job_card_id = (request.POST.get('job_card_id') or '').strip()
        reason = (request.POST.get('reason') or request.POST.get('change_reason') or '').strip()

        def _get_job_card():
            if not job_card_id:
                raise ValueError('Job Card is required.')
            return get_object_or_404(JobCard, pk=job_card_id, is_active=True)

        def _sync_status(job_card, transition_name):
            return execute_job_card_action(job_card, transition_name, actor=request.user, reason=reason)

        profile = getattr(request.user, 'profile', None)
        if action == 'delete_job_card':
            if not _user_is_admin(request.user):
                messages.error(request, 'Only administrators can delete job cards from the approval queue.')
                return redirect('qc:approval_queue')
            job_card = _get_job_card()
            job_card.delete()
            messages.success(request, f'Job Card {job_card.job_card_no} deleted.')
            return redirect('qc:approval_queue')

        if action == 'bulk_delete_job_cards':
            if not _user_is_admin(request.user):
                messages.error(request, 'Only administrators can delete job cards from the approval queue.')
                return redirect('qc:approval_queue')
            job_card_ids = request.POST.getlist('job_card_ids')
            if not job_card_ids:
                messages.error(request, 'Select at least one Job Card to delete.')
                return redirect('qc:approval_queue')
            delete_qs = JobCard.objects.filter(pk__in=job_card_ids, is_active=True)
            deleted_count = delete_qs.count()
            if deleted_count:
                delete_qs.delete()
                messages.success(request, f'Deleted {deleted_count} Job Card(s).')
            else:
                messages.error(request, 'No valid Job Cards were selected for deletion.')
            return redirect('qc:approval_queue')

        if action in {'approve_qc', 'reject_qc'}:
            if action == 'reject_qc' and not reason:
                messages.error(request, 'Rejection reason is required.')
                return redirect('qc:approval_queue')
            job_card = _get_job_card()
            if not profile or not profile.can_approve_qc():
                messages.error(request, 'You do not have permission to approve QC jobs.')
                return redirect('qc:approval_queue')

            try:
                _sync_status(job_card, action)
            except ValidationError as exc:
                error_dict = getattr(exc, 'message_dict', None)
                if error_dict:
                    for field_errors in error_dict.values():
                        for error_message in field_errors:
                            messages.error(request, error_message)
                else:
                    messages.error(request, str(exc))
                return redirect('qc:approval_queue')

            messages.success(request, f'Job Card {job_card.job_card_no} moved successfully.')
            return redirect('qc:approval_queue')

    if queue_q:
        queue_filter = (
            Q(job_card_no__icontains=queue_q)
            | Q(PO_No__icontains=queue_q)
            | Q(SKU__icontains=queue_q)
            | Q(planning_job__po_number__icontains=queue_q)
            | Q(planning_job__sku__icontains=queue_q)
        )
        qc_jobs = qc_jobs.filter(queue_filter)

    qc_jobs = qc_jobs.order_by('-updated_at', '-id')[:300]

    if queue_q and not qc_jobs:
        logger.info('QC queue search had no matches for query="%s"', queue_q)

    # Dashboard counters (display only)
    profile = getattr(request.user, 'profile', None)
    pending_qc_jobs_count = qc_jobs.count()

    context = {
        'qc_jobs': qc_jobs,
        'queue_q': queue_q,
        'pending_qc_jobs_count': pending_qc_jobs_count,
        'user_can_approve_qc': profile.can_approve_qc() if profile else False,
    }
    context['can_admin_actions'] = _user_is_admin(request.user)
    return render(request, 'qc/approval_queue.html', context)


@login_required
@permission_required('can_view_approval_queue')
def approval_history(request):
    """Read-only approval history for SKU and Job Card decisions."""
    status_filter = (request.GET.get('status') or '').strip().lower()
    q = (request.GET.get('q') or '').strip()

    job_cards = JobCard.objects.filter(is_active=True).order_by('-updated_at', '-id')
    if status_filter in {'qc_rejected', 'pm_rejected', 'released', 'qc_approved', 'production_approved'}:
        job_cards = job_cards.filter(status=status_filter)

    if q:
        # Filter at the DB level across the whole table, then cap — previously this
        # sliced to the 400 most-recently-updated job cards *before* filtering, so a
        # real match outside that window was unreachable no matter how specific the
        # search was, not just truncated.
        job_cards = list(
            job_cards.filter(
                Q(job_card_no__icontains=q) | Q(PO_No__icontains=q) | Q(SKU__icontains=q)
            )[:400]
        )
    else:
        job_cards = list(job_cards[:200])

    context = {
        'rows': job_cards,
        'q': q,
        'status_filter': status_filter,
    }
    return render(request, 'qc/approval_history.html', context)


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_edit_jobcard')
@transaction.atomic
def planning_job_status_update(request, job_id):
    if request.method != 'POST':
        return redirect('planning:job_detail', job_id=job_id)

    job = get_object_or_404(PlanningJob, id=job_id)
    current_status = _normalize_status(job.status)
    transition = (request.POST.get('transition') or '').strip()
    reason = (request.POST.get('reason') or request.POST.get('change_reason') or '').strip()
    next_url = (request.POST.get('next') or '').strip()

    transitions = {
        'approve': ('pending_qc', 'qc_approved'),
        'approve_qc': ('pending_qc', 'qc_approved'),
        'reject': ('pending_qc', 'draft'),
        'reject_qc': ('pending_qc', 'draft'),
        'release': ('qc_approved', 'released'),
        'unlock': ('qc_approved', 'pending_qc'),
    }
    if transition not in transitions:
        messages.error(request, 'Unknown status transition request.')
        return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    required_from, target_status = transitions[transition]
    if current_status == target_status:
        messages.info(request, f'Job already in {target_status} status.')
        return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    if required_from and current_status != required_from:
        messages.error(request, f'Transition not allowed from {current_status} to {target_status}.')
        return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    if transition in {'reject', 'reject_qc'} and not reason:
        messages.error(request, 'Rejection reason is required.')
        return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    if target_status in {'qc_approved', 'released'}:
        job_card = JobCard.objects.filter(planning_job=job, is_active=True).order_by('-updated_at', '-id').first()
        if not job_card:
            messages.error(request, 'Job Card is missing. Submit to QC from Planning first.')
            return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)
        try:
            job_card.full_clean()
        except ValidationError as exc:
            error_dict = getattr(exc, 'message_dict', None)
            if error_dict:
                for field_errors in error_dict.values():
                    for error_message in field_errors:
                        messages.error(request, error_message)
            else:
                messages.error(request, str(exc))
            return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    if target_status in {'qc_approved', 'released', 'in_production', 'completed'}:
        validation_errors = job.qc_validation_errors()
        if validation_errors:
            for field_errors in validation_errors.values():
                for error_message in field_errors if isinstance(field_errors, list) else [field_errors]:
                    messages.error(request, error_message)
            return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    job.status = target_status
    if target_status in {'released'}:
        job.issued_to_production = True
    if target_status in {'draft', 'pending_qc'}:
        job.issued_to_production = False

    try:
        job.save(update_fields=['status', 'issued_to_production', 'updated_at'])
    except ValidationError as exc:
        error_dict = getattr(exc, 'message_dict', None)
        if error_dict:
            for field_errors in error_dict.values():
                for error_message in field_errors:
                    messages.error(request, error_message)
        else:
            messages.error(request, str(exc))
        return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)

    messages.success(request, f'Job status updated: {current_status} -> {target_status}.')
    return redirect(next_url) if next_url else redirect('planning:job_detail', job_id=job.id)


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_edit_jobcard')
def po_review(request, doc_id):
    return planning_views.po_review(request, doc_id)


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_edit_jobcard')
def po_new_skus(request, doc_id):
    return planning_views.po_new_skus(request, doc_id)


@login_required
@permission_required('can_edit_jobcard')
def planner_pending_skus_redirect(request):
    target = reverse('planning:pending_skus')
    query = request.META.get('QUERY_STRING', '')
    return redirect(f'{target}?{query}' if query else target)


@login_required
@permission_required('can_edit_jobcard')
def planner_pending_skus_ignored_redirect(request):
    target = reverse('planning:pending_skus_ignored')
    query = request.META.get('QUERY_STRING', '')
    return redirect(f'{target}?{query}' if query else target)


@login_required
@permission_required('can_view_sku_master_review_queue')
def planner_pending_sku_master_entry_redirect(request):
    target = reverse('planning:pending_sku_master_entry')
    query = request.META.get('QUERY_STRING', '')
    return redirect(f'{target}?{query}' if query else target)


@login_required
@permission_required('can_view_sku_master_review_queue')
@transaction.atomic
def master_sku_review_queue(request):
    profile = getattr(request.user, 'profile', None)
    user_can_approve_qc = profile.can_approve_sku_master_review() if profile else False

    po_filter = (request.POST.get('return_po') or request.GET.get('po') or '').strip()
    q = (request.POST.get('return_q') or request.GET.get('q') or '').strip()

    def _redirect_queue():
        params = {}
        if po_filter:
            params['po'] = po_filter
        if q:
            params['q'] = q
        url = reverse('qc:master_review')
        return redirect(f'{url}?{urlencode(params)}' if params else url)

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        sku = (request.POST.get('sku') or '').strip()
        rejection_comment = (request.POST.get('rejection_comment') or '').strip()

        if action not in {'approve', 'back_to_draft'}:
            messages.error(request, 'Invalid action.')
            return _redirect_queue()

        if not user_can_approve_qc:
            messages.error(request, 'You do not have permission to review SKU master records.')
            return _redirect_queue()

        recipe = SkuRecipe.objects.filter(sku__iexact=sku, is_active=True).first()
        if not recipe:
            messages.error(request, f'SKU recipe for {sku} was not found.')
            return _redirect_queue()

        current_status = (recipe.master_data_status or 'draft').lower()

        if action == 'approve':
            if current_status not in {'pending_review', 'reviewed'}:
                messages.error(request, f'SKU {sku} is not in QC review queue.')
                return _redirect_queue()

            missing_required = _missing_required_master_fields(recipe, recipe.job_name, allow_missing_plate_set_no=True)
            if missing_required:
                messages.error(
                    request,
                    f'SKU {sku} cannot be approved. Missing required fields: {", ".join(missing_required)}.',
                )
                return _redirect_queue()

            if current_status == 'pending_review':
                recipe.reviewed_by = request.user
                recipe.reviewed_at = timezone.now()

            recipe.master_data_status = 'approved'
            recipe.approved_by = request.user
            recipe.approved_at = timezone.now()
            recipe.rejection_comment = ''
            recipe.last_rejected_by = None
            recipe.last_rejected_at = None
            recipe.save(update_fields=[
                'master_data_status', 'reviewed_by', 'reviewed_at',
                'approved_by', 'approved_at', 'rejection_comment',
                'last_rejected_by', 'last_rejected_at', 'updated_at',
            ])

            sync_result = _sync_new_jobs_for_approved_sku(sku, actor=request.user)
            approval_warnings = _warning_master_fields(recipe, recipe.job_name)
            notice = ''
            if approval_warnings:
                notice = (
                    f' Notice: {", ".join(approval_warnings)} not set yet '
                    '(usually assigned at plate making).'
                )
            messages.success(
                request,
                f'SKU {sku} approved. Planning jobs refreshed: updated {sync_result["updated"]}, locked {sync_result["locked"]}, missing draft jobs {sync_result.get("missing_jobs", 0)}.{notice}',
            )
            return _redirect_queue()

        if current_status == 'draft':
            messages.info(request, f'SKU {sku} is already in Draft.')
            return _redirect_queue()

        if not rejection_comment:
            messages.error(request, 'Reason is required to send back this SKU.')
            return _redirect_queue()

        recipe.master_data_status = 'draft'
        recipe.reviewed_by = None
        recipe.reviewed_at = None
        recipe.approved_by = None
        recipe.approved_at = None
        recipe.rejection_comment = rejection_comment
        recipe.last_rejected_by = request.user
        recipe.last_rejected_at = timezone.now()
        recipe.save(update_fields=[
            'master_data_status', 'reviewed_by', 'reviewed_at',
            'approved_by', 'approved_at', 'rejection_comment',
            'last_rejected_by', 'last_rejected_at', 'updated_at',
        ])
        messages.warning(request, f'SKU {sku} sent back to Draft. Reason: {rejection_comment}')
        return _redirect_queue()

    po_docs = PoDocument.objects.exclude(extracted_payload__isnull=True).order_by('-created_at')[:400]
    deduped_docs = []
    seen_po_numbers = set()
    for doc in po_docs:
        payload = doc.extracted_payload or {}
        po_number = (payload.get('po_number') or '').strip().upper()
        if po_number:
            if po_number in seen_po_numbers:
                continue
            seen_po_numbers.add(po_number)
        deduped_docs.append(doc)

    all_rows = _collect_pending_sku_rows(deduped_docs[:200])
    sku_values = sorted({row['sku'] for row in all_rows if row.get('sku')})

    recipes_by_sku = {}
    if sku_values:
        # A per-SKU Q(...) |= chain blows past SQLite's expression-tree depth
        # limit (1000) once enough PO documents are scanned — Upper(...) IN
        # (...) has no such ceiling.
        sku_keys = {sku.strip().upper() for sku in sku_values}
        recipes = SkuRecipe.objects.annotate(sku_upper=Upper('sku')).filter(sku_upper__in=sku_keys, is_active=True)
        recipes_by_sku = {recipe.sku.upper(): recipe for recipe in recipes}

    review_rows = []
    po_summary_map = {}
    seen_review_skus = set()
    for row in all_rows:
        recipe = recipes_by_sku.get(_sku_key(row.get('sku')))
        if not recipe:
            continue

        recipe_status = (recipe.master_data_status or '').strip().lower()
        if recipe_status not in {'pending_review', 'reviewed'}:
            continue

        row['recipe'] = recipe
        row['recipe_status'] = recipe_status
        row['missing_required_fields'] = _missing_required_master_fields(
            recipe, row.get('job_name') or '', allow_missing_plate_set_no=True,
        )
        row['warning_master_fields'] = _warning_master_fields(recipe, row.get('job_name') or '')

        po_number = row.get('po_number') or '-'
        po_summary_map.setdefault(po_number, {'po_number': po_number, 'count': 0})
        po_summary_map[po_number]['count'] += 1
        review_rows.append(row)
        seen_review_skus.add(_sku_key(row.get('sku')))

    # Include submitted master records even when they are not tied to the latest PO pending rows.
    submitted_recipes = SkuRecipe.objects.filter(
        is_active=True,
        master_data_status__in=['pending_review', 'reviewed'],
    ).order_by('sku')
    for recipe in submitted_recipes:
        sku_key = _sku_key(recipe.sku)
        if sku_key in seen_review_skus:
            continue
        review_rows.append({
            'po_doc_id': None,
            'po_number': '-',
            'sku': recipe.sku,
            'job_name': recipe.job_name,
            'qty': '-',
            'delivery_date': '-',
            'recipe': recipe,
            'recipe_status': recipe.master_data_status,
            'missing_required_fields': _missing_required_master_fields(
                recipe, recipe.job_name, allow_missing_plate_set_no=True,
            ),
            'warning_master_fields': _warning_master_fields(recipe, recipe.job_name),
        })
        seen_review_skus.add(sku_key)

    if po_filter:
        review_rows = [row for row in review_rows if (row.get('po_number') or '') == po_filter]
    if q:
        q_upper = q.upper()
        review_rows = [
            row for row in review_rows
            if q_upper in (row.get('sku') or '').upper()
            or q_upper in (row.get('po_number') or '').upper()
            or q_upper in (row.get('job_name') or '').upper()
        ]

    review_rows.sort(key=lambda row: (row['po_number'], row['sku']))

    context = {
        'review_rows': review_rows,
        'review_count': len(review_rows),
        'po_summary': sorted(po_summary_map.values(), key=lambda x: x['po_number']),
        'po_filter': po_filter,
        'q': q,
        'user_can_approve_qc': user_can_approve_qc,
    }
    return render(request, 'qc/master_sku_review_queue.html', context)


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_edit_jobcard')
@transaction.atomic
def pending_skus(request):
    """Central queue of SKUs that are still missing in SKU Recipe master data."""
    is_admin_user = _user_is_admin(request.user)
    if request.method == 'POST':
        action = (request.POST.get('action') or 'save').strip()
        sku = (request.POST.get('sku') or '').strip()
        po_doc_id = request.POST.get('po_doc_id')
        return_po = (request.POST.get('return_po') or '').strip()
        return_q = (request.POST.get('return_q') or '').strip()

        def _redirect_pending():
            params = {}
            if return_po:
                params['po'] = return_po
            if return_q:
                params['q'] = return_q
            url = reverse('qc:pending_skus')
            return redirect(f'{url}?{urlencode(params)}' if params else url)

        if action in {'delete', 'archive'}:
            recipe_id = request.POST.get('recipe_id')
            try:
                recipe_obj = SkuRecipe.objects.get(id=int(recipe_id))
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
                return _redirect_pending()

            if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                messages.error(request, 'Approved records can only be managed by admin users.')
                return _redirect_pending()

            if action == 'delete':
                recipe_obj.delete()
                messages.success(request, f'SKU recipe {recipe_obj.sku} deleted.')
            else:
                recipe_obj.is_active = False
                recipe_obj.archived_by = request.user
                recipe_obj.archived_at = timezone.now()
                recipe_obj.archive_reason = (request.POST.get('archive_reason') or '').strip()
                recipe_obj.save(update_fields=['is_active', 'archived_by', 'archived_at', 'archive_reason', 'updated_at'])
                messages.success(request, f'SKU recipe {recipe_obj.sku} archived.')

            return _redirect_pending()

        if action == 'ignore':
            try:
                po_doc = PoDocument.objects.get(id=int(po_doc_id)) if po_doc_id else None
            except (TypeError, ValueError, PoDocument.DoesNotExist):
                po_doc = None

            if not po_doc or not sku:
                messages.error(request, 'Invalid SKU or PO reference for ignore action.')
                return _redirect_pending()

            payload = po_doc.extracted_payload or {}
            ignored = { _sku_key(s) for s in (payload.get('new_skus_ignored') or []) if s }
            ignored.add(_sku_key(sku))
            payload['new_skus_ignored'] = sorted(ignored)
            po_doc.extracted_payload = payload
            po_doc.save(update_fields=['extracted_payload'])

            po_number = (payload.get('po_number') or '').strip()
            if po_number:
                PlanningJob.objects.filter(
                    po_number__iexact=po_number,
                    sku__iexact=sku,
                    status__iexact='draft',
                    is_active=True,
                ).update(is_active=False, updated_at=timezone.now())

            messages.success(request, f'SKU {sku} will be ignored and removed from pending processing.')
            return _redirect_pending()

        if not sku:
            messages.error(request, 'SKU is required.')
            return _redirect_pending()

        if action in {'submit_review', 'review', 'approve', 'back_to_draft'}:
            recipe = SkuRecipe.objects.filter(sku__iexact=sku).first()
            if not recipe:
                messages.error(request, f'No SKU recipe found for {sku}. Save recipe data first.')
                return _redirect_pending()

            current_status = (recipe.master_data_status or 'draft').lower()
            rejection_comment = (request.POST.get('rejection_comment') or '').strip()

            if action == 'submit_review':
                if current_status != 'draft':
                    messages.error(request, f'SKU {sku} can only move to Pending Review from Draft.')
                    return _redirect_pending()
                recipe.master_data_status = 'pending_review'
                recipe.reviewed_by = None
                recipe.reviewed_at = None
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.save(update_fields=['master_data_status', 'reviewed_by', 'reviewed_at', 'approved_by', 'approved_at', 'updated_at'])
                messages.success(request, f'SKU {sku} moved to Pending Review.')
                return _redirect_pending()

            if action == 'review':
                if current_status != 'pending_review':
                    messages.error(request, f'SKU {sku} can only move to Reviewed from Pending Review.')
                    return _redirect_pending()
                recipe.master_data_status = 'reviewed'
                recipe.reviewed_by = request.user
                recipe.reviewed_at = timezone.now()
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.rejection_comment = ''
                recipe.last_rejected_by = None
                recipe.last_rejected_at = None
                recipe.save(update_fields=[
                    'master_data_status', 'reviewed_by', 'reviewed_at',
                    'approved_by', 'approved_at', 'rejection_comment',
                    'last_rejected_by', 'last_rejected_at', 'updated_at',
                ])
                messages.success(request, f'SKU {sku} reviewed successfully.')
                return _redirect_pending()

            if action == 'approve':
                if current_status not in {'pending_review', 'reviewed'}:
                    messages.error(request, f'SKU {sku} can only be Approved from Pending Review or Reviewed status.')
                    return _redirect_pending()
                missing_required = _missing_required_master_fields(recipe, allow_missing_plate_set_no=True)
                if missing_required:
                    messages.error(
                        request,
                        f'SKU {sku} cannot be approved. Missing required master data: {", ".join(missing_required)}.',
                    )
                    return _redirect_pending()
                if current_status == 'pending_review':
                    recipe.reviewed_by = request.user
                    recipe.reviewed_at = timezone.now()
                recipe.master_data_status = 'approved'
                recipe.approved_by = request.user
                recipe.approved_at = timezone.now()
                recipe.rejection_comment = ''
                recipe.last_rejected_by = None
                recipe.last_rejected_at = None
                recipe.save(update_fields=[
                    'master_data_status', 'reviewed_by', 'reviewed_at',
                    'approved_by', 'approved_at', 'rejection_comment',
                    'last_rejected_by', 'last_rejected_at', 'updated_at',
                ])
                sync_result = _sync_new_jobs_for_approved_sku(sku, actor=request.user)
                approval_warnings = _warning_master_fields(recipe)
                notice = ''
                if approval_warnings:
                    notice = (
                        f' Notice: {", ".join(approval_warnings)} not set yet '
                        '(usually assigned at plate making).'
                    )
                messages.success(
                    request,
                    f'SKU {sku} approved for master data usage. Planning jobs refreshed: updated {sync_result["updated"]}, locked {sync_result["locked"]}, missing draft jobs {sync_result.get("missing_jobs", 0)}.{notice}',
                )
                return _redirect_pending()

            if action == 'back_to_draft':
                if current_status == 'draft':
                    messages.info(request, f'SKU {sku} is already in Draft.')
                    return _redirect_pending()
                if not rejection_comment:
                    messages.error(request, 'Reason is required to send SKU back to Draft.')
                    return _redirect_pending()
                recipe.master_data_status = 'draft'
                recipe.reviewed_by = None
                recipe.reviewed_at = None
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.rejection_comment = rejection_comment
                recipe.last_rejected_by = request.user
                recipe.last_rejected_at = timezone.now()
                recipe.save(update_fields=[
                    'master_data_status', 'reviewed_by', 'reviewed_at',
                    'approved_by', 'approved_at', 'rejection_comment',
                    'last_rejected_by', 'last_rejected_at', 'updated_at',
                ])
                messages.warning(request, f'SKU {sku} sent back to Draft. Reason: {rejection_comment}')
                return _redirect_pending()

        job_name = (request.POST.get('job_name') or '').strip()
        material = (request.POST.get('material') or '').strip()
        color_spec = (request.POST.get('color_spec') or '').strip()
        application = (request.POST.get('application') or '').strip()
        department = (request.POST.get('department') or '').strip()
        print_sheet_size = (request.POST.get('print_sheet_size') or '').strip()
        purchase_sheet_size = (request.POST.get('purchase_sheet_size') or '').strip()
        purchase_sheet_ups = _to_optional_decimal(request.POST.get('purchase_sheet_ups'))
        ups = _to_optional_decimal(request.POST.get('ups'))
        purchase_material = (request.POST.get('purchase_material') or '').strip()
        daily_demand = _to_optional_decimal(request.POST.get('daily_demand'))
        awc_no = (request.POST.get('awc_no') or '').strip()
        plate_set_no = (request.POST.get('plate_set_no') or '').strip()
        die_cutting = (request.POST.get('die_cutting') or '').strip()

        unit_cost_raw = (request.POST.get('default_unit_cost') or '').strip()
        unit_cost = None
        if unit_cost_raw:
            try:
                unit_cost = Decimal(unit_cost_raw)
            except InvalidOperation:
                unit_cost = None

        if not job_name and not material:
            messages.error(request, 'Please enter at least Job Name or Material before saving.')
            return _redirect_pending()

        SkuRecipe.objects.update_or_create(
            sku=sku,
            defaults={
                'job_name': job_name,
                'material': material,
                'color_spec': color_spec,
                'application': application,
                'department': department,
                'print_sheet_size': print_sheet_size,
                'purchase_sheet_size': purchase_sheet_size,
                'purchase_sheet_ups': purchase_sheet_ups,
                'ups': ups,
                'purchase_material': purchase_material,
                'default_unit_cost': unit_cost,
                'daily_demand': daily_demand,
                'awc_no': awc_no,
                'plate_set_no': plate_set_no,
                'die_cutting': die_cutting,
                'created_by': request.user,
                'master_data_status': 'draft',
                'reviewed_by': None,
                'reviewed_at': None,
                'approved_by': None,
                'approved_at': None,
            },
        )

        if po_doc_id:
            try:
                po_doc = PoDocument.objects.filter(id=int(po_doc_id)).first()
            except (TypeError, ValueError):
                po_doc = None
            if po_doc:
                payload = po_doc.extracted_payload or {}
                configured = set(payload.get('new_skus_configured') or [])
                configured.add(sku)
                payload['new_skus_configured'] = sorted(configured)
                po_doc.extracted_payload = payload
                po_doc.save(update_fields=['extracted_payload'])

        messages.success(request, f'SKU recipe saved for {sku}.')
        return _redirect_pending()

    po_filter = (request.GET.get('po') or '').strip()
    q = (request.GET.get('q') or '').strip()

    po_docs = PoDocument.objects.exclude(extracted_payload__isnull=True).order_by('-created_at')[:400]
    deduped_docs = []
    seen_po_numbers = set()
    for doc in po_docs:
        payload = doc.extracted_payload or {}
        po_number = (payload.get('po_number') or '').strip().upper()
        if po_number:
            if po_number in seen_po_numbers:
                continue
            seen_po_numbers.add(po_number)
        deduped_docs.append(doc)
    po_docs = deduped_docs[:200]
    all_pending_rows = _collect_pending_sku_rows(po_docs)

    po_summary_map = {}
    for row in all_pending_rows:
        po_key = row.get('po_number') or '-'
        current = po_summary_map.get(po_key)
        if not current:
            po_summary_map[po_key] = {
                'po_number': po_key,
                'count': 1,
                'po_doc_id': row.get('po_doc_id'),
            }
        else:
            current['count'] += 1

    pending_rows = all_pending_rows
    if po_filter:
        pending_rows = [row for row in pending_rows if (row.get('po_number') or '') == po_filter]
    if q:
        q_upper = q.upper()
        pending_rows = [
            row
            for row in pending_rows
            if q_upper in (row.get('sku') or '').upper()
            or q_upper in (row.get('po_number') or '').upper()
            or q_upper in (row.get('job_name') or '').upper()
        ]

    sku_values = sorted({row['sku'] for row in pending_rows if row.get('sku')})
    recipes_by_sku = {}
    if sku_values:
        # A per-SKU Q(...) |= chain blows past SQLite's expression-tree depth
        # limit (1000) once enough PO documents are scanned — Upper(...) IN
        # (...) has no such ceiling.
        sku_keys = {sku.strip().upper() for sku in sku_values}
        recipes = SkuRecipe.objects.annotate(sku_upper=Upper('sku')).filter(sku_upper__in=sku_keys)
        recipes_by_sku = {recipe.sku.upper(): recipe for recipe in recipes}

    for row in pending_rows:
        recipe = recipes_by_sku.get(_sku_key(row.get('sku')))
        row['recipe'] = recipe
        row['recipe_status'] = recipe.master_data_status if recipe else 'missing'
        row['missing_required_fields'] = _missing_required_master_fields(
            recipe, row.get('job_name') or '', allow_missing_plate_set_no=True,
        )
        row['warning_master_fields'] = _warning_master_fields(recipe, row.get('job_name') or '')

    pending_rows.sort(key=lambda row: (row['po_number'], row['sku']))
    approval_rows = [
        row for row in pending_rows
        if row.get('recipe') and row.get('recipe_status') in {'pending_review', 'reviewed'}
    ]

    context = {
        'pending_rows': pending_rows,
        'approval_rows': approval_rows,
        'pending_count': len(pending_rows),
        'approval_count': len(approval_rows),
        'po_summary': sorted(po_summary_map.values(), key=lambda x: x['po_number']),
        'po_filter': po_filter,
        'q': q,
        'can_admin_actions': is_admin_user,
    }
    return render(request, 'planning/pending_skus.html', context)


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_edit_jobcard')
@transaction.atomic
def pending_skus_ignored(request):
    """Display pending SKUs that were marked ignored and no longer appear in the active pending queue."""
    is_admin_user = _user_is_admin(request.user)
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        sku = (request.POST.get('sku') or '').strip()
        po_doc_id = request.POST.get('po_doc_id')
        if action == 'unignore':
            try:
                po_doc = PoDocument.objects.get(id=int(po_doc_id)) if po_doc_id else None
            except (TypeError, ValueError, PoDocument.DoesNotExist):
                po_doc = None

            if not po_doc or not sku:
                messages.error(request, 'Invalid PO or SKU for unignore action.')
                return redirect('qc:pending_skus_ignored')

            payload = po_doc.extracted_payload or {}
            ignored = [s for s in (payload.get('new_skus_ignored') or []) if s]
            normalized = _sku_key(sku)
            kept = [s for s in ignored if _sku_key(s) != normalized]
            payload['new_skus_ignored'] = sorted(kept)
            po_doc.extracted_payload = payload
            po_doc.save(update_fields=['extracted_payload'])
            messages.success(request, f'SKU {sku} restored to the pending queue.')
            return redirect('qc:pending_skus_ignored')

    po_filter = (request.GET.get('po') or '').strip()
    q = (request.GET.get('q') or '').strip()

    docs = PoDocument.objects.exclude(extracted_payload__isnull=True).order_by('-created_at')[:400]
    rows = []
    for doc in docs:
        payload = doc.extracted_payload or {}
        ignored_skus = [s for s in (payload.get('new_skus_ignored') or []) if s]
        if not ignored_skus:
            continue

        po_number = payload.get('po_number') or '-'
        for sku in ignored_skus:
            if not sku:
                continue
            recipe = SkuRecipe.objects.filter(sku__iexact=sku).first()
            rows.append({
                'po_doc_id': doc.id,
                'po_number': po_number,
                'supplier': payload.get('supplier_name') or '-',
                'sku': sku,
                'job_name': recipe.job_name if recipe else '-',
                'recipe_status': recipe.master_data_status if recipe else 'missing',
                'recipe': recipe,
                'uploaded_at': doc.created_at,
            })

    if po_filter:
        rows = [row for row in rows if row['po_number'] == po_filter]
    if q:
        q_upper = q.upper()
        rows = [
            row
            for row in rows
            if q_upper in (row['sku'] or '').upper()
            or q_upper in (row['po_number'] or '').upper()
            or q_upper in (row['job_name'] or '').upper()
        ]

    po_summary = {}
    for row in rows:
        po_summary.setdefault(row['po_number'], {'po_number': row['po_number'], 'count': 0})
        po_summary[row['po_number']]['count'] += 1

    rows.sort(key=lambda row: (row['po_number'], row['sku']))

    return render(request, 'planning/pending_skus_ignored.html', {
        'rows': rows,
        'po_summary': sorted(po_summary.values(), key=lambda x: x['po_number']),
        'po_filter': po_filter,
        'q': q,
        'can_admin_actions': is_admin_user,
    })


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_edit_jobcard')
@transaction.atomic
def pending_sku_master_entry(request):
    """Open a focused form for one pending SKU and send it through master-data approval flow."""
    sku = (request.GET.get('sku') or request.POST.get('sku') or '').strip()
    po_doc_id_raw = request.GET.get('po_doc_id') or request.POST.get('po_doc_id')
    return_po = (request.GET.get('return_po') or request.POST.get('return_po') or '').strip()
    return_q = (request.GET.get('return_q') or request.POST.get('return_q') or '').strip()

    def _redirect_pending():
        params = {}
        if return_po:
            params['po'] = return_po
        if return_q:
            params['q'] = return_q
        url = reverse('qc:pending_skus')
        return redirect(f'{url}?{urlencode(params)}' if params else url)

    try:
        po_doc_id = int(po_doc_id_raw)
    except (TypeError, ValueError):
        po_doc_id = None

    if not sku or not po_doc_id:
        messages.error(request, 'Missing SKU or PO reference for master-data entry.')
        return _redirect_pending()

    po_doc = PoDocument.objects.filter(id=po_doc_id).first()
    if not po_doc:
        messages.error(request, 'PO document was not found.')
        return _redirect_pending()

    payload = po_doc.extracted_payload or {}
    po_number = payload.get('po_number') or '-'
    items = _sanitize_po_payload_items(payload)
    is_admin_user = _user_is_admin(request.user)

    suggested_item = None
    sku_key = _sku_key(sku)
    for item in items:
        if _sku_key(item.get('sku')) == sku_key:
            suggested_item = item
            break
    suggested_item = suggested_item or {}
    po_job_name = (suggested_item.get('job_name') or '').strip() or sku
    po_department = (payload.get('department') or '').strip()
    po_unit_cost = _to_decimal(suggested_item.get('unit_cost'))
    po_color_spec = _normalize_color_spec_input(
        suggested_item.get('color_spec') or suggested_item.get('colour') or suggested_item.get('color') or ''
    )
    po_application = _normalize_application_input(suggested_item.get('application') or payload.get('application') or '')

    recipe = SkuRecipe.objects.filter(sku__iexact=sku).first()

    if request.method == 'POST':
        action = (request.POST.get('action') or 'save_draft').strip()
        if recipe and action == 'delete':
            if recipe.master_data_status == 'approved' and not is_admin_user:
                messages.error(request, 'Approved records can only be deleted by admin users.')
            else:
                recipe.delete()
                messages.success(request, f'SKU Recipe "{recipe.sku}" deleted.')
            return _redirect_pending()

        if recipe and action == 'archive':
            if recipe.master_data_status == 'approved' and not is_admin_user:
                messages.error(request, 'Approved records can only be archived by admin users.')
                return _redirect_pending()
            recipe.is_active = False
            recipe.archived_by = request.user
            recipe.archived_at = timezone.now()
            recipe.archive_reason = (request.POST.get('archive_reason') or '').strip()
            recipe.save(update_fields=['is_active', 'archived_by', 'archived_at', 'archive_reason', 'updated_at'])
            messages.success(request, f'SKU Recipe "{recipe.sku}" archived.')
            return _redirect_pending()

        posted = request.POST.copy()
        posted['job_name'] = po_job_name
        posted['sku'] = sku
        if not (posted.get('default_unit_cost') or '').strip() and po_unit_cost is not None:
            posted['default_unit_cost'] = str(po_unit_cost)
        if not (posted.get('color_spec') or '').strip() and po_color_spec:
            posted['color_spec'] = po_color_spec
        if not (posted.get('application') or '').strip() and po_application:
            posted['application'] = po_application
        form = SkuRecipeForm(posted, instance=recipe)
        if form.is_valid():
            action = (request.POST.get('action') or 'save_draft').strip()
            obj = form.save(commit=False)
            obj.sku = sku
            obj.job_name = po_job_name
            if not recipe:
                obj.created_by = request.user

            if action == 'submit_review':
                missing_required = _missing_required_master_fields(obj, allow_missing_plate_set_no=True)
                if missing_required:
                    messages.error(
                        request,
                        f'SKU {sku} cannot be sent for QC review. Missing required data: {", ".join(missing_required)}.',
                    )
                else:
                    obj.master_data_status = 'pending_review'
                    obj.reviewed_by = None
                    obj.reviewed_at = None
                    obj.approved_by = None
                    obj.approved_at = None
                    obj.save()

                    configured = set(payload.get('new_skus_configured') or [])
                    configured.add(sku)
                    payload['new_skus_configured'] = sorted(configured)
                    po_doc.extracted_payload = payload
                    po_doc.save(update_fields=['extracted_payload'])

                    messages.success(request, f'SKU {sku} submitted for QC review.')
                    return _redirect_pending()
            else:
                obj.master_data_status = 'draft'
                obj.reviewed_by = None
                obj.reviewed_at = None
                obj.approved_by = None
                obj.approved_at = None
                obj.save()

                configured = set(payload.get('new_skus_configured') or [])
                configured.add(sku)
                payload['new_skus_configured'] = sorted(configured)
                po_doc.extracted_payload = payload
                po_doc.save(update_fields=['extracted_payload'])

                messages.success(request, f'SKU {sku} saved as Draft.')
                return _redirect_pending()
    else:
        initial = {
            'sku': sku,
            'job_name': po_job_name,
            'default_unit_cost': (recipe.default_unit_cost if recipe else None) or po_unit_cost,
            'color_spec': (recipe.color_spec if recipe else '') or po_color_spec,
            'application': (recipe.application if recipe else '') or po_application,
        }
        form = SkuRecipeForm(instance=recipe, initial=initial)

    form.fields['sku'].widget.attrs['readonly'] = True

    current_recipe = recipe
    if request.method == 'POST' and form.is_valid() and 'obj' in locals():
        current_recipe = obj

    mismatch_alerts = []
    if current_recipe:
        cost_alert = _build_cost_mismatch_note(current_recipe.default_unit_cost, po_unit_cost)
        if cost_alert:
            mismatch_alerts.append(cost_alert)

    context = {
        'form': form,
        'sku': sku,
        'po_doc_id': po_doc_id,
        'po_number': po_number,
        'return_po': return_po,
        'return_q': return_q,
        'suggested_job_name': po_job_name,
        'suggested_qty': _format_display_qty(suggested_item.get('quantity')),
        'suggested_delivery_date': suggested_item.get('delivery_date') or '-',
        'recipe_status': (current_recipe.master_data_status if current_recipe else 'missing'),
        'missing_required_fields': _missing_required_master_fields(
            current_recipe, po_job_name, allow_missing_plate_set_no=True,
        ),
        'warning_master_fields': _warning_master_fields(current_recipe, po_job_name),
        'mismatch_alerts': mismatch_alerts,
    }
    context['can_admin_actions'] = is_admin_user
    return render(request, 'planning/pending_sku_master_entry.html', context)


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_edit_jobcard')
def sku_recipes_list(request):
    """List all SKU recipes with search; handles delete via POST."""
    is_admin_user = _user_is_admin(request.user)
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        recipe_id = request.POST.get('recipe_id')
        redirect_url = request.path
        if request.GET:
            redirect_url += '?' + request.GET.urlencode()

        if action == 'delete':
            try:
                recipe_obj = SkuRecipe.objects.get(id=int(recipe_id))
                if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                    messages.error(request, 'Approved records can only be deleted by admin users.')
                    return redirect(redirect_url)
                recipe_obj.delete()
                messages.success(request, 'SKU Recipe deleted.')
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
            return redirect(redirect_url)

        if action == 'archive':
            try:
                recipe_obj = SkuRecipe.objects.get(id=int(recipe_id))
                if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                    messages.error(request, 'Approved records can only be archived by admin users.')
                    return redirect(redirect_url)
                recipe_obj.is_active = False
                recipe_obj.archived_by = request.user
                recipe_obj.archived_at = timezone.now()
                recipe_obj.archive_reason = (request.POST.get('archive_reason') or '').strip()
                recipe_obj.save(update_fields=['is_active', 'archived_by', 'archived_at', 'archive_reason', 'updated_at'])
                messages.success(request, 'SKU Recipe archived.')
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
            return redirect(redirect_url)

        if action in {'bulk_archive', 'bulk_delete'}:
            selected_ids = []
            for raw_id in request.POST.getlist('selected_ids'):
                try:
                    selected_ids.append(int(raw_id))
                except (TypeError, ValueError):
                    continue

            if not selected_ids:
                messages.error(request, 'Select at least one SKU recipe first.')
                return redirect(redirect_url)

            processed = 0
            skipped_locked = 0
            failures = []
            for recipe_obj in SkuRecipe.objects.filter(id__in=selected_ids, is_active=True):
                if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                    skipped_locked += 1
                    continue

                if action == 'bulk_archive':
                    recipe_obj.is_active = False
                    recipe_obj.archived_by = request.user
                    recipe_obj.archived_at = timezone.now()
                    recipe_obj.archive_reason = ''
                    recipe_obj.save(update_fields=['is_active', 'archived_by', 'archived_at', 'archive_reason', 'updated_at'])
                    processed += 1
                else:
                    try:
                        recipe_obj.delete()
                        processed += 1
                    except Exception as exc:
                        failures.append(f'{recipe_obj.sku}: {str(exc)}')

            if action == 'bulk_archive':
                messages.success(request, f'Bulk archive complete. Archived {processed}, skipped {skipped_locked}.')
            else:
                messages.success(request, f'Bulk delete complete. Deleted {processed}, skipped {skipped_locked}.')
            if failures:
                messages.error(request, 'Some items could not be processed: ' + '; '.join(failures))
            return redirect(redirect_url)

        if action in {'submit_review', 'review', 'approve', 'back_to_draft'}:
            try:
                recipe = SkuRecipe.objects.get(id=int(recipe_id))
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
                return redirect(redirect_url)

            current_status = (recipe.master_data_status or 'draft').lower()

            if action == 'submit_review':
                if current_status != 'draft':
                    messages.error(request, f'SKU {recipe.sku} can only be submitted for review from Draft.')
                    return redirect(redirect_url)
                recipe.master_data_status = 'pending_review'
                recipe.reviewed_by = None
                recipe.reviewed_at = None
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.save(update_fields=['master_data_status', 'reviewed_by', 'reviewed_at', 'approved_by', 'approved_at', 'updated_at'])
                messages.success(request, f'SKU {recipe.sku} submitted for review.')
                return redirect(redirect_url)

            if action == 'review':
                if current_status != 'pending_review':
                    messages.error(request, f'SKU {recipe.sku} can only move to Reviewed from Pending Review.')
                    return redirect(redirect_url)
                recipe.master_data_status = 'reviewed'
                recipe.reviewed_by = request.user
                recipe.reviewed_at = timezone.now()
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.save(update_fields=['master_data_status', 'reviewed_by', 'reviewed_at', 'approved_by', 'approved_at', 'updated_at'])
                messages.success(request, f'SKU {recipe.sku} moved to Reviewed.')
                return redirect(redirect_url)

            if action == 'approve':
                if current_status != 'reviewed':
                    messages.error(request, f'SKU {recipe.sku} can only be Approved from Reviewed status.')
                    return redirect(redirect_url)
                missing_required = _missing_required_master_fields(recipe, recipe.job_name, allow_missing_plate_set_no=True)
                if missing_required:
                    messages.error(
                        request,
                        f'SKU {recipe.sku} cannot be approved. Missing required master data: {", ".join(missing_required)}.',
                    )
                    return redirect(redirect_url)
                recipe.master_data_status = 'approved'
                recipe.approved_by = request.user
                recipe.approved_at = timezone.now()
                recipe.save(update_fields=['master_data_status', 'approved_by', 'approved_at', 'updated_at'])
                approval_warnings = _warning_master_fields(recipe, recipe.job_name)
                if approval_warnings:
                    messages.warning(
                        request,
                        f'SKU {recipe.sku} approved. Notice: {", ".join(approval_warnings)} not set yet '
                        '(usually assigned at plate making).',
                    )
                else:
                    messages.success(request, f'SKU {recipe.sku} approved.')
                return redirect(redirect_url)

            if action == 'back_to_draft':
                if current_status == 'draft':
                    messages.info(request, f'SKU {recipe.sku} is already in Draft.')
                    return redirect(redirect_url)
                if current_status == 'approved' and not is_admin_user:
                    messages.error(request, 'Approved records can only be reverted by admin users.')
                    return redirect(redirect_url)
                comment = (request.POST.get('rejection_comment') or '').strip()
                if not comment:
                    messages.error(request, 'Please provide a reason when sending a record back to Draft.')
                    return redirect(redirect_url)
                recipe.master_data_status = 'draft'
                recipe.reviewed_by = None
                recipe.reviewed_at = None
                recipe.approved_by = None
                recipe.approved_at = None
                recipe.rejection_comment = comment
                recipe.last_rejected_by = request.user
                recipe.last_rejected_at = timezone.now()
                recipe.save(update_fields=[
                    'master_data_status', 'reviewed_by', 'reviewed_at',
                    'approved_by', 'approved_at', 'rejection_comment',
                    'last_rejected_by', 'last_rejected_at', 'updated_at',
                ])
                if comment:
                    messages.warning(request, f'SKU {recipe.sku} sent back to Draft. Reason: {comment}')
                else:
                    messages.success(request, f'SKU {recipe.sku} moved back to Draft.')
                return redirect(redirect_url)

    q = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    qs = SkuRecipe.objects.filter(is_active=True)
    if q:
        qs = qs.filter(
            Q(sku__icontains=q)
            | Q(job_name__icontains=q)
            | Q(material__icontains=q)
            | Q(application__icontains=q)
        )
    if status_filter in ('draft', 'pending_review', 'reviewed', 'approved'):
        qs = qs.filter(master_data_status=status_filter)
    paginator = Paginator(qs, 50)
    recipes = paginator.get_page(request.GET.get('page'))

    bulk_highlights = request.session.pop('sku_recipe_bulk_highlights', {})
    for recipe in recipes:
        meta = bulk_highlights.get(str(recipe.id), {})
        recipe.bulk_highlight_type = meta.get('type', '')
        recipe.bulk_highlight_fields = meta.get('fields', [])

    return render(request, 'planning/sku_recipes.html', {
        'recipes': recipes,
        'q': q,
        'status_filter': status_filter,
        'can_edit_approved': is_admin_user,
        'can_admin_actions': is_admin_user,
    })


# MOVED TO QC APP (temporary compatibility layer)
@login_required
def sku_recipes_status(request, status=None):
    """List SKU recipes filtered by a fixed status for role-specific views."""
    if status not in {'draft', 'pending_review', 'reviewed', 'approved'}:
        raise Http404('Unknown SKU recipe status view.')
    request.GET = request.GET.copy()
    request.GET['status'] = status
    return sku_recipes_list(request)


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_edit_jobcard')
def sku_recipes_archived(request):
    """List archived SKU recipes."""
    is_admin_user = _user_is_admin(request.user)
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        redirect_url = request.path
        if request.GET:
            redirect_url += '?' + request.GET.urlencode()

        if action == 'restore':
            recipe_id = request.POST.get('recipe_id')
            try:
                recipe_obj = SkuRecipe.objects.get(id=int(recipe_id), is_active=False)
                if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                    messages.error(request, 'Approved recipes can only be restored by admin users.')
                    return redirect(redirect_url)
                recipe_obj.is_active = True
                recipe_obj.save(update_fields=['is_active', 'updated_at'])
                messages.success(request, 'SKU Recipe restored to active master list.')
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
            return redirect(redirect_url)

        if action == 'delete':
            if not is_admin_user:
                messages.error(request, 'Only admin users can permanently delete archived recipes.')
                return redirect(redirect_url)

            recipe_id = request.POST.get('recipe_id')
            try:
                recipe_obj = SkuRecipe.objects.get(id=int(recipe_id), is_active=False)
                recipe_obj.delete()
                messages.success(request, 'Archived SKU Recipe permanently deleted.')
            except (TypeError, ValueError, SkuRecipe.DoesNotExist):
                messages.error(request, 'Invalid recipe ID.')
            return redirect(redirect_url)

        if action == 'bulk_restore':
            selected_ids = []
            for raw_id in request.POST.getlist('selected_ids'):
                try:
                    selected_ids.append(int(raw_id))
                except (TypeError, ValueError):
                    continue

            if not selected_ids:
                messages.error(request, 'Select at least one archived SKU recipe first.')
                return redirect(redirect_url)

            restored = 0
            skipped_locked = 0
            for recipe_obj in SkuRecipe.objects.filter(id__in=selected_ids, is_active=False):
                if recipe_obj.master_data_status == 'approved' and not is_admin_user:
                    skipped_locked += 1
                    continue
                recipe_obj.is_active = True
                recipe_obj.save(update_fields=['is_active', 'updated_at'])
                restored += 1

            messages.success(request, f'Bulk restore complete. Restored {restored}, skipped {skipped_locked}.')
            return redirect(redirect_url)

        if action == 'bulk_delete':
            if not is_admin_user:
                messages.error(request, 'Only admin users can permanently delete archived recipes.')
                return redirect(redirect_url)

            selected_ids = []
            for raw_id in request.POST.getlist('selected_ids'):
                try:
                    selected_ids.append(int(raw_id))
                except (TypeError, ValueError):
                    continue

            if not selected_ids:
                messages.error(request, 'Select at least one archived SKU recipe first.')
                return redirect(redirect_url)

            deleted = 0
            failures = []
            for recipe_obj in SkuRecipe.objects.filter(id__in=selected_ids, is_active=False):
                try:
                    recipe_obj.delete()
                    deleted += 1
                except Exception as exc:
                    failures.append(f'{recipe_obj.sku}: {str(exc)}')

            messages.success(request, f'Bulk delete complete. Deleted {deleted}.')
            if failures:
                messages.error(request, 'Some items could not be deleted: ' + '; '.join(failures))
            return redirect(redirect_url)

    q = (request.GET.get('q') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    qs = SkuRecipe.objects.filter(is_active=False)
    if q:
        qs = qs.filter(
            Q(sku__icontains=q)
            | Q(job_name__icontains=q)
            | Q(material__icontains=q)
            | Q(machine_name__icontains=q)
        )
    if status_filter in ('draft', 'pending_review', 'reviewed', 'approved'):
        qs = qs.filter(master_data_status=status_filter)
    paginator = Paginator(qs, 50)
    recipes = paginator.get_page(request.GET.get('page'))
    return render(request, 'planning/sku_recipes_archived.html', {
        'recipes': recipes,
        'q': q,
        'status_filter': status_filter,
        'can_restore_approved': is_admin_user,
        'can_delete_archived': is_admin_user,
    })


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_edit_jobcard')
def sku_recipe_edit(request, recipe_id=None):
    """Create or edit a single SKU recipe."""
    if recipe_id:
        recipe = get_object_or_404(SkuRecipe, id=recipe_id, is_active=True)
        page_title = f'Edit SKU Recipe — {recipe.sku}'
    else:
        recipe = None
        page_title = 'Add New SKU Recipe'

    is_admin_user = _user_is_admin(request.user)
    can_edit_approved = not (recipe and recipe.master_data_status == 'approved')
    can_admin_actions = is_admin_user

    if request.method == 'POST':
        action = request.POST.get('action', '').strip()
        _do_sync_on_approve = False
        if recipe and action == 'delete':
            if recipe.master_data_status == 'approved' and not is_admin_user:
                messages.error(request, 'Approved records can only be deleted by admin users.')
            else:
                recipe.delete()
                messages.success(request, f'SKU Recipe "{recipe.sku}" deleted.')
            return redirect('planning:sku_recipes')

        if recipe and action == 'archive':
            if recipe.master_data_status == 'approved' and not is_admin_user:
                messages.error(request, 'Approved records can only be archived by admin users.')
                return redirect('planning:sku_recipes')
            recipe.is_active = False
            recipe.archived_by = request.user
            recipe.archived_at = timezone.now()
            recipe.archive_reason = (request.POST.get('archive_reason') or '').strip()
            recipe.save(update_fields=['is_active', 'archived_by', 'archived_at', 'archive_reason', 'updated_at'])
            messages.success(request, f'SKU Recipe "{recipe.sku}" archived.')
            return redirect('planning:sku_recipes')

        if recipe and recipe.master_data_status == 'approved' and action != 'reopen_sku':
            messages.error(request, 'Approved SKU is locked. Use Reopen SKU before making edits.')
            return render(request, 'planning/sku_recipe_edit.html', {'form': SkuRecipeForm(instance=recipe), 'recipe': recipe, 'page_title': page_title, 'can_edit_approved': can_edit_approved, 'can_admin_actions': can_admin_actions})

        form = SkuRecipeForm(request.POST, instance=recipe)
        if form.is_valid():
            obj = form.save(commit=False)
            if not recipe_id:
                obj.created_by = request.user

            current_status = (recipe.master_data_status if recipe else 'draft')
            obj.master_data_status = current_status
            obj.reviewed_by = recipe.reviewed_by if recipe else None
            obj.reviewed_at = recipe.reviewed_at if recipe else None
            obj.approved_by = recipe.approved_by if recipe else None
            obj.approved_at = recipe.approved_at if recipe else None

            if recipe_id and action:
                if action == 'submit_review' and current_status == 'draft':
                    obj.master_data_status = 'pending_review'
                    obj.reviewed_by = None
                    obj.reviewed_at = None
                    obj.approved_by = None
                    obj.approved_at = None
                    messages.success(request, f'SKU Recipe "{obj.sku}" submitted for review. Status: Pending Review.')
                elif action == 'review' and current_status == 'pending_review':
                    obj.master_data_status = 'reviewed'
                    obj.reviewed_by = request.user
                    obj.reviewed_at = timezone.now()
                    obj.approved_by = None
                    obj.approved_at = None
                    messages.success(request, f'SKU Recipe "{obj.sku}" reviewed and submitted for approval.')
                elif action == 'approve' and current_status == 'reviewed':
                    missing = _missing_required_master_fields(obj, allow_missing_plate_set_no=True)
                    if missing:
                        messages.error(request, f'Cannot approve. Missing required fields: {", ".join(missing)}.')
                        return render(request, 'planning/sku_recipe_edit.html', {'form': form, 'recipe': obj, 'page_title': page_title, 'can_edit_approved': can_edit_approved, 'can_admin_actions': can_admin_actions})
                    obj.master_data_status = 'approved'
                    obj.approved_by = request.user
                    obj.approved_at = timezone.now()
                    _do_sync_on_approve = True
                    approval_warnings = _warning_master_fields(obj)
                    if approval_warnings:
                        messages.warning(
                            request,
                            f'SKU Recipe "{obj.sku}" approved. Notice: {", ".join(approval_warnings)} not set yet '
                            '(usually assigned at plate making).',
                        )
                    else:
                        messages.success(request, f'SKU Recipe "{obj.sku}" approved for master data usage.')
                elif action == 'back_to_draft' and current_status in ('pending_review', 'reviewed'):
                    comment = (request.POST.get('rejection_comment') or '').strip()
                    if not comment:
                        messages.error(request, 'Please provide a reason when sending a record back to Draft.')
                        return render(request, 'planning/sku_recipe_edit.html', {'form': form, 'recipe': obj, 'page_title': page_title, 'can_edit_approved': can_edit_approved})
                    obj.master_data_status = 'draft'
                    obj.reviewed_by = None
                    obj.reviewed_at = None
                    obj.approved_by = None
                    obj.approved_at = None
                    obj.rejection_comment = comment
                    obj.last_rejected_by = request.user
                    obj.last_rejected_at = timezone.now()
                    messages.success(request, f'SKU Recipe "{obj.sku}" moved back to Draft.')
                elif action == 'reopen_sku' and current_status == 'approved':
                    comment = (request.POST.get('rejection_comment') or '').strip()
                    if not comment:
                        messages.error(request, 'Please provide a reopen reason before moving approved SKU back to Draft.')
                        return render(request, 'planning/sku_recipe_edit.html', {'form': form, 'recipe': obj, 'page_title': page_title, 'can_edit_approved': can_edit_approved})
                    obj.master_data_status = 'draft'
                    obj.reviewed_by = None
                    obj.reviewed_at = None
                    obj.approved_by = None
                    obj.approved_at = None
                    obj.rejection_comment = comment
                    obj.last_rejected_by = request.user
                    obj.last_rejected_at = timezone.now()
                    messages.success(request, f'SKU Recipe "{obj.sku}" reopened to Draft.')
                else:
                    messages.info(request, f'SKU Recipe "{obj.sku}" saved without changing workflow status.')
            else:
                if recipe_id:
                    messages.success(request, f'SKU Recipe "{obj.sku}" saved.')
                else:
                    obj.master_data_status = 'draft'
                    obj.reviewed_by = None
                    obj.reviewed_at = None
                    obj.approved_by = None
                    obj.approved_at = None
                    messages.success(request, f'SKU Recipe "{obj.sku}" saved as Draft. Submit for approval from SKU Recipe Master.')

            obj.save()
            if _do_sync_on_approve:
                try:
                    sync_result = _sync_new_jobs_for_approved_sku(obj.sku, actor=request.user)
                    messages.success(
                        request,
                        f'Planning jobs refreshed for approved SKU: updated {sync_result["updated"]}, locked {sync_result["locked"]}, missing draft jobs {sync_result.get("missing_jobs", 0)}.',
                    )
                except Exception:
                    logger.exception('Approved SKU sync to planning failed for %s', obj.sku)
                    messages.error(request, 'Error while sending approved SKU to Planning; check logs.')
            return redirect('planning:sku_recipes')
        else:
            messages.error(request, 'There are errors in the form. Please correct the highlighted fields and try again.')
    else:
        form = SkuRecipeForm(instance=recipe)

    return render(request, 'planning/sku_recipe_edit.html', {
        'form': form,
        'recipe': recipe,
        'page_title': page_title,
        'can_edit_approved': can_edit_approved,
        'can_admin_actions': can_admin_actions,
    })


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_edit_jobcard')
def sku_recipe_bulk_upload(request):
    """Bulk upload SKU recipes from CSV/XLSX into Draft status for approval workflow."""
    if request.method == 'POST':
        upload_file = request.FILES.get('upload_file')
        if not upload_file:
            messages.error(request, 'Please choose a CSV or XLSX file to upload.')
            return redirect('planning:sku_recipe_bulk_upload')

        name = (upload_file.name or '').lower()
        rows = []
        header_to_field = {
            'SKU': 'sku',
            'JOB NAME': 'job_name',
            'Material': 'material',
            'Color': 'color_spec',
            'Application': 'application',
            'Size W mm': 'size_w_mm',
            'Size H mm': 'size_h_mm',
            'Size W Inch': 'size_w_inch',
            'Size H Inch': 'size_h_inch',
            'Ups': 'ups',
            'Print Sheet Size': 'print_sheet_size',
            'Purchase Sheet Size': 'purchase_sheet_size',
            'Purchase Sheet ups': 'purchase_sheet_ups',
            'Purchase Material': 'purchase_material',
            'Machine Name': 'machine_name',
            'Machine': 'machine_name',
            'Cost': 'default_unit_cost',
            'Default Unit Cost': 'default_unit_cost',
            'Daily Demand': 'daily_demand',
            'AWC No.': 'awc_no',
            'AWC No': 'awc_no',
            'Plate Set No': 'plate_set_no',
            'Die': 'die_cutting',
            'Notes': 'notes',
        }
        int_clean_fields = {'size_w_mm', 'size_h_mm', 'size_w_inch', 'size_h_inch'}
        def clean_intlike(val):
            try:
                if val is None or str(val).strip() == '':
                    return ''
                ival = int(float(val))
                return str(ival)
            except:
                return str(val) if val is not None else ''
        try:
            if name.endswith('.csv'):
                decoded = upload_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                for row in reader:
                    rows.append(row)
            elif name.endswith('.xlsx'):
                try:
                    import openpyxl
                except ImportError:
                    messages.error(request, 'openpyxl is required for XLSX upload.')
                    return redirect('planning:sku_recipe_bulk_upload')
                wb = openpyxl.load_workbook(upload_file, data_only=True)
                ws = wb.active
                header_row_idx = None
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    values = [str(c).strip() if c is not None else '' for c in row]
                    if 'SKU' in values and 'JOB NAME' in values:
                        header_row_idx = i
                        header = values
                        break
                if not header_row_idx:
                    messages.error(request, 'Could not find header row in Excel file. Make sure it matches the template.')
                    return redirect('planning:sku_recipe_bulk_upload')
                for values in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
                    row = {}
                    for idx, key in enumerate(header):
                        if key:
                            row[key] = values[idx] if idx < len(values) else None
                    rows.append(row)
            else:
                messages.error(request, 'Unsupported file type. Please upload CSV or XLSX.')
                return redirect('planning:sku_recipe_bulk_upload')
        except Exception as exc:
            messages.error(request, f'Could not read upload file: {exc}')
            return redirect('planning:sku_recipe_bulk_upload')

        if not rows:
            messages.error(request, 'No rows found in upload file.')
            return redirect('planning:sku_recipe_bulk_upload')

        created = 0
        updated = 0
        failed = 0
        sample_errors = []
        bulk_highlights = {}
        highlight_fields = {
            'sku', 'job_name', 'material', 'color_spec', 'application',
            'size_w_mm', 'size_h_mm', 'ups', 'print_sheet_size',
            'purchase_sheet_size', 'purchase_sheet_ups', 'purchase_material',
            'machine_name', 'default_unit_cost', 'daily_demand',
            'awc_no', 'plate_set_no', 'die_cutting', 'notes',
        }

        for idx, source in enumerate(rows, start=2):
            payload = {}
            lam_fnb = False
            for header, field in header_to_field.items():
                value = source.get(header, '')
                if field == 'application':
                    if 'f+b' in (str(value or '').lower()):
                        lam_fnb = True
                    payload[field] = _normalize_application_input(value)
                elif field == 'color_spec':
                    payload[field] = _normalize_color_spec_input(value)
                elif field in {'size_w_mm', 'size_h_mm'}:
                    try:
                        if value is None or str(value).strip() == '':
                            payload[field] = ''
                        else:
                            payload[field] = str(int(float(value)))
                    except:
                        payload[field] = str(value) if value is not None else ''
                elif field in int_clean_fields:
                    payload[field] = clean_intlike(value)
                else:
                    payload[field] = '' if value is None else str(value).strip()
            payload['lamination_front_and_back'] = lam_fnb
            if not payload.get('sku'):
                continue
            existing = SkuRecipe.objects.filter(sku__iexact=payload['sku']).first()
            form = SkuRecipeForm(payload, instance=existing)
            if not form.is_valid():
                failed += 1
                if len(sample_errors) < 8:
                    error_text = '; '.join(
                        f"{name}: {', '.join([str(msg) for msg in msgs])}"
                        for name, msgs in form.errors.items()
                    )
                    sample_errors.append(f'Row {idx} ({payload["sku"]}): {error_text}')
                continue
            obj = form.save(commit=False)
            if not existing:
                obj.created_by = request.user
            obj.master_data_status = 'draft'
            obj.reviewed_by = None
            obj.reviewed_at = None
            obj.approved_by = None
            obj.approved_at = None
            obj.save()

            if existing:
                changed = [field for field in form.changed_data if field in highlight_fields]
                if not changed:
                    changed = [
                        field for field in highlight_fields
                        if str(getattr(existing, field, '') or '').strip() != str(form.cleaned_data.get(field, '') or '').strip()
                    ]
                if not changed:
                    changed = ['sku']
                bulk_highlights[str(obj.id)] = {
                    'type': 'updated',
                    'fields': changed,
                }
                updated += 1
            else:
                created_fields = [
                    field for field in form.cleaned_data
                    if field in highlight_fields and form.cleaned_data.get(field) not in (None, '')
                ]
                bulk_highlights[str(obj.id)] = {
                    'type': 'created',
                    'fields': created_fields,
                }
                created += 1

        if bulk_highlights:
            request.session['sku_recipe_bulk_highlights'] = bulk_highlights

        if created or updated:
            messages.success(
                request,
                f'Bulk upload complete. Draft recipes created {created}, updated {updated}, failed {failed}.',
            )
        if failed and sample_errors:
            messages.error(request, 'Sample row errors: ' + ' | '.join(sample_errors))

        return redirect('planning:sku_recipes')

    return render(request, 'planning/sku_recipe_bulk_upload.html')


# MOVED TO QC APP (temporary compatibility layer)
@login_required
@permission_required('can_view_jobcard')
def sku_recipe_template_download(request):
    """Return a CSV template for bulk SKU recipe upload."""
    headers = [
        'Sno.', 'SKU', 'JOB NAME', 'Order Status', 'Material', 'Color', 'Application',
        'Size W mm', 'Size H mm', 'Size W Inch', 'Size H Inch', 'Ups', 'Print Sheet Size',
        'Purchase Sheet Size', 'Purchase Sheet ups', 'Purchase Material', 'Machine',
        'Default Unit Cost', 'Daily Demand', 'AWC No', 'Plate Set No', 'Die', 'Notes'
    ]
    sample_row = [
        '1', 'SKU-001', 'Sample Job Name', 'Repeat', 'Art Card 300gsm', '4 color', 'UV',
        '100', '150', '3.94', '5.91', '4', '720x1020', '720x1020', '2', 'Local',
        'Heidelberg SM52', '5.00', '500', 'AWC-001', 'PLT-001', 'YES', 'Sample notes'
    ]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(sample_row)
    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="sku_recipe_upload_template.csv"'
    return response
