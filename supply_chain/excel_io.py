import csv
import io
from datetime import datetime

from django.db.models import Q
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from .models import RawMaterialSku, StockDemand, StockTransaction
from .raw_material_sku import import_raw_material_skus as _import_raw_material_skus

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


TRANSACTION_HEADERS = [
    'Month',
    'Date',
    'GIN / JC',
    'Raw Material SKU',
    'Material Name',
    'Purchase Sheet Size',
    'UOM',
    'Sheet Packing/Pcs',
    'Sheet Qty/Pcs',
    'Pkt/Rim Qty',
]

DEMAND_HEADERS = [
    'Raw Material SKU',
    'Material Name',
    'Purchase Sheet Size',
    'UOM',
    'Sheet Packing/Pcs',
    'Month',
    'Sheet Qty/Pcs',
    'Pkt/Rim Qty',
]

RAW_MATERIAL_SKU_HEADERS = [
    'Raw Material SKU',
    'Material Name',
    'Purchase Sheet Size',
    'UOM',
    'Sheet Packing/Pcs',
    'Unit Cost',
    'Safety Stock',
    'Max Stock Level',
    'Lead Time (Days)',
    'Active',
]

ITEM_HEADERS = RAW_MATERIAL_SKU_HEADERS

ITEM_REQUEST_HEADERS = [
    'IR-ID',
    'Status',
    'Type',
    'Department',
    'Item Title',
    'Machine',
    'UOM',
    'Required Quantity',
    'Local/Import',
    'Part Number',
    'Estimated Unit Price',
    'Cost Centre',
    'Raised By',
    'Request Date',
]

ITEM_WISE_CONSUMPTION_HEADERS = [
    'Raw Material SKU',
    'Material Name',
    'Purchase Sheet Size',
    'UOM',
    'Sheet Packing/Pcs',
    'Month',
    'Sheet Qty/Pcs',
    'Pkt/Rim Qty',
    'Consumption Value',
]

KPI_HEADERS = [
    'Raw Material SKU',
    'Material Name',
    'Purchase Sheet Size',
    'ABC',
    'FSN',
    'Closing Stock',
    'Monthly Demand',
    'Daily Demand',
    'Reorder Point',
    'Safety Stock',
    'Inventory Value',
    'Inventory Turnover',
    'Annual Consumption Value',
    'Stockout',
    'Reorder Alert',
    'Safety Stock Alert',
    'Overstock',
    'Dead Stock',
    'Slow Moving',
    'Inventory Accuracy %',
]

PHYSICAL_COUNT_HEADERS = [
    'Count Date',
    'Raw Material SKU',
    'Material Name',
    'Purchase Sheet Size',
    'Physical Sheet Qty/Pcs',
    'System Sheet Qty/Pcs',
    'Variance',
    'Inventory Accuracy %',
    'Notes',
]

MONTH_WISE_CONSUMPTION_HEADERS = [
    'Month',
    'Raw Material SKU',
    'Material Name',
    'Purchase Sheet Size',
    'UOM',
    'Sheet Packing/Pcs',
    'Sheet Qty/Pcs',
    'Pkt/Rim Qty',
    'Consumption Value',
]

DEMAND_GAP_MATERIAL_HEADERS = [
    'Material',
    'Purchase Sheet Size',
    'Raw Material SKU',
    'On Hand (Sheets)',
    'Total Demand (Sheets)',
    'Gap (Sheets)',
    'Status',
    'Print Jobs',
    'Cut & Pack Jobs',
    'Job Count',
]

DEMAND_GAP_JOB_HEADERS = [
    'JC Number',
    'Status',
    'Process Type',
    'Material',
    'Purchase Sheet Size',
    'Order Qty',
    'Purchase Sheets (Planning)',
    'Job Demand (Sheets)',
    'Print Sheets Consumed',
    'Packed Pcs',
    'Dispatched Pcs',
    'Remaining From Print',
    'Remaining From Pack',
    'Remaining From Dispatch',
    'Mapped',
]


def _parse_date(value):
    if value is None or value == '':
        return timezone.now().date()
    if hasattr(value, 'date'):
        return value.date()
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return timezone.now().date()


def _as_text(value):
    """Coerce Excel/CSV cell values to trimmed text (handles ints/floats)."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _as_int(value, default=0):
    if value is None or value == '':
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _get_raw_material_sku_from_row(row):
    sku = _as_text(row.get('Raw Material SKU') or row.get('Item ID'))
    if not sku:
        return None
    return RawMaterialSku.objects.filter(sku__iexact=sku).select_related('material').first()


def _get_item_by_id(item_id):
    return _get_raw_material_sku_from_row({'Raw Material SKU': item_id})


def _sku_row_values(item):
    return [
        item.sku,
        item.material.name,
        item.purchase_sheet_size,
        item.uom,
        item.sheet_packing_pcs,
    ]


def _read_rows(upload_file):
    name = (upload_file.name or '').lower()
    rows = []

    if name.endswith('.csv'):
        decoded = upload_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        for row in reader:
            rows.append({(k or '').strip(): v for k, v in row.items()})
        return rows

    if name.endswith('.xlsx'):
        if not EXCEL_AVAILABLE:
            raise ImportError('openpyxl is required for XLSX upload.')
        wb = openpyxl.load_workbook(upload_file, data_only=True)
        ws = wb.active
        header = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header = [str(c).strip() if c is not None else '' for c in row]
        if not header:
            return []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not any(values):
                continue
            row = {}
            for idx, key in enumerate(header):
                if key:
                    row[key] = values[idx] if idx < len(values) else None
            rows.append(row)
        return rows

    raise ValueError('Unsupported file type. Please upload CSV or XLSX.')


def _write_workbook(filename, headers, data_rows):
    if not EXCEL_AVAILABLE:
        raise ImportError('openpyxl is required for Excel export.')

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_num, row in enumerate(data_rows, 2):
        for col_num, value in enumerate(row, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def export_transactions(transaction_type, queryset, filename):
    rows = []
    for txn in queryset:
        rows.append([
            txn.month_str or '',
            txn.date.isoformat() if txn.date else '',
            txn.gin_jc or '',
            *(_sku_row_values(txn.raw_material_sku)),
            txn.sheet_qty_pcs,
            txn.pkt_rim_qty,
        ])
    return _write_workbook(filename, TRANSACTION_HEADERS, rows)


def export_demands(queryset, filename):
    rows = []
    for demand in queryset:
        rows.append([
            *(_sku_row_values(demand.raw_material_sku)),
            demand.month_str or '',
            demand.sheet_qty_pcs,
            demand.pkt_rim_qty,
        ])
    return _write_workbook(filename, DEMAND_HEADERS, rows)


def export_items(queryset, filename):
    rows = []
    for item in queryset:
        rows.append([
            item.sku,
            item.material.name,
            item.purchase_sheet_size,
            item.uom,
            item.sheet_packing_pcs,
            float(item.unit_cost),
            item.safety_stock,
            item.max_stock_level,
            item.lead_time_days,
            'Yes' if item.is_active else 'No',
        ])
    return _write_workbook(filename, RAW_MATERIAL_SKU_HEADERS, rows)


def export_item_requests(queryset, filename='item_requests.xlsx'):
    rows = []
    for req in queryset:
        rows.append([
            req.request_no or '',
            req.get_status_display(),
            req.request_type.name,
            req.department.name,
            req.item_title,
            req.machine_display,
            req.uom,
            float(req.required_quantity),
            req.get_local_import_display(),
            req.part_number,
            float(req.estimated_unit_price) if req.estimated_unit_price is not None else '',
            req.cost_centre,
            req.raised_by.username,
            req.request_date.isoformat() if req.request_date else '',
        ])
    return _write_workbook(filename, ITEM_REQUEST_HEADERS, rows)


def export_raw_material_sku_template(filename='raw_material_sku_template.xlsx'):
    example = [[
        'RM-TAF-2536',
        'Taffeta',
        '25x36',
        'Sheets',
        1,
        12.5,
        100,
        10000,
        7,
        'Yes',
    ]]
    return _write_workbook(filename, RAW_MATERIAL_SKU_HEADERS, example)


def _normalize_raw_material_import_row(row):
    return {
        'sku': _as_text(row.get('Raw Material SKU') or row.get('Item ID')),
        'material_name': _as_text(row.get('Material Name') or row.get('Item Type')),
        'purchase_sheet_size': _as_text(row.get('Purchase Sheet Size')),
        'uom': _as_text(row.get('UOM')),
        'sheet_packing_pcs': row.get('Sheet Packing/Pcs'),
        'unit_cost': row.get('Unit Cost'),
        'safety_stock': row.get('Safety Stock'),
        'max_stock_level': row.get('Max Stock Level'),
        'lead_time_days': row.get('Lead Time (Days)'),
        'is_active': row.get('Active'),
    }


def import_raw_material_skus(upload_file):
    rows = [_normalize_raw_material_import_row(row) for row in _read_rows(upload_file)]
    return _import_raw_material_skus(rows)


def export_item_wise_consumption(rows, filename='item_wise_monthly_consumption.xlsx'):
    data_rows = [
        [
            row['item_id'],
            row['item_type'],
            row.get('purchase_sheet_size', ''),
            row['uom'],
            row['sheet_packing_pcs'],
            row['month'],
            row['sheet_qty_pcs'],
            row['pkt_rim_qty'],
            row['consumption_value'],
        ]
        for row in rows
    ]
    return _write_workbook(filename, ITEM_WISE_CONSUMPTION_HEADERS, data_rows)


def export_month_wise_consumption(rows, filename='month_wise_item_consumption.xlsx'):
    data_rows = [
        [
            row['month'],
            row['item_id'],
            row['item_type'],
            row.get('purchase_sheet_size', ''),
            row['uom'],
            row['sheet_packing_pcs'],
            row['sheet_qty_pcs'],
            row['pkt_rim_qty'],
            row['consumption_value'],
        ]
        for row in rows
    ]
    return _write_workbook(filename, MONTH_WISE_CONSUMPTION_HEADERS, data_rows)


def export_kpi_dashboard(rows, filename='supply_chain_kpis.xlsx'):
    data_rows = [
        [
            row['item'].sku or '',
            row['item'].material.name,
            row['item'].purchase_sheet_size,
            row['abc_class'],
            row['fsn_class'],
            row['closing'],
            row['monthly_demand'],
            row['variable_daily_demand'],
            row['reorder_point'],
            row['item'].safety_stock,
            row['inventory_value'],
            row['inventory_turnover'],
            row['annual_consumption_value'],
            'Yes' if row['stockout'] else 'No',
            'Yes' if row['reorder'] else 'No',
            'Yes' if row['safety_stock_alert'] else 'No',
            'Yes' if row['overstock'] else 'No',
            'Yes' if row['dead_stock'] else 'No',
            'Yes' if row['slow_moving'] else 'No',
            row['inventory_accuracy'] if row.get('inventory_accuracy') is not None else '',
        ]
        for row in rows
    ]
    return _write_workbook(filename, KPI_HEADERS, data_rows)


def export_physical_counts(queryset, filename='physical_stock_counts.xlsx'):
    data_rows = []
    for count in queryset:
        variance = count.physical_sheet_qty - count.system_sheet_qty
        data_rows.append([
            count.count_date.isoformat() if count.count_date else '',
            count.raw_material_sku.sku,
            count.raw_material_sku.material.name,
            count.raw_material_sku.purchase_sheet_size,
            count.physical_sheet_qty,
            count.system_sheet_qty,
            variance,
            float(count.accuracy_percent),
            count.notes,
        ])
    return _write_workbook(filename, PHYSICAL_COUNT_HEADERS, data_rows)


def import_transactions(transaction_type, upload_file):
    rows = _read_rows(upload_file)
    created = 0
    skipped_missing_sku = 0
    skipped_duplicates = 0
    seen_in_file = set()

    for row in rows:
        item = _get_raw_material_sku_from_row(row)
        if not item:
            skipped_missing_sku += 1
            continue

        txn_date = _parse_date(row.get('Date'))
        gin_jc = _as_text(row.get('GIN / JC') or row.get('GIN/JC')) or None
        sheet_qty_pcs = _as_int(row.get('Sheet Qty/Pcs'))
        pkt_rim_qty = _as_int(row.get('Pkt/Rim Qty'))
        month_str = _as_text(row.get('Month')) or None

        dedupe_key = (
            transaction_type,
            item.pk,
            txn_date.isoformat(),
            (gin_jc or '').strip().lower(),
            sheet_qty_pcs,
            pkt_rim_qty,
        )
        if dedupe_key in seen_in_file:
            skipped_duplicates += 1
            continue
        seen_in_file.add(dedupe_key)

        if gin_jc:
            duplicate_qs = StockTransaction.objects.filter(
                raw_material_sku=item,
                transaction_type=transaction_type,
                source='MANUAL',
                date=txn_date,
                sheet_qty_pcs=sheet_qty_pcs,
                pkt_rim_qty=pkt_rim_qty,
                gin_jc__iexact=gin_jc,
            )
        else:
            duplicate_qs = StockTransaction.objects.filter(
                raw_material_sku=item,
                transaction_type=transaction_type,
                source='MANUAL',
                date=txn_date,
                sheet_qty_pcs=sheet_qty_pcs,
                pkt_rim_qty=pkt_rim_qty,
            ).filter(Q(gin_jc__isnull=True) | Q(gin_jc=''))

        if duplicate_qs.exists():
            skipped_duplicates += 1
            continue

        StockTransaction.objects.create(
            raw_material_sku=item,
            transaction_type=transaction_type,
            source='MANUAL',
            month_str=month_str,
            date=txn_date,
            gin_jc=gin_jc,
            sheet_qty_pcs=sheet_qty_pcs,
            pkt_rim_qty=pkt_rim_qty,
        )
        created += 1

    return {
        'created': created,
        'skipped_missing_sku': skipped_missing_sku,
        'skipped_duplicates': skipped_duplicates,
        'skipped': skipped_missing_sku + skipped_duplicates,
    }


def import_demands(upload_file):
    rows = _read_rows(upload_file)
    created = 0
    skipped_missing_sku = 0
    skipped_duplicates = 0
    seen_in_file = set()

    for row in rows:
        item = _get_raw_material_sku_from_row(row)
        if not item:
            skipped_missing_sku += 1
            continue

        month_str = _as_text(row.get('Month')) or None
        sheet_qty_pcs = _as_int(row.get('Sheet Qty/Pcs'))
        pkt_rim_qty = _as_int(row.get('Pkt/Rim Qty'))
        dedupe_key = (item.pk, (month_str or '').lower(), sheet_qty_pcs, pkt_rim_qty)
        if dedupe_key in seen_in_file:
            skipped_duplicates += 1
            continue
        seen_in_file.add(dedupe_key)

        if month_str:
            duplicate_qs = StockDemand.objects.filter(
                raw_material_sku=item,
                month_str__iexact=month_str,
                sheet_qty_pcs=sheet_qty_pcs,
                pkt_rim_qty=pkt_rim_qty,
            )
        else:
            duplicate_qs = StockDemand.objects.filter(
                raw_material_sku=item,
                sheet_qty_pcs=sheet_qty_pcs,
                pkt_rim_qty=pkt_rim_qty,
            ).filter(Q(month_str__isnull=True) | Q(month_str=''))

        if duplicate_qs.exists():
            skipped_duplicates += 1
            continue

        StockDemand.objects.create(
            raw_material_sku=item,
            month_str=month_str,
            sheet_qty_pcs=sheet_qty_pcs,
            pkt_rim_qty=pkt_rim_qty,
        )
        created += 1

    return {
        'created': created,
        'skipped_missing_sku': skipped_missing_sku,
        'skipped_duplicates': skipped_duplicates,
        'skipped': skipped_missing_sku + skipped_duplicates,
    }


def export_demand_gap_materials(material_rows, filename='demand_gap_materials.xlsx'):
    data_rows = []
    for row in material_rows:
        sku = ''
        if row.get('raw_material_sku'):
            sku = row['raw_material_sku'].sku
        data_rows.append([
            row.get('material_name') or '',
            row.get('purchase_sheet_size') or '',
            sku,
            row.get('on_hand') if row.get('on_hand') is not None else '',
            row.get('total_demand') or 0,
            row.get('gap') if row.get('gap') is not None else '',
            row.get('gap_status') or '',
            row.get('print_job_count') or 0,
            row.get('cut_pack_job_count') or 0,
            row.get('job_count') or 0,
        ])
    return _write_workbook(filename, DEMAND_GAP_MATERIAL_HEADERS, data_rows)


def export_demand_gap_jobs(job_rows, filename='demand_gap_jobs.xlsx'):
    data_rows = []
    for row in job_rows:
        data_rows.append([
            row.get('jc_number') or '',
            row.get('status_label') or row.get('status') or '',
            row.get('process_label') or row.get('process_type') or '',
            row.get('material_name') or '',
            row.get('purchase_sheet_size') or '',
            row.get('order_qty') or 0,
            row.get('purchase_sheets_planning') or '',
            row.get('job_demand_sheets') or 0,
            row.get('consumed_print_sheets') or 0,
            row.get('packed_pcs') or 0,
            row.get('dispatched_pcs') or 0,
            row.get('remaining_from_print') if row.get('remaining_from_print') is not None else '',
            row.get('remaining_from_pack') if row.get('remaining_from_pack') is not None else '',
            row.get('remaining_from_dispatch') if row.get('remaining_from_dispatch') is not None else '',
            'Yes' if row.get('is_mapped') else 'No',
        ])
    return _write_workbook(filename, DEMAND_GAP_JOB_HEADERS, data_rows)
